import os
import tkinter
from win32com import client
from tkinter import *
from openpyxl import load_workbook
from tkinter import ttk, filedialog

altura = 750
largura = 1200

root = Tk()
root.title('Formulário de preenchimento da planilha auxiliar')
root.geometry(str(largura)+'x'+str(altura))

fr_idiomas = LabelFrame(root, borderwidth=1, relief='solid', text='  Selecione um idioma/Select a language:  ')
fr_idiomas.place(x=2, y=0, width=350, height=80)

fr_info = LabelFrame(root, borderwidth=1, relief='solid', text='  Informações/Informations:  ')
fr_info.place(x=355, y=0, width=843, height=80)

info_texto = Label(fr_info, text='Escolha um idioma para continuar/Select a language to continue.', font= '20')
info_texto.place(x=180, y=20)

# Imprime apresentação em português
fr_apresent_pt = LabelFrame(root, borderwidth=1, relief='solid', text='  Português:  ')
fr_apresent_pt.place(x=150, y=120, width=440, height=450)

Label(fr_apresent_pt, text='Bem-vindo ao formulário de preenchimento de \ndados para criação de planilha auxiliar!', font="-weight bold -size 13").place(x=10, y=10)

Label(fr_apresent_pt, text='Aviso:', font= '20').place(x=5, y=80)
Label(fr_apresent_pt, text='Esta é uma versão preliminar, sendo suscetível a "bugs".', font= '20').place(x=5, y=110)
Label(fr_apresent_pt, text='Acesse a aba "Executar Programa" para mais informações.', font= '20').place(x=5, y=140)

Label(fr_apresent_pt, text='Instruções básicas de uso:', font= '20').place(x=5, y=190)
Label(fr_apresent_pt, text='1) Escolha o idioma desejado no quadro acima;', font= '20').place(x=5, y=220)
Label(fr_apresent_pt, text='2) Percorra as abas preenchendo todos os dados;', font= '20').place(x=5, y=250)
Label(fr_apresent_pt, text='3) Escolha uma pasta na aba "Executar Programa";', font= '20').place(x=5, y=280)
Label(fr_apresent_pt, text='4) Clique no botão "Imprimir Planilha";', font= '20').place(x=5, y=310)
Label(fr_apresent_pt, text='3) Acesse a pasta de salvamento escolhida;', font= '20').place(x=5, y=340)
Label(fr_apresent_pt, text='3) Abra a planilha "planilha_resposta_mat_12345678.xlsx".', font= '20').place(x=5, y=370)

# Imprime apresentação em inglês
fr_apresent_en = LabelFrame(root, borderwidth=1, relief='solid', text='  English:  ')
fr_apresent_en.place(x=600, y=120, width=440, height=450)

Label(fr_apresent_en, text='Welcome to the data fill form for auxiliary \ndatasheet automatic creation!', font="-weight bold -size 13").place(x=30, y=10)

Label(fr_apresent_en, text='Warning:', font= '20').place(x=5, y=80)
Label(fr_apresent_en, text='This is a preliminary version, being susceptible to "bugs".', font= '20').place(x=5, y=110)
Label(fr_apresent_en, text='Go to the "Run Script" tab for more information.', font= '20').place(x=5, y=140)

Label(fr_apresent_en, text='Basic instructions for use:', font= '20').place(x=5, y=190)
Label(fr_apresent_en, text='1) Choose a language in the frame above;', font= '20').place(x=5, y=220)
Label(fr_apresent_en, text='2) Go through the tabs and fill in all fields;', font= '20').place(x=5, y=250)
Label(fr_apresent_en, text='3) Choose a folder in the "Run Script" tab;', font= '20').place(x=5, y=280)
Label(fr_apresent_en, text='4) Click the "Print Datasheet" button;', font= '20').place(x=5, y=310)
Label(fr_apresent_en, text='3) Go to the saving folder chosen before;', font= '20').place(x=5, y=340)
Label(fr_apresent_en, text='3) Open the "output_sheet_mat_12345678.xlsx" datasheet.', font= '20').place(x=5, y=370)

# Inicializando flags, dicionários e variáveis das 11 abas
flag51 = flag52 = flag53 = flag54 = True
flag61 = flag62 = flag63 = flag64 = flag65 = flag66 = flag67 = flag68 = flag69 = flag610 = flag611 = flag612 = True
flag71 = flag72 = flag73 = flag74 = flag75 = flag76 = flag77 = flag78 = flag79 = flag710 = flag711 = flag712 = True
flag81 = flag82 = flag83 = flag84 = True
flag94 = flag95 = flag96 = True
tagX2_61 = tagX2_62 = tagX2_63 = tagX2_64 = tagX2_71 = tagX2_72 = tagX2_73 = tagX2_74 = ''
dados_aba1 = dados_aba2 = dados_aba3 = dados_aba4 = dados_aba5 = dados_aba6 = dados_aba7 = dados_aba8 = dados_aba9 = dados_aba10 = {}
a = b = c = d = e = f = g = h = i = j = k = k1 = k2 = l = tagX_aba10 = releAux = conduite = qtd_conduite = luva = qtd_luva = 'Falha'
conMacho = qtd_conMacho = abracadeira = qtd_abracadeira = parafuso = qtd_parafuso = fixTermom = qtd_fixTermom = 'Falha'
cj_fix_le = cj_fix_ld = cj_tampa_resist = cabo_lig = term_olhal = cond_G1 = con_MGR_G1 = abrac_cond_G1 = paraf_abrac_G1 = 'Falha'
cj_fix_termom_aba3 = conduite_aba3 = qtd_conduite_aba3 = luva_aba3 = qtd_luva_aba3 = conMacho_aba3 = qtd_conMacho_aba3 = 'Falha'
adap_isolante_aba3 = qtd_adap_isolante_aba3 = abracadeira_aba3 = qtd_abracadeira_aba3 = parafuso_aba3 = qtd_parafuso_aba3 = sup_inclinado_aba3 = rosca_fix_res_aba3 = 'Falha'
cj_fix_termom_aba4 = conduite_aba4 = qtd_conduite_aba4 = luva_aba4 = qtd_luva_aba4 = conMacho_aba4 = qtd_conMacho_aba4 = 'Falha'
adap_isolante_aba4 = qtd_adap_isolante_aba4 = rosca_fix_res_aba4 = sup_inclinado_aba4 = abracadeira_aba4 = qtd_abracadeira_aba4 = parafuso_aba4 = qtd_parafuso_aba4 = 'Falha'
pasta_escolhida = ''
texto_padrao11 = texto_padrao12 = texto_padrao13 = texto_padrao14 = True
texto_padrao_lt21 = True
texto_padrao_lt31 = True
texto_padrao_lt41 = True
texto_padrao_lt51 = texto_padrao_lt52 = texto_padrao_lt53 = texto_padrao_lt54 = texto_padrao51 = texto_padrao52 = texto_padrao53 = texto_padrao54 = True
texto_padrao_lt61 = texto_padrao_lt62 = texto_padrao_lt63 = texto_padrao_lt64 = texto_padrao61 = texto_padrao62 = texto_padrao63 = texto_padrao64 = True
texto_padrao_lt71 = texto_padrao_lt72 = texto_padrao_lt73 = texto_padrao_lt74 = texto_padrao71 = texto_padrao72 = texto_padrao73 = texto_padrao74 = True
texto_padrao_lt91 = True
texto_padrao_lt101 = True
erro = False

# Inicializando o botão de imprimir
btn_imprimir = Button(root, text='')

def compila_dados():
   global bd, pa, aba_trilhos, aba_cx_peq, aba_cx_med, aba_cx_gde, aba_cx_emb, aba_compGerais, aba_sensorTemp, aba_termometros, aba_resist_aquec, aba_pa
   global m, n, s, a1, a2, a3, a4, a5, a6, a7, a8, a9, a10, a11, flag_botao, incremento, pos_inic, celulas
   global dados_aba1, dados_aba2, dados_aba3, dados_aba4, dados_aba5, dados_aba6, dados_aba7, dados_aba8, dados_aba9, dados_aba10, dados_aba11
   global a, b, c, d, e, f, g, h, i, j, k, k1, k2, l, tagX_aba10, releAux, cj_fix_le, cj_fix_ld, cj_tampa_resist, cabo_lig, term_olhal, con_MGR_G1, abrac_cond_G1, paraf_abrac_G1, cond_G1
   global conduite, qtd_conduite, luva, qtd_luva, conMacho, qtd_conMacho, abracadeira, qtd_abracadeira, parafuso, qtd_parafuso, fixTermom, qtd_fixTermom
   global cj_fix_termom_aba3, conduite_aba3, qtd_conduite_aba3, luva_aba3, qtd_luva_aba3, conMacho_aba3, qtd_conMacho_aba3, abracadeira_aba3, qtd_abracadeira_aba3, parafuso_aba3, qtd_parafuso_aba3, sup_inclinado_aba3, rosca_fix_res_aba3, adap_isolante_aba3, qtd_adap_isolante_aba3
   global adap_isolante_aba4, qtd_adap_isolante_aba4, rosca_fix_res_aba4, sup_inclinado_aba4, abracadeira_aba4, qtd_abracadeira_aba4, parafuso_aba4, qtd_parafuso_aba4, cj_fix_termom_aba4, conduite_aba4, qtd_conduite_aba4, luva_aba4, qtd_luva_aba4, conMacho_aba4, qtd_conMacho_aba4
   global fpath, excel, sheets, work_sheets
   global erro, res, top

   # Carregando as planilhas de banco de dados (bd) e folha resposta (pa)
   bd = load_workbook('banco_de_dados.xlsx')
   pa = load_workbook('planilha_resposta_modelo.xlsx')

   # Selecionando as abas da planilha de banco de dados
   aba_trilhos = bd['Trilhos']
   aba_cx_peq = bd['Caixa Pequena']
   aba_cx_med = bd['Caixa Média']
   aba_cx_gde = bd['Caixa Grande']
   aba_cx_emb = bd['Caixa Embutida']
   aba_compGerais = bd['componentes_gerais']
   aba_sensorTemp = bd['Sensor de temperatura']
   aba_termometros = bd['Termômetros']
   aba_resist_aquec = bd['Resistores aquecimento']

   # Selecionando a aba 'Folha_resposta' da planilha auxiliar
   aba_pa = pa['Folha_resposta']

   # Compila os dados da aba 1 (Informações de projeto)
   dados_aba1 = {'opcoes11': opcoes11.get(), 'opcoes12': opcoes12.get(), 'opcoes13': opcoes13.get(), 'opcoes14': opcoes14.get(), 'texto11': texto11.get(), 'texto12': texto12.get(), 'texto13': texto13.get(), 'texto14': texto14.get()}


   # Compila os dados da aba 2 (Carcaça/Estator)
   dados_aba2 = {'opcoes21': opcoes21.get(), 'opcoes22': opcoes22.get(), 'opcoes23': opcoes23.get(), 'opcoes24': opcoes24.get(), 'varcb1': varcb1.get(), 'varcb2': varcb2.get(), 'lt21': lt21.get(), 'inc21': inc21.get()} 
   

   # Compila os dados da aba 3 (mancal LA)
   dados_aba3 = {'opcoes31': opcoes31.get(), 'opcoes32': opcoes32.get(), 'opcoes33': opcoes33.get(), 'opcoes34': opcoes34.get(), 'opcoes35': opcoes35.get(), 'opcoes36': opcoes36.get(), 'opcoes310': opcoes310.get(), 'opcoes311': opcoes311.get(), 'opcoes312': opcoes312.get(), 'opcoes313': opcoes313.get(), 'opcoes314': opcoes314.get(), 'lt31': lt31.get(), 'inc31': inc31.get()}
   

   # Compila os dados da aba 4 (Mancal LNA)
   dados_aba4 = {'opcoes41': opcoes41.get(), 'opcoes42': opcoes42.get(), 'opcoes43': opcoes43.get(), 'opcoes44': opcoes44.get(), 'opcoes45': opcoes45.get(), 'opcoes46': opcoes46.get(), 'opcoes410': opcoes410.get(), 'opcoes411': opcoes411.get(), 'opcoes412': opcoes412.get(), 'opcoes413': opcoes413.get(), 'opcoes414': opcoes414.get(), 'lt41': lt41.get(), 'inc41': inc41.get()}


   # Compila os dados da aba 5 (Cx. Acess. Pequena)
   if flag51 is False and flag52 is False and flag53 is False and flag54 is False:
      dados_aba5 = {'opcoes51': opcoes51.get(), 'opcoes52': opcoes52.get(), 'opcoes53': opcoes53.get()}

   elif flag51 is True and flag52 is False and flag53 is False and flag54 is False:
      dados_aba5 = {'opcoes51': opcoes51.get(), 'opcoes52': opcoes52.get(), 'opcoes53': opcoes53.get(), 'opcoes54': opcoes54.get(), 'opcoes55': opcoes55.get(), 'opcoes57': opcoes57.get(), 'opcoes58': opcoes58.get(), 'opcoes59': opcoes59.get(), 'opcoes510': opcoes510.get(), 'texto51': texto51.get(), 'lt51': lt51.get(), 'inc51': inc51.get()}
   
   elif flag51 is True and flag52 is True and flag53 is False and flag54 is False:
      dados_aba5 = {'opcoes51': opcoes51.get(), 'opcoes52': opcoes52.get(), 'opcoes53': opcoes53.get(), 'opcoes54': opcoes54.get(), 'opcoes55': opcoes55.get(), 'opcoes57': opcoes57.get(), 'opcoes58': opcoes58.get(), 'opcoes59': opcoes59.get(), 'opcoes510': opcoes510.get(), 'opcoes511': opcoes511.get(), 'opcoes512': opcoes512.get(), 'opcoes514': opcoes514.get(), 'opcoes515': opcoes515.get(), 'opcoes516': opcoes516.get(), 'opcoes517': opcoes517.get(), 'texto51': texto51.get(), 'texto52': texto52.get(), 'lt51': lt51.get(), 'inc51': inc51.get(), 'lt52': lt52.get(), 'inc52': inc52.get()}
   
   elif flag51 is True and flag52 is True and flag53 is True and flag54 is False:
      dados_aba5 = {'opcoes51': opcoes51.get(), 'opcoes52': opcoes52.get(), 'opcoes53': opcoes53.get(), 'opcoes54': opcoes54.get(), 'opcoes55': opcoes55.get(), 'opcoes57': opcoes57.get(), 'opcoes58': opcoes58.get(), 'opcoes59': opcoes59.get(), 'opcoes510': opcoes510.get(), 'opcoes511': opcoes511.get(), 'opcoes512': opcoes512.get(), 'opcoes514': opcoes514.get(), 'opcoes515': opcoes515.get(), 'opcoes516': opcoes516.get(), 'opcoes517': opcoes517.get(), 'opcoes518': opcoes518.get(), 'opcoes519': opcoes519.get(), 'opcoes521': opcoes521.get(), 'opcoes522': opcoes522.get(), 'opcoes523': opcoes523.get(), 'opcoes524': opcoes524.get(), 'texto51': texto51.get(), 'texto52': texto52.get(), 'texto53': texto53.get(), 'lt51': lt51.get(), 'inc51': inc51.get(), 'lt52': lt52.get(), 'inc52': inc52.get(), 'lt53': lt53.get(), 'inc53': inc53.get()}

   elif flag51 is True and flag52 is True and flag53 is True and flag54 is True:
      dados_aba5 = {'opcoes51': opcoes51.get(), 'opcoes52': opcoes52.get(), 'opcoes53': opcoes53.get(), 'opcoes54': opcoes54.get(), 'opcoes55': opcoes55.get(), 'opcoes57': opcoes57.get(), 'opcoes58': opcoes58.get(), 'opcoes59': opcoes59.get(), 'opcoes510': opcoes510.get(), 'opcoes511': opcoes511.get(), 'opcoes512': opcoes512.get(), 'opcoes514': opcoes514.get(), 'opcoes515': opcoes515.get(), 'opcoes516': opcoes516.get(), 'opcoes517': opcoes517.get(), 'opcoes518': opcoes518.get(), 'opcoes519': opcoes519.get(), 'opcoes521': opcoes521.get(), 'opcoes522': opcoes522.get(), 'opcoes523': opcoes523.get(), 'opcoes524': opcoes524.get(), 'opcoes525': opcoes525.get(), 'opcoes526': opcoes526.get(), 'opcoes528': opcoes528.get(), 'opcoes529': opcoes529.get(), 'opcoes530': opcoes530.get(), 'opcoes531': opcoes531.get(), 'texto51': texto51.get(), 'texto52': texto52.get(), 'texto53': texto53.get(), 'texto54': texto54.get(), 'lt51': lt51.get(), 'inc51': inc51.get(), 'lt52': lt52.get(), 'inc52': inc52.get(), 'lt53': lt53.get(), 'inc53': inc53.get(), 'lt54': lt54.get(), 'inc54': inc54.get()}


   # Compila os dados da aba 6 (Cx. Acess. Média)
   if flag61 is False and flag62 is False and flag63 is False and flag64 is False:
      dados_aba6 = {'opcoes61': opcoes61.get(), 'opcoes62': opcoes62.get(), 'opcoes63': opcoes63.get()}

   elif flag61 is True and flag62 is False and flag63 is False and flag64 is False:
      dados_aba6 = {'opcoes61': opcoes61.get(), 'opcoes62': opcoes62.get(), 'opcoes63': opcoes63.get(), 'opcoes64': opcoes64.get(), 'opcoes65': opcoes65.get(), 'opcoes66': opcoes66.get(), 'opcoes67': opcoes67.get(), 'opcoes68': opcoes68.get(), 'opcoes69': opcoes69.get(), 'opcoes610': opcoes610.get(), 'opcoes611': opcoes611.get(), 'opcoes612': opcoes612.get(), 'opcoes613': opcoes613.get(), 'opcoes614': opcoes614.get(), 'texto61': texto61.get(), 'lt61': lt61.get(), 'inc61': inc61.get()}

   elif flag61 is True and flag62 is True and flag63 is False and flag64 is False:
      dados_aba6 = {'opcoes61': opcoes61.get(), 'opcoes62': opcoes62.get(), 'opcoes63': opcoes63.get(), 'opcoes64': opcoes64.get(), 'opcoes65': opcoes65.get(), 'opcoes66': opcoes66.get(), 'opcoes67': opcoes67.get(), 'opcoes68': opcoes68.get(), 'opcoes69': opcoes69.get(), 'opcoes610': opcoes610.get(), 'opcoes611': opcoes611.get(), 'opcoes612': opcoes612.get(), 'opcoes613': opcoes613.get(), 'opcoes614': opcoes614.get(), 'opcoes615': opcoes615.get(), 'opcoes616': opcoes616.get(), 'opcoes617': opcoes617.get(), 'opcoes618': opcoes618.get(), 'opcoes619': opcoes619.get(), 'opcoes620': opcoes620.get(), 'opcoes621': opcoes621.get(), 'opcoes622': opcoes622.get(), 'opcoes623': opcoes623.get(), 'opcoes624': opcoes624.get(), 'opcoes625': opcoes625.get(), 'texto61': texto61.get(), 'texto62': texto62.get(), 'lt61': lt61.get(), 'inc61': inc61.get(), 'lt62': lt62.get(), 'inc62': inc62.get()}

   elif flag61 is True and flag62 is True and flag63 is True and flag64 is False:
      dados_aba6 = {'opcoes61': opcoes61.get(), 'opcoes62': opcoes62.get(), 'opcoes63': opcoes63.get(), 'opcoes64': opcoes64.get(), 'opcoes65': opcoes65.get(), 'opcoes66': opcoes66.get(), 'opcoes67': opcoes67.get(), 'opcoes68': opcoes68.get(), 'opcoes69': opcoes69.get(), 'opcoes610': opcoes610.get(), 'opcoes611': opcoes611.get(), 'opcoes612': opcoes612.get(), 'opcoes613': opcoes613.get(), 'opcoes614': opcoes614.get(), 'opcoes615': opcoes615.get(), 'opcoes616': opcoes616.get(), 'opcoes617': opcoes617.get(), 'opcoes618': opcoes618.get(), 'opcoes619': opcoes619.get(), 'opcoes620': opcoes620.get(), 'opcoes621': opcoes621.get(), 'opcoes622': opcoes622.get(), 'opcoes623': opcoes623.get(), 'opcoes624': opcoes624.get(), 'opcoes625': opcoes625.get(), 'opcoes626': opcoes626.get(), 'opcoes627': opcoes627.get(), 'opcoes628': opcoes628.get(), 'opcoes629': opcoes629.get(), 'opcoes630': opcoes630.get(), 'opcoes631': opcoes631.get(), 'opcoes632': opcoes632.get(), 'opcoes633': opcoes633.get(), 'opcoes634': opcoes634.get(), 'opcoes635': opcoes635.get(), 'opcoes636': opcoes636.get(), 'texto61': texto61.get(), 'texto62': texto62.get(), 'texto63': texto63.get(), 'lt61': lt61.get(), 'inc61': inc61.get(), 'lt62': lt62.get(), 'inc62': inc62.get(), 'lt63': lt63.get(), 'inc63': inc63.get()}

   elif flag61 is True and flag62 is True and flag63 is True and flag64 is True:
      dados_aba6 = {'opcoes61': opcoes61.get(), 'opcoes62': opcoes62.get(), 'opcoes63': opcoes63.get(), 'opcoes64': opcoes64.get(), 'opcoes65': opcoes65.get(), 'opcoes66': opcoes66.get(), 'opcoes67': opcoes67.get(), 'opcoes68': opcoes68.get(), 'opcoes69': opcoes69.get(), 'opcoes610': opcoes610.get(), 'opcoes611': opcoes611.get(), 'opcoes612': opcoes612.get(), 'opcoes613': opcoes613.get(), 'opcoes614': opcoes614.get(), 'opcoes615': opcoes615.get(), 'opcoes616': opcoes616.get(), 'opcoes617': opcoes617.get(), 'opcoes618': opcoes618.get(), 'opcoes619': opcoes619.get(), 'opcoes620': opcoes620.get(), 'opcoes621': opcoes621.get(), 'opcoes622': opcoes622.get(), 'opcoes623': opcoes623.get(), 'opcoes624': opcoes624.get(), 'opcoes625': opcoes625.get(), 'opcoes626': opcoes626.get(), 'opcoes627': opcoes627.get(), 'opcoes628': opcoes628.get(), 'opcoes629': opcoes629.get(), 'opcoes630': opcoes630.get(), 'opcoes631': opcoes631.get(), 'opcoes632': opcoes632.get(), 'opcoes633': opcoes633.get(), 'opcoes634': opcoes634.get(), 'opcoes635': opcoes635.get(), 'opcoes636': opcoes636.get(), 'opcoes637': opcoes637.get(), 'opcoes638': opcoes638.get(), 'opcoes639': opcoes639.get(), 'opcoes640': opcoes640.get(), 'opcoes641': opcoes641.get(), 'opcoes642': opcoes642.get(), 'opcoes643': opcoes643.get(), 'opcoes644': opcoes644.get(), 'opcoes645': opcoes645.get(), 'opcoes646': opcoes646.get(), 'opcoes647': opcoes647.get(), 'texto61': texto61.get(), 'texto62': texto62.get(), 'texto63': texto63.get(), 'texto64': texto64.get(), 'lt61': lt61.get(), 'inc61': inc61.get(), 'lt62': lt62.get(), 'inc62': inc62.get(), 'lt63': lt63.get(), 'inc63': inc63.get(), 'lt64': lt64.get(), 'inc64': inc64.get()}


   # Compila os dados da aba 7 (Cx. Acess. Grande)
   if flag71 is False and flag72 is False and flag73 is False and flag74 is False:
      dados_aba7 = {'opcoes71': opcoes71.get(), 'opcoes72': opcoes72.get(), 'opcoes73': opcoes73.get()}

   elif flag71 is True and flag72 is False and flag73 is False and flag74 is False:
      dados_aba7 = {'opcoes71': opcoes71.get(), 'opcoes72': opcoes72.get(), 'opcoes73': opcoes73.get(), 'opcoes74': opcoes74.get(), 'opcoes75': opcoes75.get(), 'opcoes76': opcoes76.get(), 'opcoes77': opcoes77.get(), 'opcoes78': opcoes78.get(), 'opcoes79': opcoes79.get(), 'opcoes710': opcoes710.get(), 'opcoes711': opcoes711.get(), 'opcoes712': opcoes712.get(), 'opcoes714': opcoes714.get(), 'texto71': texto71.get(), 'lt71': lt71.get(), 'inc71': inc71.get()}

   elif flag71 is True and flag72 is True and flag73 is False and flag74 is False:
      dados_aba7 = {'opcoes71': opcoes71.get(), 'opcoes72': opcoes72.get(), 'opcoes73': opcoes73.get(), 'opcoes74': opcoes74.get(), 'opcoes75': opcoes75.get(), 'opcoes76': opcoes76.get(), 'opcoes77': opcoes77.get(), 'opcoes78': opcoes78.get(), 'opcoes79': opcoes79.get(), 'opcoes710': opcoes710.get(), 'opcoes711': opcoes711.get(), 'opcoes712': opcoes712.get(), 'opcoes714': opcoes714.get(), 'opcoes715': opcoes715.get(), 'opcoes716': opcoes716.get(), 'opcoes717': opcoes717.get(), 'opcoes718': opcoes718.get(), 'opcoes719': opcoes719.get(), 'opcoes720': opcoes720.get(), 'opcoes721': opcoes721.get(), 'opcoes722': opcoes722.get(), 'opcoes723': opcoes723.get(), 'opcoes725': opcoes725.get(), 'texto71': texto71.get(), 'texto72': texto72.get(), 'lt71': lt71.get(), 'inc71': inc71.get(), 'lt72': lt72.get(), 'inc72': inc72.get()}

   elif flag71 is True and flag72 is True and flag73 is True and flag74 is False:
      dados_aba7 = {'opcoes71': opcoes71.get(), 'opcoes72': opcoes72.get(), 'opcoes73': opcoes73.get(), 'opcoes74': opcoes74.get(), 'opcoes75': opcoes75.get(), 'opcoes76': opcoes76.get(), 'opcoes77': opcoes77.get(), 'opcoes78': opcoes78.get(), 'opcoes79': opcoes79.get(), 'opcoes710': opcoes710.get(), 'opcoes711': opcoes711.get(), 'opcoes712': opcoes712.get(), 'opcoes714': opcoes714.get(), 'opcoes715': opcoes715.get(), 'opcoes716': opcoes716.get(), 'opcoes717': opcoes717.get(), 'opcoes718': opcoes718.get(), 'opcoes719': opcoes719.get(), 'opcoes720': opcoes720.get(), 'opcoes721': opcoes721.get(), 'opcoes722': opcoes722.get(), 'opcoes723': opcoes723.get(), 'opcoes725': opcoes725.get(), 'opcoes726': opcoes726.get(), 'opcoes727': opcoes727.get(), 'opcoes728': opcoes728.get(), 'opcoes729': opcoes729.get(), 'opcoes730': opcoes730.get(), 'opcoes731': opcoes731.get(), 'opcoes732': opcoes732.get(), 'opcoes733': opcoes733.get(), 'opcoes734': opcoes734.get(), 'opcoes736': opcoes736.get(), 'texto71': texto71.get(), 'texto72': texto72.get(), 'texto73': texto73.get(), 'lt71': lt71.get(), 'inc71': inc71.get(), 'lt72': lt72.get(), 'inc72': inc72.get(), 'lt73': lt73.get(), 'inc73': inc73.get()}

   elif flag71 is True and flag72 is True and flag73 is True and flag74 is True:
      dados_aba7 = {'opcoes71': opcoes71.get(), 'opcoes72': opcoes72.get(), 'opcoes73': opcoes73.get(), 'opcoes74': opcoes74.get(), 'opcoes75': opcoes75.get(), 'opcoes76': opcoes76.get(), 'opcoes77': opcoes77.get(), 'opcoes78': opcoes78.get(), 'opcoes79': opcoes79.get(), 'opcoes710': opcoes710.get(), 'opcoes711': opcoes711.get(), 'opcoes712': opcoes712.get(), 'opcoes714': opcoes714.get(), 'opcoes715': opcoes715.get(), 'opcoes716': opcoes716.get(), 'opcoes717': opcoes717.get(), 'opcoes718': opcoes718.get(), 'opcoes719': opcoes719.get(), 'opcoes720': opcoes720.get(), 'opcoes721': opcoes721.get(), 'opcoes722': opcoes722.get(), 'opcoes723': opcoes723.get(), 'opcoes725': opcoes725.get(), 'opcoes726': opcoes726.get(), 'opcoes727': opcoes727.get(), 'opcoes728': opcoes728.get(), 'opcoes729': opcoes729.get(), 'opcoes730': opcoes730.get(), 'opcoes731': opcoes731.get(), 'opcoes732': opcoes732.get(), 'opcoes733': opcoes733.get(), 'opcoes734': opcoes734.get(), 'opcoes736': opcoes736.get(), 'opcoes737': opcoes737.get(), 'opcoes738': opcoes738.get(), 'opcoes739': opcoes739.get(), 'opcoes740': opcoes740.get(), 'opcoes741': opcoes741.get(), 'opcoes742': opcoes742.get(), 'opcoes743': opcoes743.get(), 'opcoes744': opcoes744.get(), 'opcoes745': opcoes745.get(), 'opcoes747': opcoes747.get(), 'texto71': texto71.get(), 'texto72': texto72.get(), 'texto73': texto73.get(), 'texto74': texto74.get(), 'lt71': lt71.get(), 'inc71': inc71.get(), 'lt72': lt72.get(), 'inc72': inc72.get(), 'lt73': lt73.get(), 'inc73': inc73.get(), 'lt74': lt74.get(), 'inc74': inc74.get()}


   # Compila os dados da aba 8 (Cx. Acess. Embutida)
   if flag81 is False and flag82 is False:
      dados_aba8 = {'opcoes81': opcoes81.get(), 'opcoes82': opcoes82.get(), 'opcoes83': opcoes83.get()}
   
   elif flag81 is True and flag82 is False:
      if flag83 is True and flag84 is False:
            dados_aba8 = {'opcoes81': opcoes81.get(), 'opcoes82': opcoes82.get(), 'opcoes83': opcoes83.get(), 'opcoes84': opcoes84.get(), 'opcoes85': opcoes85.get(), 'opcoes86': opcoes86.get(), 'opcoes87': opcoes87.get(), 'opcoes88': opcoes88.get(), 'opcoes810': opcoes810.get(), 'texto81': texto81.get()}
      elif flag83 is True and flag84 is True:
            dados_aba8 = {'opcoes81': opcoes81.get(), 'opcoes82': opcoes82.get(), 'opcoes83': opcoes83.get(), 'opcoes84': opcoes84.get(), 'opcoes85': opcoes85.get(), 'opcoes86': opcoes86.get(), 'opcoes87': opcoes87.get(), 'opcoes88': opcoes88.get(), 'opcoes89': opcoes89.get(), 'opcoes810': opcoes810.get(), 'texto81': texto81.get()}

   elif flag81 is True and flag82 is True:
      if flag83 is False and flag84 is False:
            dados_aba8 = {'opcoes81': opcoes81.get(), 'opcoes82': opcoes82.get(), 'opcoes83': opcoes83.get(), 'opcoes84': opcoes84.get(), 'opcoes85': opcoes85.get(), 'opcoes86': opcoes86.get(), 'opcoes87': opcoes87.get(), 'opcoes88': opcoes88.get(), 'opcoes810': opcoes810.get(), 'opcoes811': opcoes811.get(), 'opcoes812': opcoes812.get(), 'opcoes813': opcoes813.get(), 'opcoes814': opcoes814.get(), 'opcoes815': opcoes815.get(), 'opcoes817': opcoes817.get(), 'texto81': texto81.get(), 'texto82': texto82.get()}
      elif flag83 is False and flag84 is True:
            dados_aba8 = {'opcoes81': opcoes81.get(), 'opcoes82': opcoes82.get(), 'opcoes83': opcoes83.get(), 'opcoes84': opcoes84.get(), 'opcoes85': opcoes85.get(), 'opcoes86': opcoes86.get(), 'opcoes87': opcoes87.get(), 'opcoes88': opcoes88.get(), 'opcoes810': opcoes810.get(), 'opcoes811': opcoes811.get(), 'opcoes812': opcoes812.get(), 'opcoes813': opcoes813.get(), 'opcoes814': opcoes814.get(), 'opcoes815': opcoes815.get(), 'opcoes816': opcoes816.get(), 'opcoes817': opcoes817.get(), 'texto81': texto81.get(), 'texto82': texto82.get()}
      elif flag83 is True and flag84 is False:
            dados_aba8 = {'opcoes81': opcoes81.get(), 'opcoes82': opcoes82.get(), 'opcoes83': opcoes83.get(), 'opcoes84': opcoes84.get(), 'opcoes85': opcoes85.get(), 'opcoes86': opcoes86.get(), 'opcoes87': opcoes87.get(), 'opcoes88': opcoes88.get(), 'opcoes89': opcoes89.get(), 'opcoes810': opcoes810.get(), 'opcoes811': opcoes811.get(), 'opcoes812': opcoes812.get(), 'opcoes813': opcoes813.get(), 'opcoes814': opcoes814.get(), 'opcoes815': opcoes815.get(), 'opcoes817': opcoes817.get(), 'texto81': texto81.get(), 'texto82': texto82.get()}
      elif flag83 is True and flag84 is True:
            dados_aba8 = {'opcoes81': opcoes81.get(), 'opcoes82': opcoes82.get(), 'opcoes83': opcoes83.get(), 'opcoes84': opcoes84.get(), 'opcoes85': opcoes85.get(), 'opcoes86': opcoes86.get(), 'opcoes87': opcoes87.get(), 'opcoes88': opcoes88.get(), 'opcoes89': opcoes89.get(), 'opcoes810': opcoes810.get(), 'opcoes811': opcoes811.get(), 'opcoes812': opcoes812.get(), 'opcoes813': opcoes813.get(), 'opcoes814': opcoes814.get(), 'opcoes815': opcoes815.get(), 'opcoes816': opcoes816.get(), 'opcoes817': opcoes817.get(), 'texto81': texto81.get(), 'texto82': texto82.get()}

   # Compila os dados da aba 9 (Refrigeração)
   dados_aba9 = {'opcoes91': opcoes91.get(), 'opcoes92': opcoes92.get(), 'opcoes93': opcoes93.get(), 'opcoes94': opcoes94.get(), 'opcoes95': opcoes95.get(), 'opcoes96': opcoes96.get(), 'opcoes97': opcoes97.get(), 'opcoes98': opcoes98.get(), 'lt91': lt91.get(), 'inc91': inc91.get()}
   
   # Compila os dados da aba 10 (Comp. gerais/Avulsos)
   dados_aba10 = {'opcoes101': opcoes101.get(), 'opcoes102': opcoes102.get(), 'textolongo101': textolongo101.get("1.0",'end-1c'), 'lt101': lt101.get(), 'inc101': inc101.get()}

   # Compila os dados da aba 11 (Executar)
   dados_aba11 = {'pasta_escolhida': pasta_escolhida}

   #print(dados_aba1)
   print(dados_aba2)
   print(dados_aba3)
   print(dados_aba4)
   #print(dados_aba5)
   #print(dados_aba6)
   #print(dados_aba7)
   print(dados_aba9)
   #print(dados_aba10)
   #print(dados_aba11)

   # Criando validação do formulário para ativar o botão de imprimir tabela
   flag_botao = True
   for a1 in dados_aba1.values():
      if a1 == '' or a1 == 'Selecionar...' or a1 == 'Digitar nome do projeto...' or a1 == 'Digitar material...' or a1 == 'Digitar login...' or a1 == 'Digitar ordem...':
         flag_botao = False
   for a2 in dados_aba2.values():
      if a2 == '' or a2 == 'Selecionar...' or a2 == 'Digitar número...':
         flag_botao = False
   for a3 in dados_aba3.values():
      if a3 == '' or a3 == 'Selecionar...' or a3 == 'Digitar número...':
         flag_botao = False
   for a4 in dados_aba4.values():
      if a4 == '' or a4 == 'Selecionar...' or a4 == 'Digitar número...':
         flag_botao = False
   for a5 in dados_aba5.values():
      if flag51 == True:
         if a5 == '' or a5 == 'Selecionar...' or a5 == 'Digitar nome...' or a5 == 'Digitar número...':
            flag_botao = False
   for a6 in dados_aba6.values():
      if flag61 == True:
         if a6 == '' or a6 == 'Selecionar...' or a6 == 'Digitar nome...' or a6 == 'Digitar número...':
            flag_botao = False
   for a7 in dados_aba7.values():
      if flag71 == True:
         if a7 == '' or a7 == 'Selecionar...' or a7 == 'Digitar nome...' or a7 == 'Digitar número...':
            flag_botao = False
   for a9 in dados_aba9.values():
      if a9 == 'Selecionar...' or a9 == 'Digitar número...':
         flag_botao = False
   for a10 in dados_aba10.values():
      if a10 == 'Selecionar...':
         flag_botao = False
   for a11 in dados_aba11.values():
      if a11 == '' or a11 == 'Escolher pasta...' or a11 == '(Escolha uma pasta clicando no botão ao lado)':
         flag_botao = False

   # Função que deleta popup de aviso para preencher
   def deleta_popup_erro():
      if flag_botao == False:
         top.destroy()


   # Criando popup de aviso para preencher todos os campos
   if flag_botao == False:
      top = Toplevel(root)
      top.geometry("500x250")
      top.title("Aviso")
      if var.get() == 1:
         Label(top, text= 'Preencha todos os campos e escolha uma pasta \npara salvar o resultado antes de continuar!', font= '20').place(x=80,y=80)
      elif var.get()== 2:
         Label(top, text= 'Fill in all fields and choose a folder\n to save the output file to continue!', font= '20').place(x=120,y=80)

      Button(top, text='OK', font='20', command=deleta_popup_erro).place(x=220, y=150)

   elif flag_botao is True:

      # Imprimindo dados da aba 1 (Informações do projeto) na planilha resposta ===================================================================================================
      aba_pa['E3'].value = str(dados_aba1['texto11'])
      aba_pa['E4'].value = str(dados_aba1['texto12'])
      aba_pa['E5'].value = str(dados_aba1['texto14'])
      aba_pa['E6'].value = str(dados_aba1['texto13'])


      # Regras da aba 2 (Carcaça/Estator) =========================================================================================================================================
      if varcb1.get() == 1:
         aba_pa['E11'].value = 'Especificado pelo maestro/projeto elétrico'
         aba_pa['E12'].value = 'Especificado pelo maestro/projeto elétrico'
         aba_pa['E13'].value = 'Especificado pelo maestro/projeto elétrico'
         aba_pa['E14'].value = 'Especificado pelo maestro/projeto elétrico'

         aba_pa['D11'].value = 'N/A'
         aba_pa['D12'].value = 'N/A'
         aba_pa['D13'].value = 'N/A'
         aba_pa['D14'].value = 'N/A'
         
      if varcb2.get() == 1:
         aba_pa['E16'].value = 'Especificado pelo maestro/projeto elétrico'
         aba_pa['E17'].value = 'Especificado pelo maestro/projeto elétrico'
         aba_pa['E18'].value = 'Especificado pelo maestro/projeto elétrico'
         aba_pa['E19'].value = 'Especificado pelo maestro/projeto elétrico'

         aba_pa['D16'].value = 'N/A'
         aba_pa['D17'].value = 'N/A'
         aba_pa['D18'].value = 'N/A'
         aba_pa['D19'].value = 'N/A'
      
      # Conjunto fixação resistor de aquecimento (lado esquerdo)
      if opcoes23.get() == '1' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 304':
         cj_fix_le = aba_resist_aquec['C6'].value
      elif opcoes23.get() == '1' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 316':
         cj_fix_le = aba_resist_aquec['C7'].value
      elif opcoes23.get() == '1' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 316L':
         cj_fix_le = aba_resist_aquec['C8'].value
      elif opcoes23.get() == '1' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Aço carbono':
         cj_fix_le = aba_resist_aquec['C5'].value
      elif opcoes23.get() == '2' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 304':
         cj_fix_le = aba_resist_aquec['C11'].value
      elif opcoes23.get() == '2' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 316':
         cj_fix_le = aba_resist_aquec['C12'].value
      elif opcoes23.get() == '2' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 316L':
         cj_fix_le = aba_resist_aquec['C13'].value
      elif opcoes23.get() == '2' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Aço carbono':
         cj_fix_le = aba_resist_aquec['C10'].value
      elif opcoes23.get() == '3' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 304':
         cj_fix_le = aba_resist_aquec['C16'].value
      elif opcoes23.get() == '3' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 316':
         cj_fix_le = aba_resist_aquec['C17'].value
      elif opcoes23.get() == '3' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 316L':
         cj_fix_le = aba_resist_aquec['C18'].value
      elif opcoes23.get() == '3' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Aço carbono':
         cj_fix_le = aba_resist_aquec['C15'].value
      elif opcoes23.get() == '4' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 304':
         cj_fix_le = aba_resist_aquec['C21'].value
      elif opcoes23.get() == '4' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 316':
         cj_fix_le = aba_resist_aquec['C22'].value
      elif opcoes23.get() == '4' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 316L':
         cj_fix_le = aba_resist_aquec['C23'].value
      elif opcoes23.get() == '4' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Aço carbono':
         cj_fix_le = aba_resist_aquec['C20'].value
      elif opcoes23.get() == '0':
         cj_fix_le = 'N/A'

      # Conjunto fixação resistor de aquecimento (lado direito)
      if opcoes24.get() == '1' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 304':
         cj_fix_ld = aba_resist_aquec['C6'].value
      elif opcoes24.get() == '1' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 316':
         cj_fix_ld = aba_resist_aquec['C7'].value
      elif opcoes24.get() == '1' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 316L':
         cj_fix_ld = aba_resist_aquec['C8'].value
      elif opcoes24.get() == '1' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Aço carbono':
         cj_fix_ld = aba_resist_aquec['C5'].value
      elif opcoes24.get() == '2' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 304':
         cj_fix_ld = aba_resist_aquec['C11'].value
      elif opcoes24.get() == '2' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 316':
         cj_fix_ld = aba_resist_aquec['C12'].value
      elif opcoes24.get() == '2' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 316L':
         cj_fix_ld = aba_resist_aquec['C13'].value
      elif opcoes24.get() == '2' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Aço carbono':
         cj_fix_ld = aba_resist_aquec['C10'].value
      elif opcoes24.get() == '3' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 304':
         cj_fix_ld = aba_resist_aquec['C16'].value
      elif opcoes24.get() == '3' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 316':
         cj_fix_ld = aba_resist_aquec['C17'].value
      elif opcoes24.get() == '3' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 316L':
         cj_fix_ld = aba_resist_aquec['C18'].value
      elif opcoes24.get() == '3' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Aço carbono':
         cj_fix_ld = aba_resist_aquec['C15'].value
      elif opcoes24.get() == '4' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 304':
         cj_fix_ld = aba_resist_aquec['C21'].value
      elif opcoes24.get() == '4' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 316':
         cj_fix_ld = aba_resist_aquec['C22'].value
      elif opcoes24.get() == '4' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Inox 316L':
         cj_fix_ld = aba_resist_aquec['C23'].value
      elif opcoes24.get() == '4' and opcoes21.get() == 'Tipo "W"' and opcoes13.get() == 'Aço carbono':
         cj_fix_ld = aba_resist_aquec['C20'].value
      elif opcoes24.get() == '0':
         cj_fix_ld = 'N/A'
      
      # Conjunto tampa resistores de aquecimento
      if opcoes21.get() == 'Tipo "W"' and int(opcoes14.get()) == 280:
         cj_tampa_resist = aba_resist_aquec['G4'].value
      elif opcoes21.get() == 'Tipo "W"' and int(opcoes14.get()) == 315:
         cj_tampa_resist = aba_resist_aquec['G4'].value
      elif opcoes21.get() == 'Tipo "W"' and int(opcoes14.get()) >= 355 and int(opcoes14.get()) <= 560:
         cj_tampa_resist = aba_resist_aquec['G4'].value
      elif opcoes21.get() == 'Tipo "W"' and int(opcoes14.get()) >= 630 and int(opcoes14.get()) <= 1250:
         cj_tampa_resist = aba_resist_aquec['G4'].value
      elif opcoes21.get() == 'Tipo "W"' and int(opcoes14.get()) > 1250:
         cj_tampa_resist = aba_resist_aquec['G4'].value

      # Caso seja selecionado Resistor tipo "U" com cabeçote
      if opcoes21.get() == 'Roscado tipo "U" com cabeçote' and opcoes13.get() == 'Inox 304':
         cabo_lig = aba_resist_aquec['K11'].value
         term_olhal = aba_resist_aquec['K12'].value
         cond_G1 = aba_resist_aquec['K13'].value
         con_MGR_G1 = aba_resist_aquec['K14'].value
         abrac_cond_G1 = aba_resist_aquec['K15'].value
         paraf_abrac_G1 = aba_resist_aquec['K16'].value
      elif opcoes21.get() == 'Roscado tipo "U" com cabeçote' and opcoes13.get() == 'Inox 316':
         cabo_lig = aba_resist_aquec['K18'].value
         term_olhal = aba_resist_aquec['K19'].value
         cond_G1 = aba_resist_aquec['K20'].value
         con_MGR_G1 = aba_resist_aquec['K21'].value
         abrac_cond_G1 = aba_resist_aquec['K22'].value
         paraf_abrac_G1 = aba_resist_aquec['K23'].value
      elif opcoes21.get() == 'Roscado tipo "U" com cabeçote' and opcoes13.get() == 'Inox 316L':
         cabo_lig = aba_resist_aquec['K25'].value
         term_olhal = aba_resist_aquec['K26'].value
         cond_G1 = aba_resist_aquec['K27'].value
         con_MGR_G1 = aba_resist_aquec['K28'].value
         abrac_cond_G1 = aba_resist_aquec['K29'].value
         paraf_abrac_G1 = aba_resist_aquec['K30'].value
      elif opcoes21.get() == 'Roscado tipo "U" com cabeçote' and opcoes13.get() == 'Aço carbono':
         cabo_lig = aba_resist_aquec['K4'].value
         term_olhal = aba_resist_aquec['K5'].value
         cond_G1 = aba_resist_aquec['K6'].value
         con_MGR_G1 = aba_resist_aquec['K7'].value
         abrac_cond_G1 = aba_resist_aquec['K8'].value
         paraf_abrac_G1 = aba_resist_aquec['K9'].value
      
      # Caso seja selecionado Resistor tipo "U" sem cabeçote
      if opcoes21.get() == 'Roscado tipo "U" sem cabeçote':
         cabo_lig = aba_resist_aquec['O4'].value
         term_olhal = aba_resist_aquec['O5'].value
      
      # Imprimindo os dados da aba 2 na planilha
      aba_pa['D30'].value = int(lt21.get())
      if opcoes21.get() == 'Tipo "W"':
         aba_pa['E31'].value = cj_fix_le
         aba_pa['E32'].value = cj_fix_ld
         aba_pa['E33'].value = cj_tampa_resist

         aba_pa['H31'].value = '1'
         aba_pa['H32'].value = '1'
         aba_pa['H33'].value = '2'

         # Imprimindo as posições da LT para tipo "W"
         aba_pa['D31'].value = int(lt21.get()) + int(inc21.get())
         aba_pa['D32'].value = int(lt21.get()) + 2 * int(inc21.get())
         aba_pa['D33'].value = int(lt21.get()) + 3 * int(inc21.get())
         aba_pa['D40'].value = int(lt21.get()) + 4 * int(inc21.get())

      elif opcoes21.get() == 'Roscado tipo "U" com cabeçote':
         aba_pa['E34'].value = cabo_lig
         aba_pa['H34'].value = float(aba_resist_aquec['L4'].value) * float(opcoes22.get())
         aba_pa['E35'].value = term_olhal
         aba_pa['H35'].value = float(aba_resist_aquec['L5'].value) * float(opcoes22.get())
         aba_pa['E36'].value = cond_G1
         aba_pa['H36'].value = float(aba_resist_aquec['L6'].value) * float(opcoes22.get())
         aba_pa['E37'].value = con_MGR_G1
         aba_pa['H37'].value = float(aba_resist_aquec['L7'].value) * float(opcoes22.get())
         aba_pa['E38'].value = abrac_cond_G1
         aba_pa['H38'].value = float(aba_resist_aquec['L8'].value) * float(opcoes22.get())
         aba_pa['E39'].value = paraf_abrac_G1
         aba_pa['H39'].value = float(aba_resist_aquec['L9'].value) * float(opcoes22.get())

         # Imprimindo as posições da LT para tipo "U" com cabeçote
         incremento = int(dados_aba2['inc21'])
         pos_inic = int(dados_aba2['lt21'])
         celulas = 'D' + str(34)
         for m in range(7):
            aba_pa[celulas].value = pos_inic + (m * incremento)
            celulas = 'D' + str(34 + m + 1)
      
      elif opcoes21.get() == 'Roscado tipo "U" sem cabeçote':
         aba_pa['E34'].value = cabo_lig
         aba_pa['H34'].value = float(aba_resist_aquec['L4'].value) * float(opcoes22.get())
         aba_pa['E35'].value = term_olhal
         aba_pa['H35'].value = float(aba_resist_aquec['L5'].value) * float(opcoes22.get())

         aba_pa['D34'].value = int(lt21.get()) + 4 * int(inc21.get())
         aba_pa['D35'].value = int(lt21.get()) + 5 * int(inc21.get())
         aba_pa['D40'].value = int(lt21.get()) + 10 * int(inc21.get())


      # Regras da aba 3 (Mancal L.A.) =============================================================================================================================================
      # Frame termoresistor na axial, na radial e reservatório de óleo
      if opcoes13.get() == 'Aço carbono':
         conduite_aba3 = aba_sensorTemp['B5'].value
         qtd_conduite_aba3 = float(aba_sensorTemp['C5'].value) * float(opcoes33.get())
         luva_aba3 = aba_sensorTemp['B6'].value
         qtd_luva_aba3 = float(aba_sensorTemp['C6'].value) * float(opcoes33.get())
         conMacho_aba3 = aba_sensorTemp['B7'].value
         qtd_conMacho_aba3 = float(aba_sensorTemp['C7'].value) * float(opcoes33.get())
         abracadeira_aba3 = aba_sensorTemp['B8'].value
         qtd_abracadeira_aba3 = float(aba_sensorTemp['C8'].value) * float(opcoes33.get())
         parafuso_aba3 = aba_sensorTemp['B9'].value
         qtd_parafuso_aba3 = float(aba_sensorTemp['C9'].value) * float(opcoes33.get())
      elif opcoes13.get() == 'Inox 304':
         conduite_aba3 = aba_sensorTemp['B12'].value
         qtd_conduite_aba3 = float(aba_sensorTemp['C12'].value) * float(opcoes33.get())
         luva_aba3 = aba_sensorTemp['B13'].value
         qtd_luva_aba3 = float(aba_sensorTemp['C13'].value) * float(opcoes33.get())
         conMacho_aba3 = aba_sensorTemp['B14'].value
         qtd_conMacho_aba3 = float(aba_sensorTemp['C14'].value) * float(opcoes33.get())
         abracadeira_aba3 = aba_sensorTemp['B15'].value
         qtd_abracadeira_aba3 = float(aba_sensorTemp['C15'].value) * float(opcoes33.get())
         parafuso_aba3 = aba_sensorTemp['B15'].value
         qtd_parafuso_aba3 = float(aba_sensorTemp['C15'].value) * float(opcoes33.get())
      elif opcoes13.get() == 'Inox 316':
         conduite_aba3 = aba_sensorTemp['B19'].value
         qtd_conduite_aba3 = float(aba_sensorTemp['C19'].value) * float(opcoes33.get())
         luva_aba3 = aba_sensorTemp['B20'].value
         qtd_luva_aba3 = float(aba_sensorTemp['C20'].value) * float(opcoes33.get())
         conMacho_aba3 = aba_sensorTemp['B21'].value
         qtd_conMacho_aba3 = float(aba_sensorTemp['C21'].value) * float(opcoes33.get())
         abracadeira_aba3 = aba_sensorTemp['B22'].value
         qtd_abracadeira_aba3 = float(aba_sensorTemp['C22'].value) * float(opcoes33.get())
         parafuso_aba3 = aba_sensorTemp['B23'].value
         qtd_parafuso_aba3 = float(aba_sensorTemp['C23'].value) * float(opcoes33.get())
      elif opcoes13.get() == 'Inox 316L':
         conduite_aba3 = aba_sensorTemp['B26'].value
         qtd_conduite_aba3 = float(aba_sensorTemp['C26'].value) * float(opcoes33.get())
         luva_aba3 = aba_sensorTemp['B27'].value
         qtd_luva_aba3 = float(aba_sensorTemp['C27'].value) * float(opcoes33.get())
         conMacho_aba3 = aba_sensorTemp['B28'].value
         qtd_conMacho_aba3 = float(aba_sensorTemp['C28'].value) * float(opcoes33.get())
         abracadeira_aba3 = aba_sensorTemp['B29'].value
         qtd_abracadeira_aba3 = float(aba_sensorTemp['C29'].value) * float(opcoes33.get())
         parafuso_aba3 = aba_sensorTemp['B30'].value
         qtd_parafuso_aba3 = float(aba_sensorTemp['C30'].value) * float(opcoes33.get())

      # Frame termoresistor na radial
      if opcoes31.get() == 'Isolado':
         adap_isolante_aba3 = aba_sensorTemp['H2'].value
         qtd_adap_isolante_aba3 = aba_sensorTemp['C4'].value * float(opcoes34.get())
      
      if opcoes13.get() == 'Aço carbono':
         conduite_aba3 = aba_sensorTemp['B5'].value
         qtd_conduite_aba3 = float(aba_sensorTemp['C5'].value) * float(opcoes34.get())
         luva_aba3 = aba_sensorTemp['B6'].value
         qtd_luva_aba3 = float(aba_sensorTemp['C6'].value) * float(opcoes34.get())
         conMacho_aba3 = aba_sensorTemp['B7'].value
         qtd_conMacho_aba3 = float(aba_sensorTemp['C7'].value) * float(opcoes34.get())
         abracadeira_aba3 = aba_sensorTemp['B8'].value
         qtd_abracadeira_aba3 = float(aba_sensorTemp['C8'].value) * float(opcoes34.get())
         parafuso_aba3 = aba_sensorTemp['B9'].value
         qtd_parafuso_aba3 = float(aba_sensorTemp['C9'].value) * float(opcoes34.get())
      elif opcoes13.get() == 'Inox 304':
         conduite_aba3 = aba_sensorTemp['B12'].value
         qtd_conduite_aba3 = float(aba_sensorTemp['C12'].value) * float(opcoes34.get())
         luva_aba3 = aba_sensorTemp['B13'].value
         qtd_luva_aba3 = float(aba_sensorTemp['C13'].value) * float(opcoes34.get())
         conMacho_aba3 = aba_sensorTemp['B14'].value
         qtd_conMacho_aba3 = float(aba_sensorTemp['C14'].value) * float(opcoes34.get())
         abracadeira_aba3 = aba_sensorTemp['B15'].value
         qtd_abracadeira_aba3 = float(aba_sensorTemp['C15'].value) * float(opcoes34.get())
         parafuso_aba3 = aba_sensorTemp['B15'].value
         qtd_parafuso_aba3 = float(aba_sensorTemp['C15'].value) * float(opcoes34.get())
      elif opcoes13.get() == 'Inox 316':
         conduite_aba3 = aba_sensorTemp['B19'].value
         qtd_conduite_aba3 = float(aba_sensorTemp['C19'].value) * float(opcoes34.get())
         luva_aba3 = aba_sensorTemp['B20'].value
         qtd_luva_aba3 = float(aba_sensorTemp['C20'].value) * float(opcoes34.get())
         conMacho_aba3 = aba_sensorTemp['B21'].value
         qtd_conMacho_aba3 = float(aba_sensorTemp['C21'].value) * float(opcoes34.get())
         abracadeira_aba3 = aba_sensorTemp['B22'].value
         qtd_abracadeira_aba3 = float(aba_sensorTemp['C22'].value) * float(opcoes34.get())
         parafuso_aba3 = aba_sensorTemp['B23'].value
         qtd_parafuso_aba3 = float(aba_sensorTemp['C23'].value) * float(opcoes34.get())
      elif opcoes13.get() == 'Inox 316L':
         conduite_aba3 = aba_sensorTemp['B26'].value
         qtd_conduite_aba3 = float(aba_sensorTemp['C26'].value) * float(opcoes34.get())
         luva_aba3 = aba_sensorTemp['B27'].value
         qtd_luva_aba3 = float(aba_sensorTemp['C27'].value) * float(opcoes34.get())
         conMacho_aba3 = aba_sensorTemp['B28'].value
         qtd_conMacho_aba3 = float(aba_sensorTemp['C28'].value) * float(opcoes34.get())
         abracadeira_aba3 = aba_sensorTemp['B29'].value
         qtd_abracadeira_aba3 = float(aba_sensorTemp['C29'].value) * float(opcoes34.get())
         parafuso_aba3 = aba_sensorTemp['B30'].value
         qtd_parafuso_aba3 = float(aba_sensorTemp['C30'].value) * float(opcoes34.get())

      # Frame termoresistor no reservatório de óleo
      if opcoes13.get() == 'Aço carbono':
         conduite_aba3 = aba_sensorTemp['B5'].value
         qtd_conduite_aba3 = float(aba_sensorTemp['C5'].value) * float(opcoes35.get())
         luva_aba3 = aba_sensorTemp['B6'].value
         qtd_luva_aba3 = float(aba_sensorTemp['C6'].value) * float(opcoes35.get())
         conMacho_aba3 = aba_sensorTemp['B7'].value
         qtd_conMacho_aba3 = float(aba_sensorTemp['C7'].value) * float(opcoes35.get())
         abracadeira_aba3 = aba_sensorTemp['B8'].value
         qtd_abracadeira_aba3 = float(aba_sensorTemp['C8'].value) * float(opcoes35.get())
         parafuso_aba3 = aba_sensorTemp['B9'].value
         qtd_parafuso_aba3 = float(aba_sensorTemp['C9'].value) * float(opcoes35.get())
      elif opcoes13.get() == 'Inox 304':
         conduite_aba3 = aba_sensorTemp['B12'].value
         qtd_conduite_aba3 = float(aba_sensorTemp['C12'].value) * float(opcoes35.get())
         luva_aba3 = aba_sensorTemp['B13'].value
         qtd_luva_aba3 = float(aba_sensorTemp['C13'].value) * float(opcoes35.get())
         conMacho_aba3 = aba_sensorTemp['B14'].value
         qtd_conMacho_aba3 = float(aba_sensorTemp['C14'].value) * float(opcoes35.get())
         abracadeira_aba3 = aba_sensorTemp['B15'].value
         qtd_abracadeira_aba3 = float(aba_sensorTemp['C15'].value) * float(opcoes35.get())
         parafuso_aba3 = aba_sensorTemp['B15'].value
         qtd_parafuso_aba3 = float(aba_sensorTemp['C15'].value) * float(opcoes35.get())
      elif opcoes13.get() == 'Inox 316':
         conduite_aba3 = aba_sensorTemp['B19'].value
         qtd_conduite_aba3 = float(aba_sensorTemp['C19'].value) * float(opcoes35.get())
         luva_aba3 = aba_sensorTemp['B20'].value
         qtd_luva_aba3 = float(aba_sensorTemp['C20'].value) * float(opcoes35.get())
         conMacho_aba3 = aba_sensorTemp['B21'].value
         qtd_conMacho_aba3 = float(aba_sensorTemp['C21'].value) * float(opcoes35.get())
         abracadeira_aba3 = aba_sensorTemp['B22'].value
         qtd_abracadeira_aba3 = float(aba_sensorTemp['C22'].value) * float(opcoes35.get())
         parafuso_aba3 = aba_sensorTemp['B23'].value
         qtd_parafuso_aba3 = float(aba_sensorTemp['C23'].value) * float(opcoes35.get())
      elif opcoes13.get() == 'Inox 316L':
         conduite_aba3 = aba_sensorTemp['B26'].value
         qtd_conduite_aba3 = float(aba_sensorTemp['C26'].value) * float(opcoes35.get())
         luva_aba3 = aba_sensorTemp['B27'].value
         qtd_luva_aba3 = float(aba_sensorTemp['C27'].value) * float(opcoes35.get())
         conMacho_aba3 = aba_sensorTemp['B28'].value
         qtd_conMacho_aba3 = float(aba_sensorTemp['C28'].value) * float(opcoes35.get())
         abracadeira_aba3 = aba_sensorTemp['B29'].value
         qtd_abracadeira_aba3 = float(aba_sensorTemp['C29'].value) * float(opcoes35.get())
         parafuso_aba3 = aba_sensorTemp['B30'].value
         qtd_parafuso_aba3 = float(aba_sensorTemp['C30'].value) * float(opcoes35.get())

      if opcoes36.get() == 'G 3/4"':
         rosca_fix_res_aba3 = aba_termometros['K2'].value
      elif opcoes36.get() == 'G 1.1/4"':
         rosca_fix_res_aba3 = aba_termometros['K3'].value
      elif opcoes36.get() == 'G 1"':
         rosca_fix_res_aba3 = aba_termometros['K4'].value
      
      # Imprimindo os dados de termorresistor para instalação axial
      aba_pa['E47'].value = conduite_aba3
      aba_pa['H47'].value = qtd_conduite_aba3
      aba_pa['E48'].value = luva_aba3
      aba_pa['H48'].value = qtd_luva_aba3
      aba_pa['E49'].value = conMacho_aba3
      aba_pa['H49'].value = qtd_conMacho_aba3
      aba_pa['E50'].value = abracadeira_aba3
      aba_pa['H50'].value = qtd_abracadeira_aba3
      aba_pa['E51'].value = parafuso_aba3
      aba_pa['H51'].value = qtd_parafuso_aba3

      pos_inic = int(dados_aba3['lt31'])
      incremento = int(dados_aba3['inc31'])
      celulas = 'D' + str(45)
      for m in range(7):
         aba_pa[celulas].value = pos_inic + (m * incremento)
         celulas = 'D' + str(45 + m + 1)

      # Imprimindo os dados de termorresistor para instalação radial
      aba_pa['E55'].value = adap_isolante_aba3
      aba_pa['H55'].value = qtd_adap_isolante_aba3
      aba_pa['E56'].value = conduite_aba3
      aba_pa['H56'].value = qtd_conduite_aba3
      aba_pa['E57'].value = luva_aba3
      aba_pa['H57'].value = qtd_luva_aba3
      aba_pa['E58'].value = conMacho_aba3
      aba_pa['H58'].value = qtd_conMacho_aba3
      aba_pa['E59'].value = abracadeira_aba3
      aba_pa['H59'].value = qtd_abracadeira_aba3
      aba_pa['E60'].value = parafuso_aba3
      aba_pa['H60'].value = qtd_parafuso_aba3

      pos_inic = int(aba_pa['D51'].value + incremento)
      incremento = int(dados_aba3['inc31'])
      celulas = 'D' + str(53)
      for m in range(8):
         aba_pa[celulas].value = pos_inic + (m * incremento)
         celulas = 'D' + str(53 + m + 1)

      # Imprimindo os dados de termorresistor para instalação no rservatório de óleo
      aba_pa['E64'].value = conduite_aba3
      aba_pa['H64'].value = qtd_conduite_aba3
      aba_pa['E65'].value = luva_aba3
      aba_pa['H65'].value = qtd_luva_aba3
      aba_pa['E66'].value = conMacho_aba3
      aba_pa['H66'].value = qtd_conMacho_aba3
      aba_pa['E67'].value = abracadeira_aba3
      aba_pa['H67'].value = qtd_abracadeira_aba3
      aba_pa['E68'].value = parafuso_aba3
      aba_pa['H68'].value = qtd_parafuso_aba3
      aba_pa['E69'].value = rosca_fix_res_aba3
      aba_pa['H69'].value = opcoes35.get()

      pos_inic = int(aba_pa['D60'].value + incremento)
      incremento = int(dados_aba3['inc31'])
      celulas = 'D' + str(62)
      for m in range(9):
         aba_pa[celulas].value = pos_inic + (m * incremento)
         celulas = 'D' + str(62 + m + 1)

      # Frame termômetro na radial
      if opcoes310.get() == 'Sem contato' and opcoes13.get() == 'Aço carbono' and opcoes31.get() == 'Isolado':
         cj_fix_termom_aba3 = aba_termometros['D3'].value
      elif opcoes310.get() == 'Sem contato' and opcoes13.get() == 'Inox 304' and opcoes31.get() == 'Isolado':
         cj_fix_termom_aba3 = aba_termometros['D4'].value
      elif opcoes310.get() == 'Sem contato' and opcoes13.get() == 'Inox 316' and opcoes31.get() == 'Isolado':
         cj_fix_termom_aba3 = aba_termometros['D5'].value
      elif opcoes310.get() == 'Sem contato' and opcoes13.get() == 'Inox 316L' and opcoes31.get() == 'Isolado':
         cj_fix_termom_aba3 = aba_termometros['D6'].value
      elif opcoes310.get() == 'Sem contato' and opcoes13.get() == 'Aço carbono' and opcoes31.get() == 'Não isolado':
         cj_fix_termom_aba3 = aba_termometros['D7'].value
      elif opcoes310.get() == 'Sem contato' and opcoes13.get() == 'Inox 304' and opcoes31.get() == 'Não isolado':
         cj_fix_termom_aba3 = aba_termometros['D8'].value
      elif opcoes310.get() == 'Sem contato' and opcoes13.get() == 'Inox 316' and opcoes31.get() == 'Não isolado':
         cj_fix_termom_aba3 = aba_termometros['D9'].value
      elif opcoes310.get() == 'Sem contato' and opcoes13.get() == 'Inox 316L' and opcoes31.get() == 'Não isolado':
         cj_fix_termom_aba3 = aba_termometros['D10'].value
      
      if opcoes310.get() == 'Com contato' and opcoes13.get() == 'Aço carbono' and opcoes31.get() == 'Isolado':
         cj_fix_termom_aba3 = aba_termometros['D3'].value
      elif opcoes310.get() == 'Com contato' and opcoes13.get() == 'Inox 304' and opcoes31.get() == 'Isolado':
         cj_fix_termom_aba3 = aba_termometros['D4'].value
      elif opcoes310.get() == 'Com contato' and opcoes13.get() == 'Inox 316' and opcoes31.get() == 'Isolado':
         cj_fix_termom_aba3 = aba_termometros['D5'].value
      elif opcoes310.get() == 'Com contato' and opcoes13.get() == 'Inox 316L' and opcoes31.get() == 'Isolado':
         cj_fix_termom_aba3 = aba_termometros['D6'].value
      elif opcoes310.get() == 'Com contato' and opcoes13.get() == 'Aço carbono' and opcoes31.get() == 'Não isolado':
         cj_fix_termom_aba3 = aba_termometros['D7'].value
      elif opcoes310.get() == 'Com contato' and opcoes13.get() == 'Inox 304' and opcoes31.get() == 'Não isolado':
         cj_fix_termom_aba3 = aba_termometros['D8'].value
      elif opcoes310.get() == 'Com contato' and opcoes13.get() == 'Inox 316' and opcoes31.get() == 'Não isolado':
         cj_fix_termom_aba3 = aba_termometros['D9'].value
      elif opcoes310.get() == 'Com contato' and opcoes13.get() == 'Inox 316L' and opcoes31.get() == 'Não isolado':
         cj_fix_termom_aba3 = aba_termometros['D10'].value
      
      if opcoes310.get() == 'Com contato' and opcoes13.get() == 'Aço carbono':
         conduite_aba3 = aba_sensorTemp['B5'].value
         qtd_conduite_aba3 = float(aba_sensorTemp['C5'].value) * float(opcoes312.get())
         luva_aba3 = aba_sensorTemp['B6'].value
         qtd_luva_aba3 = float(aba_sensorTemp['C6'].value) * float(opcoes312.get())
         conMacho_aba3 = aba_sensorTemp['B7'].value
         qtd_conMacho_aba3 = float(aba_sensorTemp['C7'].value) * float(opcoes312.get())
         abracadeira_aba3 = aba_sensorTemp['B8'].value
         qtd_abracadeira_aba3 = float(aba_sensorTemp['C8'].value) * float(opcoes312.get())
         parafuso_aba3 = aba_sensorTemp['B9'].value
         qtd_parafuso_aba3 = float(aba_sensorTemp['C9'].value) * float(opcoes312.get())
      elif opcoes310.get() == 'Com contato' and opcoes13.get() == 'Inox 304':
         conduite_aba3 = aba_sensorTemp['B12'].value
         qtd_conduite_aba3 = float(aba_sensorTemp['C12'].value) * float(opcoes312.get())
         luva_aba3 = aba_sensorTemp['B13'].value
         qtd_luva_aba3 = float(aba_sensorTemp['C13'].value) * float(opcoes312.get())
         conMacho_aba3 = aba_sensorTemp['B14'].value
         qtd_conMacho_aba3 = float(aba_sensorTemp['C14'].value) * float(opcoes312.get())
         abracadeira_aba3 = aba_sensorTemp['B15'].value
         qtd_abracadeira_aba3 = float(aba_sensorTemp['C15'].value) * float(opcoes312.get())
         parafuso_aba3 = aba_sensorTemp['B16'].value
         qtd_parafuso_aba3 = float(aba_sensorTemp['C16'].value) * float(opcoes312.get())
      elif opcoes310.get() == 'Com contato' and opcoes13.get() == 'Inox 316':
         conduite_aba3 = aba_sensorTemp['B19'].value
         qtd_conduite_aba3 = float(aba_sensorTemp['C19'].value) * float(opcoes312.get())
         luva_aba3 = aba_sensorTemp['B20'].value
         qtd_luva_aba3 = float(aba_sensorTemp['C20'].value) * float(opcoes312.get())
         conMacho_aba3 = aba_sensorTemp['B21'].value
         qtd_conMacho_aba3 = float(aba_sensorTemp['C21'].value) * float(opcoes312.get())
         abracadeira_aba3 = aba_sensorTemp['B22'].value
         qtd_abracadeira_aba3 = float(aba_sensorTemp['C22'].value) * float(opcoes312.get())
         parafuso_aba3 = aba_sensorTemp['B23'].value
         qtd_parafuso_aba3 = float(aba_sensorTemp['C23'].value) * float(opcoes312.get())
      elif opcoes310.get() == 'Com contato' and opcoes13.get() == 'Inox 316L':
         conduite_aba3 = aba_sensorTemp['B26'].value
         qtd_conduite_aba3 = float(aba_sensorTemp['C26'].value) * float(opcoes312.get())
         luva_aba3 = aba_sensorTemp['B27'].value
         qtd_luva_aba3 = float(aba_sensorTemp['C27'].value) * float(opcoes312.get())
         conMacho_aba3 = aba_sensorTemp['B28'].value
         qtd_conMacho_aba3 = float(aba_sensorTemp['C28'].value) * float(opcoes312.get())
         abracadeira_aba3 = aba_sensorTemp['B29'].value
         qtd_abracadeira_aba3 = float(aba_sensorTemp['C29'].value) * float(opcoes312.get())
         parafuso_aba3 = aba_sensorTemp['B30'].value
         qtd_parafuso_aba3 = float(aba_sensorTemp['C30'].value) * float(opcoes312.get())

      if opcoes311.get() == 'Sim':
         sup_inclinado_aba3 = aba_termometros['H2'].value
      elif opcoes311.get() == 'Não':
         sup_inclinado_aba3 = aba_termometros['H3'].value

      # Imprimindo os dados da aba 3 na planilha para termômetro na radial
      aba_pa['E74'].value = cj_fix_termom_aba3
      aba_pa['H74'].value = opcoes312.get()
      aba_pa['E75'].value = conduite_aba3
      aba_pa['H75'].value = qtd_conduite_aba3 * int(opcoes312.get())
      aba_pa['E76'].value = luva_aba3
      aba_pa['H76'].value = qtd_luva_aba3 * int(opcoes312.get())
      aba_pa['E77'].value = conMacho_aba3
      aba_pa['H77'].value = qtd_conMacho_aba3 * int(opcoes312.get())
      aba_pa['E78'].value = abracadeira_aba3
      aba_pa['H78'].value = qtd_abracadeira_aba3 * int(opcoes312.get())
      aba_pa['E79'].value = parafuso_aba3
      aba_pa['H79'].value = qtd_parafuso_aba3 * int(opcoes312.get())
      aba_pa['E80'].value = sup_inclinado_aba3
      aba_pa['H80'].value = opcoes312.get()

      # Imprimindo as posições da LT na planilha para termômetros
      pos_inic = int(aba_pa['D70'].value + incremento)
      incremento = int(dados_aba3['inc31'])
      celulas = 'D' + str(73)
      for m in range(8):
         aba_pa[celulas].value = pos_inic + (m * incremento)
         celulas = 'D' + str(73 + m + 1)
         
      # Frame termômetro no reservatório de óleo
      if opcoes310.get() == 'Sem contato' and opcoes13.get() == 'Aço carbono' and opcoes31.get() == 'Isolado':
         cj_fix_termom_aba3 = aba_termometros['D3'].value
      elif opcoes310.get() == 'Sem contato' and opcoes13.get() == 'Inox 304' and opcoes31.get() == 'Isolado':
         cj_fix_termom_aba3 = aba_termometros['D4'].value
      elif opcoes310.get() == 'Sem contato' and opcoes13.get() == 'Inox 316' and opcoes31.get() == 'Isolado':
         cj_fix_termom_aba3 = aba_termometros['D5'].value
      elif opcoes310.get() == 'Sem contato' and opcoes13.get() == 'Inox 316L' and opcoes31.get() == 'Isolado':
         cj_fix_termom_aba3 = aba_termometros['D6'].value
      elif opcoes310.get() == 'Sem contato' and opcoes13.get() == 'Aço carbono' and opcoes31.get() == 'Não isolado':
         cj_fix_termom_aba3 = aba_termometros['D7'].value
      elif opcoes310.get() == 'Sem contato' and opcoes13.get() == 'Inox 304' and opcoes31.get() == 'Não isolado':
         cj_fix_termom_aba3 = aba_termometros['D8'].value
      elif opcoes310.get() == 'Sem contato' and opcoes13.get() == 'Inox 316' and opcoes31.get() == 'Não isolado':
         cj_fix_termom_aba3 = aba_termometros['D9'].value
      elif opcoes310.get() == 'Sem contato' and opcoes13.get() == 'Inox 316L' and opcoes31.get() == 'Não isolado':
         cj_fix_termom_aba3 = aba_termometros['D10'].value
      elif opcoes310.get() == 'Com contato' and opcoes13.get() == 'Aço carbono' and opcoes31.get() == 'Isolado':
         cj_fix_termom_aba3 = aba_termometros['D11'].value
      elif opcoes310.get() == 'Com contato' and opcoes13.get() == 'Inox 304' and opcoes31.get() == 'Isolado':
         cj_fix_termom_aba3 = aba_termometros['D12'].value
      elif opcoes310.get() == 'Com contato' and opcoes13.get() == 'Inox 316' and opcoes31.get() == 'Isolado':
         cj_fix_termom_aba3 = aba_termometros['D13'].value
      elif opcoes310.get() == 'Com contato' and opcoes13.get() == 'Inox 316L' and opcoes31.get() == 'Isolado':
         cj_fix_termom_aba3 = aba_termometros['D14'].value
      elif opcoes310.get() == 'Com contato' and opcoes13.get() == 'Aço carbono' and opcoes31.get() == 'Não isolado':
         cj_fix_termom_aba3 = aba_termometros['D15'].value
      elif opcoes310.get() == 'Com contato' and opcoes13.get() == 'Inox 304' and opcoes31.get() == 'Não isolado':
         cj_fix_termom_aba3 = aba_termometros['D16'].value
      elif opcoes310.get() == 'Com contato' and opcoes13.get() == 'Inox 316' and opcoes31.get() == 'Não isolado':
         cj_fix_termom_aba3 = aba_termometros['D17'].value
      elif opcoes310.get() == 'Com contato' and opcoes13.get() == 'Inox 316L' and opcoes31.get() == 'Não isolado':
         cj_fix_termom_aba3 = aba_termometros['D18'].value
      
      if 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Aço carbono':
         conduite = aba_sensorTemp['B5'].value
         qtd_conduite = float(aba_sensorTemp['C5'].value) * float(opcoes313.get())
         luva = aba_sensorTemp['B6'].value
         qtd_luva = float(aba_sensorTemp['C6'].value) * float(opcoes313.get())
         conMacho = aba_sensorTemp['B7'].value
         qtd_conMacho = float(aba_sensorTemp['C7'].value) * float(opcoes313.get())
         abracadeira = aba_sensorTemp['B8'].value
         qtd_abracadeira = float(aba_sensorTemp['C8'].value) * float(opcoes313.get())
         parafuso = aba_sensorTemp['B9'].value
         qtd_parafuso = float(aba_sensorTemp['C9'].value) * float(opcoes313.get())
      elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 304':
         conduite = aba_sensorTemp['B12'].value
         qtd_conduite = float(aba_sensorTemp['C12'].value) * float(opcoes313.get())
         luva = aba_sensorTemp['B13'].value
         qtd_luva = float(aba_sensorTemp['C13'].value) * float(opcoes313.get())
         conMacho = aba_sensorTemp['B14'].value
         qtd_conMacho = float(aba_sensorTemp['C14'].value) * float(opcoes313.get())
         abracadeira = aba_sensorTemp['B15'].value
         qtd_abracadeira = float(aba_sensorTemp['C15'].value) * float(opcoes313.get())
         parafuso = aba_sensorTemp['B16'].value
         qtd_parafuso = float(aba_sensorTemp['C16'].value) * float(opcoes313.get())
      elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316':
         conduite = aba_sensorTemp['B19'].value
         qtd_conduite = float(aba_sensorTemp['C19'].value) * float(opcoes313.get())
         luva = aba_sensorTemp['B20'].value
         qtd_luva = float(aba_sensorTemp['C20'].value) * float(opcoes313.get())
         conMacho = aba_sensorTemp['B21'].value
         qtd_conMacho = float(aba_sensorTemp['C21'].value) * float(opcoes313.get())
         abracadeira = aba_sensorTemp['B22'].value
         qtd_abracadeira = float(aba_sensorTemp['C22'].value) * float(opcoes313.get())
         parafuso = aba_sensorTemp['B23'].value
         qtd_parafuso = float(aba_sensorTemp['C23'].value) * float(opcoes313.get())
      elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316L':
         conduite = aba_sensorTemp['B26'].value
         qtd_conduite = float(aba_sensorTemp['C26'].value) * float(opcoes313.get())
         luva = aba_sensorTemp['B27'].value
         qtd_luva = float(aba_sensorTemp['C27'].value) * float(opcoes313.get())
         conMacho = aba_sensorTemp['B28'].value
         qtd_conMacho = float(aba_sensorTemp['C28'].value) * float(opcoes313.get())
         abracadeira = aba_sensorTemp['B29'].value
         qtd_abracadeira = float(aba_sensorTemp['C29'].value) * float(opcoes313.get())
         parafuso = aba_sensorTemp['B30'].value
         qtd_parafuso = float(aba_sensorTemp['C30'].value) * float(opcoes313.get())
      
      if opcoes311.get() == 'Sim':
         sup_inclinado_aba3 = aba_termometros['H2'].value
      elif opcoes311.get() == 'Não':
         sup_inclinado_aba3 = aba_termometros['H3'].value
      
      if opcoes314.get() == 'G 1"':
         rosca_fix_res_aba3 = aba_termometros['K2'].value
      elif opcoes314.get() == 'G 3/4"':
         rosca_fix_res_aba3 = aba_termometros['K3'].value
      elif opcoes314.get() == 'G 1.1/4"':
         rosca_fix_res_aba3 = aba_termometros['K4'].value

      # Imprimindo células para termômetro no reservatório de óleo
      aba_pa['E83'].value = cj_fix_termom_aba3
      aba_pa['H83'].value = opcoes313.get()
      aba_pa['E84'].value = conduite
      aba_pa['H84'].value = qtd_conduite * int(opcoes313.get())
      aba_pa['E85'].value = luva
      aba_pa['H85'].value = qtd_luva * int(opcoes313.get())
      aba_pa['E86'].value = conMacho
      aba_pa['H86'].value = qtd_conMacho * int(opcoes313.get())
      aba_pa['E87'].value = abracadeira
      aba_pa['H87'].value = qtd_abracadeira * int(opcoes313.get())
      aba_pa['E88'].value = parafuso
      aba_pa['H88'].value = qtd_parafuso * int(opcoes313.get())
      aba_pa['E89'].value = sup_inclinado_aba3
      aba_pa['H89'].value = opcoes313.get()
      aba_pa['E90'].value = rosca_fix_res_aba3
      aba_pa['H90'].value = opcoes313.get()

      # Imprimindo as posições da LT na planilha para termômetros
      pos_inic = int(aba_pa['D80'].value + incremento)
      incremento = int(dados_aba3['inc31'])
      celulas = 'D' + str(82)
      for m in range(10):
         aba_pa[celulas].value = pos_inic + (m * incremento)
         celulas = 'D' + str(82 + m + 1)

      # Regras da aba 4 (Mancal L.N.A) ============================================================================================================================================
      # Frame termoresistor na axial, na radial e reservatório de óleo
      if opcoes13.get() == 'Aço carbono':
         conduite_aba4 = aba_sensorTemp['B5'].value
         qtd_conduite_aba4 = float(aba_sensorTemp['C5'].value) * float(opcoes43.get())
         luva_aba4 = aba_sensorTemp['B6'].value
         qtd_luva_aba4 = float(aba_sensorTemp['C6'].value) * float(opcoes43.get())
         conMacho_aba4 = aba_sensorTemp['B7'].value
         qtd_conMacho_aba4 = float(aba_sensorTemp['C7'].value) * float(opcoes43.get())
         abracadeira_aba4 = aba_sensorTemp['B8'].value
         qtd_abracadeira_aba4 = float(aba_sensorTemp['C8'].value) * float(opcoes43.get())
         parafuso_aba4 = aba_sensorTemp['B9'].value
         qtd_parafuso_aba4 = float(aba_sensorTemp['C9'].value) * float(opcoes43.get())
      elif opcoes13.get() == 'Inox 304':
         conduite_aba4 = aba_sensorTemp['B12'].value
         qtd_conduite_aba4 = float(aba_sensorTemp['C12'].value) * float(opcoes43.get())
         luva_aba4 = aba_sensorTemp['B13'].value
         qtd_luva_aba4 = float(aba_sensorTemp['C13'].value) * float(opcoes43.get())
         conMacho_aba4 = aba_sensorTemp['B14'].value
         qtd_conMacho_aba4 = float(aba_sensorTemp['C14'].value) * float(opcoes43.get())
         abracadeira_aba4 = aba_sensorTemp['B15'].value
         qtd_abracadeira_aba4 = float(aba_sensorTemp['C15'].value) * float(opcoes43.get())
         parafuso_aba4 = aba_sensorTemp['B15'].value
         qtd_parafuso_aba4 = float(aba_sensorTemp['C15'].value) * float(opcoes43.get())
      elif opcoes13.get() == 'Inox 316':
         conduite_aba4 = aba_sensorTemp['B19'].value
         qtd_conduite_aba4 = float(aba_sensorTemp['C19'].value) * float(opcoes43.get())
         luva_aba4 = aba_sensorTemp['B20'].value
         qtd_luva_aba4 = float(aba_sensorTemp['C20'].value) * float(opcoes43.get())
         conMacho_aba4 = aba_sensorTemp['B21'].value
         qtd_conMacho_aba4 = float(aba_sensorTemp['C21'].value) * float(opcoes43.get())
         abracadeira_aba4 = aba_sensorTemp['B22'].value
         qtd_abracadeira_aba4 = float(aba_sensorTemp['C22'].value) * float(opcoes43.get())
         parafuso_aba4 = aba_sensorTemp['B23'].value
         qtd_parafuso_aba4 = float(aba_sensorTemp['C23'].value) * float(opcoes43.get())
      elif opcoes13.get() == 'Inox 316L':
         conduite_aba4 = aba_sensorTemp['B26'].value
         qtd_conduite_aba4 = float(aba_sensorTemp['C26'].value) * float(opcoes43.get())
         luva_aba4 = aba_sensorTemp['B27'].value
         qtd_luva_aba4 = float(aba_sensorTemp['C27'].value) * float(opcoes43.get())
         conMacho_aba4 = aba_sensorTemp['B28'].value
         qtd_conMacho_aba4 = float(aba_sensorTemp['C28'].value) * float(opcoes43.get())
         abracadeira_aba4 = aba_sensorTemp['B29'].value
         qtd_abracadeira_aba4 = float(aba_sensorTemp['C29'].value) * float(opcoes43.get())
         parafuso_aba4 = aba_sensorTemp['B30'].value
         qtd_parafuso_aba4 = float(aba_sensorTemp['C30'].value) * float(opcoes43.get())

      # Frame termoresistor na radial
      if opcoes41.get() == 'Isolado':
         adap_isolante_aba4 = aba_sensorTemp['H2'].value
         qtd_adap_isolante_aba4 = aba_sensorTemp['C4'].value * float(opcoes44.get())
      
      if opcoes13.get() == 'Aço carbono':
         conduite_aba4 = aba_sensorTemp['B5'].value
         qtd_conduite_aba4 = float(aba_sensorTemp['C5'].value) * float(opcoes44.get())
         luva_aba4 = aba_sensorTemp['B6'].value
         qtd_luva_aba4 = float(aba_sensorTemp['C6'].value) * float(opcoes44.get())
         conMacho_aba4 = aba_sensorTemp['B7'].value
         qtd_conMacho_aba4 = float(aba_sensorTemp['C7'].value) * float(opcoes44.get())
         abracadeira_aba4 = aba_sensorTemp['B8'].value
         qtd_abracadeira_aba4 = float(aba_sensorTemp['C8'].value) * float(opcoes44.get())
         parafuso_aba4 = aba_sensorTemp['B9'].value
         qtd_parafuso_aba4 = float(aba_sensorTemp['C9'].value) * float(opcoes44.get())
      elif opcoes13.get() == 'Inox 304':
         conduite_aba4 = aba_sensorTemp['B12'].value
         qtd_conduite_aba4 = float(aba_sensorTemp['C12'].value) * float(opcoes44.get())
         luva_aba4 = aba_sensorTemp['B13'].value
         qtd_luva_aba4 = float(aba_sensorTemp['C13'].value) * float(opcoes44.get())
         conMacho_aba4 = aba_sensorTemp['B14'].value
         qtd_conMacho_aba4 = float(aba_sensorTemp['C14'].value) * float(opcoes44.get())
         abracadeira_aba4 = aba_sensorTemp['B15'].value
         qtd_abracadeira_aba4 = float(aba_sensorTemp['C15'].value) * float(opcoes44.get())
         parafuso_aba4 = aba_sensorTemp['B15'].value
         qtd_parafuso_aba4 = float(aba_sensorTemp['C15'].value) * float(opcoes44.get())
      elif opcoes13.get() == 'Inox 316':
         conduite_aba4 = aba_sensorTemp['B19'].value
         qtd_conduite_aba4 = float(aba_sensorTemp['C19'].value) * float(opcoes44.get())
         luva_aba4 = aba_sensorTemp['B20'].value
         qtd_luva_aba4 = float(aba_sensorTemp['C20'].value) * float(opcoes44.get())
         conMacho_aba4 = aba_sensorTemp['B21'].value
         qtd_conMacho_aba4 = float(aba_sensorTemp['C21'].value) * float(opcoes44.get())
         abracadeira_aba4 = aba_sensorTemp['B22'].value
         qtd_abracadeira_aba4 = float(aba_sensorTemp['C22'].value) * float(opcoes44.get())
         parafuso_aba4 = aba_sensorTemp['B23'].value
         qtd_parafuso_aba4 = float(aba_sensorTemp['C23'].value) * float(opcoes44.get())
      elif opcoes13.get() == 'Inox 316L':
         conduite_aba4 = aba_sensorTemp['B26'].value
         qtd_conduite_aba4 = float(aba_sensorTemp['C26'].value) * float(opcoes44.get())
         luva_aba4 = aba_sensorTemp['B27'].value
         qtd_luva_aba4 = float(aba_sensorTemp['C27'].value) * float(opcoes44.get())
         conMacho_aba4 = aba_sensorTemp['B28'].value
         qtd_conMacho_aba4 = float(aba_sensorTemp['C28'].value) * float(opcoes44.get())
         abracadeira_aba4 = aba_sensorTemp['B29'].value
         qtd_abracadeira_aba4 = float(aba_sensorTemp['C29'].value) * float(opcoes44.get())
         parafuso_aba4 = aba_sensorTemp['B30'].value
         qtd_parafuso_aba4 = float(aba_sensorTemp['C30'].value) * float(opcoes44.get())

      # Frame termoresistor no reservatório de óleo
      if opcoes13.get() == 'Aço carbono':
         conduite_aba4 = aba_sensorTemp['B5'].value
         qtd_conduite_aba4 = float(aba_sensorTemp['C5'].value) * float(opcoes45.get())
         luva_aba4 = aba_sensorTemp['B6'].value
         qtd_luva_aba4 = float(aba_sensorTemp['C6'].value) * float(opcoes45.get())
         conMacho_aba4 = aba_sensorTemp['B7'].value
         qtd_conMacho_aba4 = float(aba_sensorTemp['C7'].value) * float(opcoes45.get())
         abracadeira_aba4 = aba_sensorTemp['B8'].value
         qtd_abracadeira_aba4 = float(aba_sensorTemp['C8'].value) * float(opcoes45.get())
         parafuso_aba4 = aba_sensorTemp['B9'].value
         qtd_parafuso_aba4 = float(aba_sensorTemp['C9'].value) * float(opcoes45.get())
      elif opcoes13.get() == 'Inox 304':
         conduite_aba4 = aba_sensorTemp['B12'].value
         qtd_conduite_aba4 = float(aba_sensorTemp['C12'].value) * float(opcoes45.get())
         luva_aba4 = aba_sensorTemp['B13'].value
         qtd_luva_aba4 = float(aba_sensorTemp['C13'].value) * float(opcoes45.get())
         conMacho_aba4 = aba_sensorTemp['B14'].value
         qtd_conMacho_aba4 = float(aba_sensorTemp['C14'].value) * float(opcoes45.get())
         abracadeira_aba4 = aba_sensorTemp['B15'].value
         qtd_abracadeira_aba4 = float(aba_sensorTemp['C15'].value) * float(opcoes45.get())
         parafuso_aba4 = aba_sensorTemp['B15'].value
         qtd_parafuso_aba4 = float(aba_sensorTemp['C15'].value) * float(opcoes45.get())
      elif opcoes13.get() == 'Inox 316':
         conduite_aba4 = aba_sensorTemp['B19'].value
         qtd_conduite_aba4 = float(aba_sensorTemp['C19'].value) * float(opcoes45.get())
         luva_aba4 = aba_sensorTemp['B20'].value
         qtd_luva_aba4 = float(aba_sensorTemp['C20'].value) * float(opcoes45.get())
         conMacho_aba4 = aba_sensorTemp['B21'].value
         qtd_conMacho_aba4 = float(aba_sensorTemp['C21'].value) * float(opcoes45.get())
         abracadeira_aba4 = aba_sensorTemp['B22'].value
         qtd_abracadeira_aba4 = float(aba_sensorTemp['C22'].value) * float(opcoes45.get())
         parafuso_aba4 = aba_sensorTemp['B23'].value
         qtd_parafuso_aba4 = float(aba_sensorTemp['C23'].value) * float(opcoes45.get())
      elif opcoes13.get() == 'Inox 316L':
         conduite_aba4 = aba_sensorTemp['B26'].value
         qtd_conduite_aba4 = float(aba_sensorTemp['C26'].value) * float(opcoes45.get())
         luva_aba4 = aba_sensorTemp['B27'].value
         qtd_luva_aba4 = float(aba_sensorTemp['C27'].value) * float(opcoes45.get())
         conMacho_aba4 = aba_sensorTemp['B28'].value
         qtd_conMacho_aba4 = float(aba_sensorTemp['C28'].value) * float(opcoes45.get())
         abracadeira_aba4 = aba_sensorTemp['B29'].value
         qtd_abracadeira_aba4 = float(aba_sensorTemp['C29'].value) * float(opcoes45.get())
         parafuso_aba4 = aba_sensorTemp['B30'].value
         qtd_parafuso_aba4 = float(aba_sensorTemp['C30'].value) * float(opcoes45.get())

      if opcoes46.get() == 'G 3/4"':
         rosca_fix_res_aba4 = aba_termometros['K2'].value
      elif opcoes46.get() == 'G 1.1/4"':
         rosca_fix_res_aba4 = aba_termometros['K3'].value
      elif opcoes46.get() == 'G 1"':
         rosca_fix_res_aba4 = aba_termometros['K4'].value
      
      # Imprimindo os dados de termorresistor para instalação axial
      aba_pa['E147'].value = conduite_aba4
      aba_pa['H147'].value = qtd_conduite_aba4
      aba_pa['E148'].value = luva_aba4
      aba_pa['H148'].value = qtd_luva_aba4
      aba_pa['E149'].value = conMacho_aba4
      aba_pa['H149'].value = qtd_conMacho_aba4
      aba_pa['E150'].value = abracadeira_aba4
      aba_pa['H150'].value = qtd_abracadeira_aba4
      aba_pa['E151'].value = parafuso_aba4
      aba_pa['H151'].value = qtd_parafuso_aba4

      pos_inic = int(dados_aba4['lt41'])
      incremento = int(dados_aba4['inc41'])
      celulas = 'D' + str(145)
      for m in range(7):
         aba_pa[celulas].value = pos_inic + (m * incremento)
         celulas = 'D' + str(145 + m + 1)

      # Imprimindo os dados de termorresistor para instalação radial
      aba_pa['E155'].value = adap_isolante_aba4
      aba_pa['H155'].value = qtd_adap_isolante_aba4
      aba_pa['E156'].value = conduite_aba4
      aba_pa['H156'].value = qtd_conduite_aba4
      aba_pa['E157'].value = luva_aba4
      aba_pa['H157'].value = qtd_luva_aba4
      aba_pa['E158'].value = conMacho_aba4
      aba_pa['H158'].value = qtd_conMacho_aba4
      aba_pa['E159'].value = abracadeira_aba4
      aba_pa['H159'].value = qtd_abracadeira_aba4
      aba_pa['E160'].value = parafuso_aba4
      aba_pa['H160'].value = qtd_parafuso_aba4

      pos_inic = int(aba_pa['D151'].value + incremento)
      incremento = int(dados_aba4['inc41'])
      celulas = 'D' + str(153)
      for m in range(8):
         aba_pa[celulas].value = pos_inic + (m * incremento)
         celulas = 'D' + str(153 + m + 1)

      # Imprimindo os dados de termorresistor para instalação no rservatório de óleo
      aba_pa['E164'].value = conduite_aba4
      aba_pa['H164'].value = qtd_conduite_aba4
      aba_pa['E165'].value = luva_aba4
      aba_pa['H165'].value = qtd_luva_aba4
      aba_pa['E166'].value = conMacho_aba4
      aba_pa['H166'].value = qtd_conMacho_aba4
      aba_pa['E167'].value = abracadeira_aba4
      aba_pa['H167'].value = qtd_abracadeira_aba4
      aba_pa['E168'].value = parafuso_aba4
      aba_pa['H168'].value = qtd_parafuso_aba4
      aba_pa['E169'].value = rosca_fix_res_aba4
      aba_pa['H169'].value = opcoes45.get()

      pos_inic = int(aba_pa['D160'].value + incremento)
      incremento = int(dados_aba4['inc41'])
      celulas = 'D' + str(162)
      for m in range(9):
         aba_pa[celulas].value = pos_inic + (m * incremento)
         celulas = 'D' + str(162 + m + 1)

      # Frame termômetro na radial
      if opcoes410.get() == 'Sem contato' and opcoes13.get() == 'Aço carbono' and opcoes41.get() == 'Isolado':
         cj_fix_termom_aba4 = aba_termometros['D3'].value
      elif opcoes410.get() == 'Sem contato' and opcoes13.get() == 'Inox 304' and opcoes41.get() == 'Isolado':
         cj_fix_termom_aba4 = aba_termometros['D4'].value
      elif opcoes410.get() == 'Sem contato' and opcoes13.get() == 'Inox 316' and opcoes41.get() == 'Isolado':
         cj_fix_termom_aba4 = aba_termometros['D5'].value
      elif opcoes410.get() == 'Sem contato' and opcoes13.get() == 'Inox 316L' and opcoes41.get() == 'Isolado':
         cj_fix_termom_aba4 = aba_termometros['D6'].value
      elif opcoes410.get() == 'Sem contato' and opcoes13.get() == 'Aço carbono' and opcoes41.get() == 'Não isolado':
         cj_fix_termom_aba4 = aba_termometros['D7'].value
      elif opcoes410.get() == 'Sem contato' and opcoes13.get() == 'Inox 304' and opcoes41.get() == 'Não isolado':
         cj_fix_termom_aba4 = aba_termometros['D8'].value
      elif opcoes410.get() == 'Sem contato' and opcoes13.get() == 'Inox 316' and opcoes41.get() == 'Não isolado':
         cj_fix_termom_aba4 = aba_termometros['D9'].value
      elif opcoes410.get() == 'Sem contato' and opcoes13.get() == 'Inox 316L' and opcoes41.get() == 'Não isolado':
         cj_fix_termom_aba4 = aba_termometros['D10'].value
      
      if opcoes410.get() == 'Com contato' and opcoes13.get() == 'Aço carbono' and opcoes41.get() == 'Isolado':
         cj_fix_termom_aba4 = aba_termometros['D3'].value
      elif opcoes410.get() == 'Com contato' and opcoes13.get() == 'Inox 304' and opcoes41.get() == 'Isolado':
         cj_fix_termom_aba4 = aba_termometros['D4'].value
      elif opcoes410.get() == 'Com contato' and opcoes13.get() == 'Inox 316' and opcoes41.get() == 'Isolado':
         cj_fix_termom_aba4 = aba_termometros['D5'].value
      elif opcoes410.get() == 'Com contato' and opcoes13.get() == 'Inox 316L' and opcoes41.get() == 'Isolado':
         cj_fix_termom_aba4 = aba_termometros['D6'].value
      elif opcoes410.get() == 'Com contato' and opcoes13.get() == 'Aço carbono' and opcoes41.get() == 'Não isolado':
         cj_fix_termom_aba4 = aba_termometros['D7'].value
      elif opcoes410.get() == 'Com contato' and opcoes13.get() == 'Inox 304' and opcoes41.get() == 'Não isolado':
         cj_fix_termom_aba4 = aba_termometros['D8'].value
      elif opcoes410.get() == 'Com contato' and opcoes13.get() == 'Inox 316' and opcoes41.get() == 'Não isolado':
         cj_fix_termom_aba4 = aba_termometros['D9'].value
      elif opcoes410.get() == 'Com contato' and opcoes13.get() == 'Inox 316L' and opcoes41.get() == 'Não isolado':
         cj_fix_termom_aba4 = aba_termometros['D10'].value
      
      if opcoes410.get() == 'Com contato' and opcoes13.get() == 'Aço carbono':
         conduite_aba4 = aba_sensorTemp['B5'].value
         qtd_conduite_aba4 = float(aba_sensorTemp['C5'].value) * float(opcoes412.get())
         luva_aba4 = aba_sensorTemp['B6'].value
         qtd_luva_aba4 = float(aba_sensorTemp['C6'].value) * float(opcoes412.get())
         conMacho_aba4 = aba_sensorTemp['B7'].value
         qtd_conMacho_aba4 = float(aba_sensorTemp['C7'].value) * float(opcoes412.get())
         abracadeira_aba4 = aba_sensorTemp['B8'].value
         qtd_abracadeira_aba4 = float(aba_sensorTemp['C8'].value) * float(opcoes412.get())
         parafuso_aba4 = aba_sensorTemp['B9'].value
         qtd_parafuso_aba4 = float(aba_sensorTemp['C9'].value) * float(opcoes412.get())
      elif opcoes410.get() == 'Com contato' and opcoes13.get() == 'Inox 304':
         conduite_aba4 = aba_sensorTemp['B12'].value
         qtd_conduite_aba4 = float(aba_sensorTemp['C12'].value) * float(opcoes412.get())
         luva_aba4 = aba_sensorTemp['B13'].value
         qtd_luva_aba4 = float(aba_sensorTemp['C13'].value) * float(opcoes412.get())
         conMacho_aba4 = aba_sensorTemp['B14'].value
         qtd_conMacho_aba4 = float(aba_sensorTemp['C14'].value) * float(opcoes412.get())
         abracadeira_aba4 = aba_sensorTemp['B15'].value
         qtd_abracadeira_aba4 = float(aba_sensorTemp['C15'].value) * float(opcoes412.get())
         parafuso_aba4 = aba_sensorTemp['B16'].value
         qtd_parafuso_aba4 = float(aba_sensorTemp['C16'].value) * float(opcoes412.get())
      elif opcoes410.get() == 'Com contato' and opcoes13.get() == 'Inox 316':
         conduite_aba4 = aba_sensorTemp['B19'].value
         qtd_conduite_aba4 = float(aba_sensorTemp['C19'].value) * float(opcoes412.get())
         luva_aba4 = aba_sensorTemp['B20'].value
         qtd_luva_aba4 = float(aba_sensorTemp['C20'].value) * float(opcoes412.get())
         conMacho_aba4 = aba_sensorTemp['B21'].value
         qtd_conMacho_aba4 = float(aba_sensorTemp['C21'].value) * float(opcoes412.get())
         abracadeira_aba4 = aba_sensorTemp['B22'].value
         qtd_abracadeira_aba4 = float(aba_sensorTemp['C22'].value) * float(opcoes412.get())
         parafuso_aba4 = aba_sensorTemp['B23'].value
         qtd_parafuso_aba4 = float(aba_sensorTemp['C23'].value) * float(opcoes412.get())
      elif opcoes410.get() == 'Com contato' and opcoes13.get() == 'Inox 316L':
         conduite_aba4 = aba_sensorTemp['B26'].value
         qtd_conduite_aba4 = float(aba_sensorTemp['C26'].value) * float(opcoes412.get())
         luva_aba4 = aba_sensorTemp['B27'].value
         qtd_luva_aba4 = float(aba_sensorTemp['C27'].value) * float(opcoes412.get())
         conMacho_aba4 = aba_sensorTemp['B28'].value
         qtd_conMacho_aba4 = float(aba_sensorTemp['C28'].value) * float(opcoes412.get())
         abracadeira_aba4 = aba_sensorTemp['B29'].value
         qtd_abracadeira_aba4 = float(aba_sensorTemp['C29'].value) * float(opcoes412.get())
         parafuso_aba4 = aba_sensorTemp['B30'].value
         qtd_parafuso_aba4 = float(aba_sensorTemp['C30'].value) * float(opcoes412.get())

      if opcoes411.get() == 'Sim':
         sup_inclinado_aba4 = aba_termometros['H2'].value
      elif opcoes411.get() == 'Não':
         sup_inclinado_aba4 = aba_termometros['H3'].value

      # Imprimindo os dados da aba 3 na planilha para termômetro na radial
      aba_pa['E174'].value = cj_fix_termom_aba4
      aba_pa['H174'].value = opcoes412.get()
      aba_pa['E175'].value = conduite_aba4
      aba_pa['H175'].value = qtd_conduite_aba4 * int(opcoes412.get())
      aba_pa['E176'].value = luva_aba4
      aba_pa['H176'].value = qtd_luva_aba4 * int(opcoes412.get())
      aba_pa['E177'].value = conMacho_aba4
      aba_pa['H177'].value = qtd_conMacho_aba4 * int(opcoes412.get())
      aba_pa['E178'].value = abracadeira_aba4
      aba_pa['H178'].value = qtd_abracadeira_aba4 * int(opcoes412.get())
      aba_pa['E179'].value = parafuso_aba4
      aba_pa['H179'].value = qtd_parafuso_aba4 * int(opcoes412.get())
      aba_pa['E180'].value = sup_inclinado_aba4
      aba_pa['H180'].value = opcoes412.get()

      # Imprimindo as posições da LT na planilha para termômetros
      pos_inic = int(aba_pa['D170'].value + incremento)
      incremento = int(dados_aba4['inc41'])
      celulas = 'D' + str(173)
      for m in range(8):
         aba_pa[celulas].value = pos_inic + (m * incremento)
         celulas = 'D' + str(173 + m + 1)
         
      # Frame termômetro no reservatório de óleo
      if opcoes410.get() == 'Sem contato' and opcoes13.get() == 'Aço carbono' and opcoes41.get() == 'Isolado':
         cj_fix_termom_aba4 = aba_termometros['D3'].value
      elif opcoes410.get() == 'Sem contato' and opcoes13.get() == 'Inox 304' and opcoes41.get() == 'Isolado':
         cj_fix_termom_aba4 = aba_termometros['D4'].value
      elif opcoes410.get() == 'Sem contato' and opcoes13.get() == 'Inox 316' and opcoes41.get() == 'Isolado':
         cj_fix_termom_aba4 = aba_termometros['D5'].value
      elif opcoes410.get() == 'Sem contato' and opcoes13.get() == 'Inox 316L' and opcoes41.get() == 'Isolado':
         cj_fix_termom_aba4 = aba_termometros['D6'].value
      elif opcoes410.get() == 'Sem contato' and opcoes13.get() == 'Aço carbono' and opcoes41.get() == 'Não isolado':
         cj_fix_termom_aba4 = aba_termometros['D7'].value
      elif opcoes410.get() == 'Sem contato' and opcoes13.get() == 'Inox 304' and opcoes41.get() == 'Não isolado':
         cj_fix_termom_aba4 = aba_termometros['D8'].value
      elif opcoes410.get() == 'Sem contato' and opcoes13.get() == 'Inox 316' and opcoes41.get() == 'Não isolado':
         cj_fix_termom_aba4 = aba_termometros['D9'].value
      elif opcoes410.get() == 'Sem contato' and opcoes13.get() == 'Inox 316L' and opcoes41.get() == 'Não isolado':
         cj_fix_termom_aba4 = aba_termometros['D10'].value
      elif opcoes410.get() == 'Com contato' and opcoes13.get() == 'Aço carbono' and opcoes41.get() == 'Isolado':
         cj_fix_termom_aba4 = aba_termometros['D11'].value
      elif opcoes410.get() == 'Com contato' and opcoes13.get() == 'Inox 304' and opcoes41.get() == 'Isolado':
         cj_fix_termom_aba4 = aba_termometros['D12'].value
      elif opcoes410.get() == 'Com contato' and opcoes13.get() == 'Inox 316' and opcoes41.get() == 'Isolado':
         cj_fix_termom_aba4 = aba_termometros['D13'].value
      elif opcoes410.get() == 'Com contato' and opcoes13.get() == 'Inox 316L' and opcoes41.get() == 'Isolado':
         cj_fix_termom_aba4 = aba_termometros['D14'].value
      elif opcoes410.get() == 'Com contato' and opcoes13.get() == 'Aço carbono' and opcoes41.get() == 'Não isolado':
         cj_fix_termom_aba4 = aba_termometros['D15'].value
      elif opcoes410.get() == 'Com contato' and opcoes13.get() == 'Inox 304' and opcoes41.get() == 'Não isolado':
         cj_fix_termom_aba4 = aba_termometros['D16'].value
      elif opcoes410.get() == 'Com contato' and opcoes13.get() == 'Inox 316' and opcoes41.get() == 'Não isolado':
         cj_fix_termom_aba4 = aba_termometros['D17'].value
      elif opcoes410.get() == 'Com contato' and opcoes13.get() == 'Inox 316L' and opcoes41.get() == 'Não isolado':
         cj_fix_termom_aba4 = aba_termometros['D18'].value
      
      if 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Aço carbono':
         conduite = aba_sensorTemp['B5'].value
         qtd_conduite = float(aba_sensorTemp['C5'].value) * float(opcoes413.get())
         luva = aba_sensorTemp['B6'].value
         qtd_luva = float(aba_sensorTemp['C6'].value) * float(opcoes413.get())
         conMacho = aba_sensorTemp['B7'].value
         qtd_conMacho = float(aba_sensorTemp['C7'].value) * float(opcoes413.get())
         abracadeira = aba_sensorTemp['B8'].value
         qtd_abracadeira = float(aba_sensorTemp['C8'].value) * float(opcoes413.get())
         parafuso = aba_sensorTemp['B9'].value
         qtd_parafuso = float(aba_sensorTemp['C9'].value) * float(opcoes413.get())
      elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 304':
         conduite = aba_sensorTemp['B12'].value
         qtd_conduite = float(aba_sensorTemp['C12'].value) * float(opcoes413.get())
         luva = aba_sensorTemp['B13'].value
         qtd_luva = float(aba_sensorTemp['C13'].value) * float(opcoes413.get())
         conMacho = aba_sensorTemp['B14'].value
         qtd_conMacho = float(aba_sensorTemp['C14'].value) * float(opcoes413.get())
         abracadeira = aba_sensorTemp['B15'].value
         qtd_abracadeira = float(aba_sensorTemp['C15'].value) * float(opcoes413.get())
         parafuso = aba_sensorTemp['B16'].value
         qtd_parafuso = float(aba_sensorTemp['C16'].value) * float(opcoes413.get())
      elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316':
         conduite = aba_sensorTemp['B19'].value
         qtd_conduite = float(aba_sensorTemp['C19'].value) * float(opcoes413.get())
         luva = aba_sensorTemp['B20'].value
         qtd_luva = float(aba_sensorTemp['C20'].value) * float(opcoes413.get())
         conMacho = aba_sensorTemp['B21'].value
         qtd_conMacho = float(aba_sensorTemp['C21'].value) * float(opcoes413.get())
         abracadeira = aba_sensorTemp['B22'].value
         qtd_abracadeira = float(aba_sensorTemp['C22'].value) * float(opcoes413.get())
         parafuso = aba_sensorTemp['B23'].value
         qtd_parafuso = float(aba_sensorTemp['C23'].value) * float(opcoes413.get())
      elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316L':
         conduite = aba_sensorTemp['B26'].value
         qtd_conduite = float(aba_sensorTemp['C26'].value) * float(opcoes413.get())
         luva = aba_sensorTemp['B27'].value
         qtd_luva = float(aba_sensorTemp['C27'].value) * float(opcoes413.get())
         conMacho = aba_sensorTemp['B28'].value
         qtd_conMacho = float(aba_sensorTemp['C28'].value) * float(opcoes413.get())
         abracadeira = aba_sensorTemp['B29'].value
         qtd_abracadeira = float(aba_sensorTemp['C29'].value) * float(opcoes413.get())
         parafuso = aba_sensorTemp['B30'].value
         qtd_parafuso = float(aba_sensorTemp['C30'].value) * float(opcoes413.get())
      
      if opcoes411.get() == 'Sim':
         sup_inclinado_aba4 = aba_termometros['H2'].value
      elif opcoes411.get() == 'Não':
         sup_inclinado_aba4 = aba_termometros['H3'].value
      
      if opcoes414.get() == 'G 1"':
         rosca_fix_res_aba4 = aba_termometros['K2'].value
      elif opcoes414.get() == 'G 3/4"':
         rosca_fix_res_aba4 = aba_termometros['K3'].value
      elif opcoes414.get() == 'G 1.1/4"':
         rosca_fix_res_aba4 = aba_termometros['K4'].value

      # Imprimindo células para termômetro no reservatório de óleo
      aba_pa['E183'].value = cj_fix_termom_aba4
      aba_pa['H183'].value = opcoes413.get()
      aba_pa['E184'].value = conduite
      aba_pa['H184'].value = qtd_conduite * int(opcoes413.get())
      aba_pa['E185'].value = luva
      aba_pa['H185'].value = qtd_luva * int(opcoes413.get())
      aba_pa['E186'].value = conMacho
      aba_pa['H186'].value = qtd_conMacho * int(opcoes413.get())
      aba_pa['E187'].value = abracadeira
      aba_pa['H187'].value = qtd_abracadeira * int(opcoes413.get())
      aba_pa['E188'].value = parafuso
      aba_pa['H188'].value = qtd_parafuso * int(opcoes413.get())
      aba_pa['E189'].value = sup_inclinado_aba4
      aba_pa['H189'].value = opcoes413.get()
      aba_pa['E190'].value = rosca_fix_res_aba4
      aba_pa['H190'].value = opcoes413.get()

      # Imprimindo as posições da LT na planilha para termômetros
      pos_inic = int(aba_pa['D180'].value + incremento)
      incremento = int(dados_aba4['inc41'])
      celulas = 'D' + str(182)
      for m in range(10):
         aba_pa[celulas].value = pos_inic + (m * incremento)
         celulas = 'D' + str(182 + m + 1)

      # Regras da aba 5 (Cx. Acess. Pequena) ======================================================================================================================================
      s = 0
      for n in range(4):
         if (n + 1) == 1:
            s = 4
         elif (n + 1) == 2:
            s = 11
         elif (n + 1) == 3:
            s = 18
         elif (n + 1) == 4:
            s = 25

         # Conjunto tampa caixa de ligação
         if 'opcoes52' in dados_aba5 and dados_aba5['opcoes52'] == 'Ferro fundido':
            a = aba_cx_peq['B3'].value
         elif 'opcoes52' in dados_aba5 and dados_aba5['opcoes52'] == 'Inox 316':
            a = aba_cx_peq['B4'].value
         elif 'opcoes52' in dados_aba5 and dados_aba5['opcoes52'] == 'Inox 304':
            a = aba_cx_peq['B5'].value

         # Placa de fixação de trilho
         if ('opcoes5' + str(s + 5)) in dados_aba5 and ('opcoes5' + str(s + 6)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 5)] == 'Não' and dados_aba5['opcoes5' + str(s + 6)] == 'Área Segura':
            b = aba_cx_peq['E3'].value
         elif ('opcoes5' + str(s + 5)) in dados_aba5 and ('opcoes5' + str(s + 6)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 5)] == 'Sim' and dados_aba5['opcoes5' + str(s + 6)] == 'Área Segura':
            b = aba_cx_peq['E4'].value
         elif ('opcoes5' + str(s + 5)) in dados_aba5 and ('opcoes5' + str(s + 6)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 5)] == 'Sim' and dados_aba5['opcoes5' + str(s + 6)] == 'Ex-e':
            b = aba_cx_peq['E5'].value
         elif ('opcoes5' + str(s + 5)) in dados_aba5 and ('opcoes5' + str(s + 6)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 5)] == 'Não' and dados_aba5['opcoes5' + str(s + 6)] == 'Ex-e':
            b = aba_cx_peq['E5'].value
         elif ('opcoes5' + str(s + 5)) in dados_aba5 and ('opcoes5' + str(s + 6)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 5)] == 'Não' and dados_aba5['opcoes5' + str(s + 6)] == 'Ex-p':
            b = aba_cx_peq['E6'].value

         # Conj. fixação caixa de acessórios
         if dados_aba1['opcoes13'] == 'Aço carbono':
            c = aba_cx_peq['H3'].value
         elif dados_aba1['opcoes13'] == 'Inox 304':
            c = aba_cx_peq['H4'].value
         elif dados_aba1['opcoes13'] == 'Inox 316':
            c = aba_cx_peq['H5'].value
         elif dados_aba1['opcoes13'] == 'Inox 316L':
            c = aba_cx_peq['H6'].value

         # Tampão roscado
         if ('opcoes5' + str(s + 1)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'G 1"':
            d = aba_cx_peq['K3'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'NPT 1"':
            d = aba_cx_peq['K4'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'M20x1,5':
            d = aba_cx_peq['K5'].value
         else:
            d = 'N/A'

         # Parafuso fixação M6x20
         if 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Aço carbono':
            e = aba_cx_peq['N3'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 304':
            e = aba_cx_peq['N4'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316':
            e = aba_cx_peq['N5'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316L':
            e = aba_cx_peq['N6'].value

         # Parafuso fixação triho
         if 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Aço carbono':
            f = aba_cx_peq['Q3'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 304':
            f = aba_cx_peq['Q4'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316':
            f = aba_cx_peq['Q5'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316L':
            f = aba_cx_peq['Q6'].value

         # Ilho de borracha
         if ('opcoes5' + str(s + 6)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 6)] == 'Área Segura':
            g = aba_cx_peq['T3'].value
         elif ('opcoes5' + str(s + 5)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 5)] == 'Sim':
            g = aba_cx_peq['T4'].value
         elif ('opcoes5' + str(s + 6)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 6)] == 'Ex-e':
            g = aba_cx_peq['T4'].value
         elif ('opcoes5' + str(s + 6)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 6)] == 'Ex-p':
            g = aba_cx_peq['T4'].value

         # Caixa de ligação fundida
         if ('opcoes5' + str(s)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes52') in dados_aba5 and dados_aba5['opcoes52'] == 'Inox 304' and dados_aba5['opcoes5' + str(s)] == 2 and ('opcoes5' + str(s + 1)) == 'Sem rosca':
            h = aba_cx_peq['Z3'].value
         elif ('opcoes5' + str(s)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes52') in dados_aba5 and dados_aba5['opcoes52'] == 'Inox 316' and dados_aba5['opcoes5' + str(s)] == 2 and ('opcoes5' + str(s + 1)) == 'Sem rosca':
            h = aba_cx_peq['Z4'].value
         elif ('opcoes5' + str(s)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes52') in dados_aba5 and dados_aba5['opcoes52'] == 'Ferro fundido' and dados_aba5['opcoes5' + str(s)] == 2 and ('opcoes5' + str(s + 1)) == 'Sem rosca':
            h = aba_cx_peq['Z5'].value
         elif ('opcoes5' + str(s)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and dados_aba5['opcoes52'] == 'Inox 304' and dados_aba5['opcoes5' + str(s)] == 1 and dados_aba5['opcoes5' + str(s + 1)] == 'G 1"':
            h = aba_cx_peq['Z6'].value
         elif ('opcoes5' + str(s)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and dados_aba5['opcoes52'] == 'Inox 316' and dados_aba5['opcoes5' + str(s)] == 1 and dados_aba5['opcoes5' + str(s + 1)] == 'G 1"':
            h = aba_cx_peq['Z7'].value
         elif ('opcoes5' + str(s)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and dados_aba5['opcoes52'] == 'Ferro fundido' and dados_aba5['opcoes5' + str(s)] == 1 and dados_aba5['opcoes5' + str(s + 1)] == 'G 1"':
            h = aba_cx_peq['Z8'].value
         elif ('opcoes5' + str(s)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and dados_aba5['opcoes52'] == 'Inox 304' and dados_aba5['opcoes5' + str(s)] == 1 and dados_aba5['opcoes5' + str(s + 1)] == 'NPT 1"':
            h = aba_cx_peq['Z9'].value
         elif ('opcoes5' + str(s)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and dados_aba5['opcoes52'] == 'Inox 316' and dados_aba5['opcoes5' + str(s)] == 1 and dados_aba5['opcoes5' + str(s + 1)] == 'NPT 1"':
            h = aba_cx_peq['Z10'].value
         elif ('opcoes5' + str(s)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and dados_aba5['opcoes52'] == 'Ferro fundido' and dados_aba5['opcoes5' + str(s)] == 1 and dados_aba5['opcoes5' + str(s + 1)] == 'NPT 1"':
            h = aba_cx_peq['Z11'].value
         elif ('opcoes5' + str(s)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and dados_aba5['opcoes52'] == 'Inox 304' and dados_aba5['opcoes5' + str(s)] == 1 and dados_aba5['opcoes5' + str(s + 1)] == 'M20x1,5':
            h = aba_cx_peq['Z12'].value
         elif ('opcoes5' + str(s)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and dados_aba5['opcoes52'] == 'Inox 316' and dados_aba5['opcoes5' + str(s)] == 1 and dados_aba5['opcoes5' + str(s + 1)] == 'M20x1,5':
            h = aba_cx_peq['Z13'].value
         elif ('opcoes5' + str(s)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 1)) in dados_aba5 and dados_aba5['opcoes52'] == 'Ferro fundido' and dados_aba5['opcoes5' + str(s)] == 1 and dados_aba5['opcoes5' + str(s + 1)] == 'M20x1,5':
            h = aba_cx_peq['Z14'].value

         # Conj. placa cega s/ parafusos
         if ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'Sem rosca' and dados_aba5['opcoes5' + str(s + 3)] == 'Ferro fundido':
            i = aba_cx_peq['AD3'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'Sem rosca' and dados_aba5['opcoes5' + str(s + 3)] == 'Alumínio (AlSi6Cu4)':
            i = aba_cx_peq['AD4'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'Sem rosca' and dados_aba5['opcoes5' + str(s + 3)] == 'Alumínio (AlSi10Mg)':
            i = aba_cx_peq['AD5'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'Sem rosca' and dados_aba5['opcoes5' + str(s + 3)] == 'Inox 316':
            i = aba_cx_peq['AD6'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'Sem rosca' and dados_aba5['opcoes5' + str(s + 3)] == 'Inox 304':
            i = aba_cx_peq['AD7'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'G 1"' and dados_aba5['opcoes5' + str(s + 3)] == 'Ferro fundido':
            i = aba_cx_peq['AD8'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'NPT 1"' and dados_aba5['opcoes5' + str(s + 3)] == 'Ferro fundido':
            i = aba_cx_peq['AD9'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'M20x1,5' and dados_aba5['opcoes5' + str(s + 3)] == 'Ferro fundido':
            i = aba_cx_peq['AD10'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'G 1"' and dados_aba5['opcoes5' + str(s + 3)] == 'Alumínio (AlSi6Cu4)':
            i = aba_cx_peq['AD11'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'NPT 1"' and dados_aba5['opcoes5' + str(s + 3)] == 'Alumínio (AlSi6Cu4)':
            i = aba_cx_peq['AD12'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'M20x1,5' and dados_aba5['opcoes5' + str(s + 3)] == 'Alumínio (AlSi6Cu4)':
            i = aba_cx_peq['AD13'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'G 1"' and dados_aba5['opcoes5' + str(s + 3)] == 'Alumínio (AlSi10Mg)':
            i = aba_cx_peq['AD14'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'NPT 1"' and dados_aba5['opcoes5' + str(s + 3)] == 'Alumínio (AlSi10Mg)':
            i = aba_cx_peq['AD15'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'M20x1,5' and dados_aba5['opcoes5' + str(s + 3)] == 'Alumínio (AlSi10Mg)':
            i = aba_cx_peq['AD16'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'G 1"' and dados_aba5['opcoes5' + str(s + 3)] == 'Inox 304':
            i = aba_cx_peq['AD17'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'NPT 1"' and dados_aba5['opcoes5' + str(s + 3)] == 'Inox 304':
            i = aba_cx_peq['AD18'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'M20x1,5' and dados_aba5['opcoes5' + str(s + 3)] == 'Inox 304':
            i = aba_cx_peq['AD19'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'G 1"' and dados_aba5['opcoes5' + str(s + 3)] == 'Inox 316':
            i = aba_cx_peq['AD20'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'NPT 1"' and dados_aba5['opcoes5' + str(s + 3)] == 'Inox 316':
            i = aba_cx_peq['AD21'].value
         elif ('opcoes5' + str(s + 1)) in dados_aba5 and ('opcoes5' + str(s + 3)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 1)] == 'M20x1,5' and dados_aba5['opcoes5' + str(s + 3)] == 'Inox 316':
            i = aba_cx_peq['AD22'].value
         elif ('opcoes5' + str(s)) in dados_aba5 and dados_aba5['opcoes5' + str(s)] == 1:
            i = 'N/A'
         
         # Identificador trilho
         j = aba_trilhos['D3'].value

         # Tag X*
         if ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X1':
            k = aba_trilhos['B2'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X2':
            k = aba_trilhos['B3'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X3':
            k = aba_trilhos['B4'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X4':
            k = aba_trilhos['B5'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X5':
            k = aba_trilhos['B6'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X6':
            k = aba_trilhos['B7'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X7':
            k = aba_trilhos['B8'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X8':
            k = aba_trilhos['B9'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X9':
            k = aba_trilhos['B10'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X10':
            k = aba_trilhos['B11'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X11':
            k = aba_trilhos['B12'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X12':
            k = aba_trilhos['B13'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X13':
            k = aba_trilhos['B14'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X14':
            k = aba_trilhos['B15'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X15':
            k = aba_trilhos['B16'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X16':
            k = aba_trilhos['B17'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X17':
            k = aba_trilhos['B18'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X18':
            k = aba_trilhos['B19'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X19':
            k = aba_trilhos['B20'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X20':
            k = aba_trilhos['B21'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X21':
            k = aba_trilhos['B22'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X22':
            k = aba_trilhos['B23'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X23':
            k = aba_trilhos['B24'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X24':
            k = aba_trilhos['B25'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X25':
            k = aba_trilhos['B26'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X26':
            k = aba_trilhos['B27'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X27':
            k = aba_trilhos['B28'].value
         elif ('opcoes5' + str(s + 4)) in dados_aba5 and dados_aba5['opcoes5' + str(s + 4)] == 'X28':
            k = aba_trilhos['B29'].value

         # Conj. aterramento
         if 'opcoes53' in dados_aba5 and dados_aba5['opcoes53'] == 'Com aterramento':
            if 'opcoes13' in dados_aba1 and ('opcoes5' + str(s)) in dados_aba5 and dados_aba1['opcoes13'] == 'Aço carbono' and dados_aba5['opcoes5' + str(s)] == 1:
                  l = aba_cx_peq['AK3'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes5' + str(s)) in dados_aba5 and dados_aba1['opcoes13'] == 'Inox 304' and dados_aba5['opcoes5' + str(s)] == 1:
                  l = aba_cx_peq['AK4'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes5' + str(s)) in dados_aba5 and dados_aba1['opcoes13'] == 'Inox 316' and dados_aba5['opcoes5' + str(s)] == 1:
                  l = aba_cx_peq['AK5'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes5' + str(s)) in dados_aba5 and dados_aba1['opcoes13'] == 'Inox 316L' and dados_aba5['opcoes5' + str(s)] == 1:
                  l = aba_cx_peq['AK6'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes5' + str(s)) in dados_aba5 and dados_aba1['opcoes13'] == 'Aço carbono' and dados_aba5['opcoes5' + str(s)] == 2:
                  l = aba_cx_peq['AK7'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes5' + str(s)) in dados_aba5 and dados_aba1['opcoes13'] == 'Inox 304' and dados_aba5['opcoes5' + str(s)] == 2:
                  l = aba_cx_peq['AK8'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes5' + str(s)) in dados_aba5 and dados_aba1['opcoes13'] == 'Inox 316' and dados_aba5['opcoes5' + str(s)] == 2:
                  l = aba_cx_peq['AK9'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes5' + str(s)) in dados_aba5 and dados_aba1['opcoes13'] == 'Inox 316L' and dados_aba5['opcoes5' + str(s)] == 2:
                  l = aba_cx_peq['AK10'].value
         elif 'opcoes53' in dados_aba5 and dados_aba5['opcoes53'] == 'Sem aterramento':
            l = 'N/A'
         
         # Caso seja requisitado 4 caixas
         if dados_aba5['opcoes51'] == '4':
            # Preenche o nome e campos da caixa 1
            if s == 4:
               aba_pa['A244'].value = str(dados_aba5['texto51'])
               aba_pa['E246'].value = a
               aba_pa['E248'].value = b
               aba_pa['E250'].value = c
               aba_pa['E256'].value = d
               aba_pa['E258'].value = e
               aba_pa['E253'].value = f
               aba_pa['E249'].value = g
               aba_pa['E247'].value = h
               aba_pa['E257'].value = i
               aba_pa['E254'].value = j
               aba_pa['E255'].value = k
               aba_pa['E259'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba5['inc51'])
               pos_inic = int(dados_aba5['lt51'])
               celulas = 'D' + str(246)
               for m in range(14):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(246 + m + 1)

            # Preenche o nome e campos da caixa 2
            if s == 11:
               aba_pa['A260'].value = str(dados_aba5['texto52'])
               aba_pa['E261'].value = a
               aba_pa['E263'].value = b
               aba_pa['E265'].value = c
               aba_pa['E271'].value = d
               aba_pa['E273'].value = e
               aba_pa['E268'].value = f
               aba_pa['E264'].value = g
               aba_pa['E262'].value = h
               aba_pa['E272'].value = i
               aba_pa['E269'].value = j
               aba_pa['E270'].value = k
               aba_pa['E274'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba5['inc52'])
               pos_inic = int(dados_aba5['lt52'])
               celulas = 'D' + str(261)
               for m in range(14):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(261 + m + 1)

            # Preenche o nome e campos da caixa 3
            if s == 18:
               aba_pa['A275'].value = str(dados_aba5['texto53'])
               aba_pa['E276'].value = a
               aba_pa['E278'].value = b
               aba_pa['E280'].value = c
               aba_pa['E286'].value = d
               aba_pa['E288'].value = e
               aba_pa['E283'].value = f
               aba_pa['E279'].value = g
               aba_pa['E277'].value = h
               aba_pa['E287'].value = i
               aba_pa['E284'].value = j
               aba_pa['E285'].value = k
               aba_pa['E289'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba5['inc53'])
               pos_inic = int(dados_aba5['lt53'])
               celulas = 'D' + str(276)
               for m in range(14):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(276 + m + 1)

            # Preenche o nome e campos da caixa 4
            if s == 25:
               aba_pa['A290'].value = str(dados_aba5['texto54'])
               aba_pa['E291'].value = a
               aba_pa['E293'].value = b
               aba_pa['E295'].value = c
               aba_pa['E301'].value = d
               aba_pa['E303'].value = e
               aba_pa['E298'].value = f
               aba_pa['E294'].value = g
               aba_pa['E292'].value = h
               aba_pa['E302'].value = i
               aba_pa['E299'].value = j
               aba_pa['E300'].value = k
               aba_pa['E304'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba5['inc54'])
               pos_inic = int(dados_aba5['lt54'])
               celulas = 'D' + str(291)
               for m in range(14):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(291 + m + 1)

         # Caso seja requisitado 3 caixas
         if dados_aba5['opcoes51'] == '3':
            # Preenche o nome e campos da caixa 1
            if s == 4:
               aba_pa['A244'].value = str(dados_aba5['texto51'])
               aba_pa['E246'].value = a
               aba_pa['E248'].value = b
               aba_pa['E250'].value = c
               aba_pa['E256'].value = d
               aba_pa['E258'].value = e
               aba_pa['E253'].value = f
               aba_pa['E249'].value = g
               aba_pa['E247'].value = h
               aba_pa['E257'].value = i
               aba_pa['E254'].value = j
               aba_pa['E255'].value = k
               aba_pa['E259'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba5['inc51'])
               pos_inic = int(dados_aba5['lt51'])
               celulas = 'D' + str(246)
               for m in range(14):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(246 + m + 1)

            # Preenche o nome e campos da caixa 2
            if s == 11:
               aba_pa['A260'].value = str(dados_aba5['texto52'])
               aba_pa['E261'].value = a
               aba_pa['E263'].value = b
               aba_pa['E265'].value = c
               aba_pa['E271'].value = d
               aba_pa['E273'].value = e
               aba_pa['E268'].value = f
               aba_pa['E264'].value = g
               aba_pa['E262'].value = h
               aba_pa['E272'].value = i
               aba_pa['E269'].value = j
               aba_pa['E270'].value = k
               aba_pa['E274'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba5['inc52'])
               pos_inic = int(dados_aba5['lt52'])
               celulas = 'D' + str(261)
               for m in range(14):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(261 + m + 1)

            # Preenche o nome e campos da caixa 3
            if s == 18:
               aba_pa['A275'].value = str(dados_aba5['texto53'])
               aba_pa['E276'].value = a
               aba_pa['E278'].value = b
               aba_pa['E280'].value = c
               aba_pa['E286'].value = d
               aba_pa['E288'].value = e
               aba_pa['E283'].value = f
               aba_pa['E279'].value = g
               aba_pa['E277'].value = h
               aba_pa['E287'].value = i
               aba_pa['E284'].value = j
               aba_pa['E285'].value = k
               aba_pa['E289'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba5['inc53'])
               pos_inic = int(dados_aba5['lt53'])
               celulas = 'D' + str(276)
               for m in range(14):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(276 + m + 1)
            
         # Caso seja requisitado 2 caixas
         if dados_aba5['opcoes51'] == '2':
            # Preenche o nome e campos da caixa 1
            if s == 4:
               aba_pa['A244'].value = str(dados_aba5['texto51'])
               aba_pa['E246'].value = a
               aba_pa['E248'].value = b
               aba_pa['E250'].value = c
               aba_pa['E256'].value = d
               aba_pa['E258'].value = e
               aba_pa['E253'].value = f
               aba_pa['E249'].value = g
               aba_pa['E247'].value = h
               aba_pa['E257'].value = i
               aba_pa['E254'].value = j
               aba_pa['E255'].value = k
               aba_pa['E259'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba5['inc51'])
               pos_inic = int(dados_aba5['lt51'])
               celulas = 'D' + str(246)
               for m in range(14):
                  aba_pa[celulas].value = pos_inic + (m * incremento)
                  celulas = 'D' + str(246 + m + 1)

            # Preenche o nome e campos da caixa 2
            if s == 11:
               aba_pa['A260'].value = str(dados_aba5['texto52'])
               aba_pa['E261'].value = a
               aba_pa['E263'].value = b
               aba_pa['E265'].value = c
               aba_pa['E271'].value = d
               aba_pa['E273'].value = e
               aba_pa['E268'].value = f
               aba_pa['E264'].value = g
               aba_pa['E262'].value = h
               aba_pa['E272'].value = i
               aba_pa['E269'].value = j
               aba_pa['E270'].value = k
               aba_pa['E274'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba5['inc52'])
               pos_inic = int(dados_aba5['lt52'])
               celulas = 'D' + str(261)
               for m in range(14):
                  aba_pa[celulas].value = pos_inic + (m * incremento)
                  celulas = 'D' + str(261 + m + 1)

         # Caso seja requisitado 1 caixa
         if dados_aba5['opcoes51'] == '1':
            # Preenche o nome e campos da caixa 1
            if s == 4:
               aba_pa['A244'].value = str(dados_aba5['texto51'])
               aba_pa['E246'].value = a
               aba_pa['E248'].value = b
               aba_pa['E250'].value = c
               aba_pa['E256'].value = d
               aba_pa['E258'].value = e
               aba_pa['E253'].value = f
               aba_pa['E249'].value = g
               aba_pa['E247'].value = h
               aba_pa['E257'].value = i
               aba_pa['E254'].value = j
               aba_pa['E255'].value = k
               aba_pa['E259'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba5['inc51'])
               pos_inic = int(dados_aba5['lt51'])
               celulas = 'D' + str(246)
               for m in range(14):
                  aba_pa[celulas].value = pos_inic + (m * incremento)
                  celulas = 'D' + str(246 + m + 1)

         if dados_aba5['opcoes51'] == '0':
            pass


      # Regras da aba 6 (Cx. Acess. Média) =========================================================================================================================================
      s = 0
      for n in range(4):
         if (n + 1) == 1:
            s = 4
         elif (n + 1) == 2:
            s = 15
         elif (n + 1) == 3:
            s = 26
         elif (n + 1) == 4:
            s = 37
      
         # Conjunto tampa caixa de ligação
         if 'opcoes62' in dados_aba6 and ('opcoes6' + str(s + 9)) in dados_aba6 and dados_aba6['opcoes62'] == 'Inox 304' and dados_aba6['opcoes6' + str(s + 9)] == 'Não':
            a = aba_cx_med['C5'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s + 9)) in dados_aba6 and dados_aba6['opcoes62'] == 'Inox 304' and dados_aba6['opcoes6' + str(s + 9)] == 'Sim':
            a = aba_cx_med['C8'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s + 9)) in dados_aba6 and dados_aba6['opcoes62'] == 'Inox 316' and dados_aba6['opcoes6' + str(s + 9)] == 'Não':
            a = aba_cx_med['C4'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s + 9)) in dados_aba6 and dados_aba6['opcoes62'] == 'Inox 316' and dados_aba6['opcoes6' + str(s + 9)] == 'Sim':
            a = aba_cx_med['C7'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s + 9)) in dados_aba6 and dados_aba6['opcoes62'] == 'Ferro fundido' and dados_aba6['opcoes6' + str(s + 9)] == 'Não':
            a = aba_cx_med['C3'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s + 9)) in dados_aba6 and dados_aba6['opcoes62'] == 'Ferro fundido' and dados_aba6['opcoes6' + str(s + 9)] == 'Sim':
            a = aba_cx_med['C6'].value

         # Placa de fixação de trilho
         if ('opcoes6' + str(s + 5)) in dados_aba6 and ('opcoes6' + str(s + 8)) in dados_aba6 and ('opcoes6' + str(s + 10)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 5)] == '1' and dados_aba6['opcoes6' + str(s + 10)] == 'Área Segura' and dados_aba6['opcoes6' + str(s + 8)] == 'Não':
            b = aba_cx_med['F3'].value
         elif ('opcoes6' + str(s + 5)) in dados_aba6 and ('opcoes6' + str(s + 8)) in dados_aba6 and ('opcoes6' + str(s + 10)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 5)] == '2' and dados_aba6['opcoes6' + str(s + 10)] == 'Área Segura' and dados_aba6['opcoes6' + str(s + 8)] == 'Não':
            b = aba_cx_med['F4'].value
         elif ('opcoes6' + str(s + 5)) in dados_aba6 and ('opcoes6' + str(s + 8)) in dados_aba6 and ('opcoes6' + str(s + 10)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 5)] == '1' and dados_aba6['opcoes6' + str(s + 10)] == 'Área Segura' and dados_aba6['opcoes6' + str(s + 8)] == 'Sim':
            b = aba_cx_med['F5'].value
         elif ('opcoes6' + str(s + 5)) in dados_aba6 and ('opcoes6' + str(s + 8)) in dados_aba6 and ('opcoes6' + str(s + 10)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 5)] == '2' and dados_aba6['opcoes6' + str(s + 10)] == 'Área Segura' and dados_aba6['opcoes6' + str(s + 8)] == 'Sim':
            b = aba_cx_med['F6'].value
         elif ('opcoes6' + str(s + 5)) in dados_aba6 and ('opcoes6' + str(s + 10)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 5)] == '1' and dados_aba6['opcoes6' + str(s + 10)] == 'Ex-e':
            b = aba_cx_med['F7'].value
         elif ('opcoes6' + str(s + 5)) in dados_aba6 and ('opcoes6' + str(s + 10)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 5)] == '2' and dados_aba6['opcoes6' + str(s + 10)] == 'Ex-e':
            b = aba_cx_med['F8'].value
         elif ('opcoes6' + str(s + 5)) in dados_aba6 and ('opcoes6' + str(s + 8)) in dados_aba6 and ('opcoes6' + str(s + 10)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 5)] == '1' and dados_aba6['opcoes6' + str(s + 10)] == 'Ex-p' and dados_aba6['opcoes6' + str(s + 8)] == 'Não':
            b = aba_cx_med['F9'].value
         elif ('opcoes6' + str(s + 5)) in dados_aba6 and ('opcoes6' + str(s + 8)) in dados_aba6 and ('opcoes6' + str(s + 10)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 5)] == '2' and dados_aba6['opcoes6' + str(s + 10)] == 'Ex-p' and dados_aba6['opcoes6' + str(s + 8)] == 'Não':
            b = aba_cx_med['F10'].value

         # Conj. fixação caixa de acessórios
         if dados_aba1['opcoes13'] == 'Aço carbono':
            c = aba_cx_med['I3'].value
         elif dados_aba1['opcoes13'] == 'Inox 304':
            c = aba_cx_med['I4'].value
         elif dados_aba1['opcoes13'] == 'Inox 316':
            c = aba_cx_med['I5'].value
         elif dados_aba1['opcoes13'] == 'Inox 316L':
            c = aba_cx_med['I6'].value

         # Tampão roscado
         if ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 2)] == 'G 1"':
            d = aba_cx_med['L3'].value
         elif ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 2)] == 'NPT 1"':
            d = aba_cx_med['L4'].value
         elif ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 2)] == 'M20x1,5':
            d = aba_cx_med['L5'].value
         else:
            d = 'N/A'

         # Parafuso fixação M6x20
         if 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Aço carbono':
            e = aba_cx_med['O3'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 304':
            e = aba_cx_med['O4'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316':
            e = aba_cx_med['O5'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316L':
            e = aba_cx_med['O6'].value

         # Parafuso fixação triho
         if 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Aço carbono':
            f = aba_cx_med['R3'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 304':
            f = aba_cx_med['R4'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316':
            f = aba_cx_med['R5'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316L':
            f = aba_cx_med['R6'].value

         # Ilho de borracha
         if ('opcoes6' + str(s + 10)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 10)] == 'Área Segura':
            g = aba_cx_med['U3'].value
         elif ('opcoes6' + str(s + 8)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 8)] == 'Sim':
            g = aba_cx_med['U4'].value
         elif ('opcoes6' + str(s + 10)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 10)] == 'Ex-e':
            g = aba_cx_med['U4'].value
         elif ('opcoes6' + str(s + 10)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 10)] == 'Ex-p':
            g = aba_cx_med['U4'].value

         # Caixa de ligação fundida
         if 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Inox 304' and dados_aba6['opcoes6' + str(s)] == 2 and dados_aba6['opcoes6' + str(s + 2)] == 'Sem rosca' and dados_aba6['opcoes6' + str(s + 3)] == '0':
            h = aba_cx_med['AA3'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Inox 316' and dados_aba6['opcoes6' + str(s)] == 2 and dados_aba6['opcoes6' + str(s + 2)] == 'Sem rosca' and dados_aba6['opcoes6' + str(s + 3)] == '0':
            h = aba_cx_med['AA4'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Ferro fundido' and dados_aba6['opcoes6' + str(s)] == 2 and dados_aba6['opcoes6' + str(s + 2)] == 'Sem rosca' and dados_aba6['opcoes6' + str(s + 3)] == '0':
            h = aba_cx_med['AA5'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Inox 304' and dados_aba6['opcoes6' + str(s)] == 1 and dados_aba6['opcoes6' + str(s + 2)] == 'G 1"' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            h = aba_cx_med['AA6'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Inox 316' and dados_aba6['opcoes6' + str(s)] == 1 and dados_aba6['opcoes6' + str(s + 2)] == 'G 1"' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            h = aba_cx_med['AA7'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Ferro fundido' and dados_aba6['opcoes6' + str(s)] == 1 and dados_aba6['opcoes6' + str(s + 2)] == 'G 1"' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            h = aba_cx_med['AA8'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Inox 304' and dados_aba6['opcoes6' + str(s)] == 1 and dados_aba6['opcoes6' + str(s + 2)] == 'NPT 1"' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            h = aba_cx_med['AA9'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Inox 316' and dados_aba6['opcoes6' + str(s)] == 1 and dados_aba6['opcoes6' + str(s + 2)] == 'NPT 1"' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            h = aba_cx_med['AA10'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Ferro fundido' and dados_aba6['opcoes6' + str(s)] == 1 and dados_aba6['opcoes6' + str(s + 2)] == 'NPT 1"' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            h = aba_cx_med['AA11'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Inox 304' and dados_aba6['opcoes6' + str(s)] == 1 and dados_aba6['opcoes6' + str(s + 2)] == 'M20x1,5' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            h = aba_cx_med['AA12'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Inox 316' and dados_aba6['opcoes6' + str(s)] == 1 and dados_aba6['opcoes6' + str(s + 2)] == 'M20x1,5' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            h = aba_cx_med['AA13'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Ferro fundido' and dados_aba6['opcoes6' + str(s)] == 1 and dados_aba6['opcoes6' + str(s + 2)] == 'M20x1,5' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            h = aba_cx_med['AA14'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Inox 304' and dados_aba6['opcoes6' + str(s)] == 1 and dados_aba6['opcoes6' + str(s + 2)] == 'G 1"' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            h = aba_cx_med['AA15'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Inox 316' and dados_aba6['opcoes6' + str(s)] == 1 and dados_aba6['opcoes6' + str(s + 2)] == 'G 1"' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            h = aba_cx_med['AA16'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Ferro fundido' and dados_aba6['opcoes6' + str(s)] == 1 and dados_aba6['opcoes6' + str(s + 2)] == 'G 1"' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            h = aba_cx_med['AA17'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Inox 304' and dados_aba6['opcoes6' + str(s)] == 1 and dados_aba6['opcoes6' + str(s + 2)] == 'NPT 1"' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            h = aba_cx_med['AA18'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Inox 316' and dados_aba6['opcoes6' + str(s)] == 1 and dados_aba6['opcoes6' + str(s + 2)] == 'NPT 1"' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            h = aba_cx_med['AA19'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Ferro fundido' and dados_aba6['opcoes6' + str(s)] == 1 and dados_aba6['opcoes6' + str(s + 2)] == 'NPT 1"' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            h = aba_cx_med['AA20'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Inox 304' and dados_aba6['opcoes6' + str(s)] == 1 and dados_aba6['opcoes6' + str(s + 2)] == 'M20x1,5' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            h = aba_cx_med['AA21'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Inox 316' and dados_aba6['opcoes6' + str(s)] == 1 and dados_aba6['opcoes6' + str(s + 2)] == 'M20x1,5' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            h = aba_cx_med['AA22'].value
         elif 'opcoes62' in dados_aba6 and ('opcoes6' + str(s)) in dados_aba6 and ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes62'] == 'Ferro fundido' and dados_aba6['opcoes6' + str(s)] == 1 and dados_aba6['opcoes6' + str(s + 2)] == 'M20x1,5' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            h = aba_cx_med['AA23'].value

         # Conj. placa cega s/ parafusos
         if ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Ferro fundido' and dados_aba6['opcoes6' + str(s + 2)] == 'Sem rosca' and dados_aba6['opcoes6' + str(s + 3)] == '0':
            i = aba_cx_med['AE3'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Alumínio (AlSi6Cu4)' and dados_aba6['opcoes6' + str(s + 2)] == 'Sem rosca' and dados_aba6['opcoes6' + str(s + 3)] == '0':
            i = aba_cx_med['AE4'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Alumínio (AlSi10Mg)' and dados_aba6['opcoes6' + str(s + 2)] == 'Sem rosca' and dados_aba6['opcoes6' + str(s + 3)] == '0':
            i = aba_cx_med['A5'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Inox 316' and dados_aba6['opcoes6' + str(s + 2)] == 'Sem rosca' and dados_aba6['opcoes6' + str(s + 3)] == '0':
            i = aba_cx_med['AE6'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Inox 304' and dados_aba6['opcoes6' + str(s + 2)] == 'Sem rosca' and dados_aba6['opcoes6' + str(s + 3)] == '0':
            i = aba_cx_med['AE7'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Ferro fundido' and dados_aba6['opcoes6' + str(s + 2)] == 'G 1"' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            i = aba_cx_med['AE8'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Ferro fundido' and dados_aba6['opcoes6' + str(s + 2)] == 'NPT 1"' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            i = aba_cx_med['AE9'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Ferro fundido' and dados_aba6['opcoes6' + str(s + 2)] == 'M20x1,5' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            i = aba_cx_med['AE10'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Alumínio (AlSi6Cu4)' and dados_aba6['opcoes6' + str(s + 2)] == 'G 1"' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            i = aba_cx_med['AE11'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Alumínio (AlSi6Cu4)' and dados_aba6['opcoes6' + str(s + 2)] == 'NPT 1"' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            i = aba_cx_med['AE12'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Alumínio (AlSi6Cu4)' and dados_aba6['opcoes6' + str(s + 2)] == 'M20x1,5' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            i = aba_cx_med['AE13'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Alumínio (AlSi10Mg)' and dados_aba6['opcoes6' + str(s + 2)] == 'G 1"' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            i = aba_cx_med['AE14'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Alumínio (AlSi10Mg)' and dados_aba6['opcoes6' + str(s + 2)] == 'NPT 1"' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            i = aba_cx_med['AE15'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Alumínio (AlSi10Mg)' and dados_aba6['opcoes6' + str(s + 2)] == 'M20x1,5' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            i = aba_cx_med['AE16'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Inox 316' and dados_aba6['opcoes6' + str(s + 2)] == 'G 1"' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            i = aba_cx_med['AE17'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Inox 316' and dados_aba6['opcoes6' + str(s + 2)] == 'NPT 1"' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            i = aba_cx_med['AE18'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Inox 316' and dados_aba6['opcoes6' + str(s + 2)] == 'M20x1,5' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            i = aba_cx_med['AE19'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Inox 304' and dados_aba6['opcoes6' + str(s + 2)] == 'G 1"' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            i = aba_cx_med['AE20'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Inox 304' and dados_aba6['opcoes6' + str(s + 2)] == 'NPT 1"' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            i = aba_cx_med['AE21'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Inox 304' and dados_aba6['opcoes6' + str(s + 2)] == 'M20x1,5' and dados_aba6['opcoes6' + str(s + 3)] == '1':
            i = aba_cx_med['AE22'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Ferro fundido' and dados_aba6['opcoes6' + str(s + 2)] == 'G 1"' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            i = aba_cx_med['AE23'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Ferro fundido' and dados_aba6['opcoes6' + str(s + 2)] == 'NPT 1"' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            i = aba_cx_med['AE24'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Ferro fundido' and dados_aba6['opcoes6' + str(s + 2)] == 'M20x1,5' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            i = aba_cx_med['AE25'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Alumínio (AlSi6Cu4)' and dados_aba6['opcoes6' + str(s + 2)] == 'G 1"' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            i = aba_cx_med['AE26'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Alumínio (AlSi6Cu4)' and dados_aba6['opcoes6' + str(s + 2)] == 'NPT 1"' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            i = aba_cx_med['AE27'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Alumínio (AlSi6Cu4)' and dados_aba6['opcoes6' + str(s + 2)] == 'M20x1,5' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            i = aba_cx_med['AE28'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Alumínio (AlSi10Mg)' and dados_aba6['opcoes6' + str(s + 2)] == 'G 1"' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            i = aba_cx_med['AE29'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Alumínio (AlSi10Mg)' and dados_aba6['opcoes6' + str(s + 2)] == 'NPT 1"' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            i = aba_cx_med['AE30'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Alumínio (AlSi10Mg)' and dados_aba6['opcoes6' + str(s + 2)] == 'M20x1,5' and dados_aba6['opcoes6' + str(s + 32)] == '2':
            i = aba_cx_med['AE31'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Inox 316' and dados_aba6['opcoes6' + str(s + 2)] == 'G 1"' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            i = aba_cx_med['AE32'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Inox 316' and dados_aba6['opcoes6' + str(s + 2)] == 'NPT 1"' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            i = aba_cx_med['AE33'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Inox 316' and dados_aba6['opcoes6' + str(s + 2)] == 'M20x1,5' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            i = aba_cx_med['AE34'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Inox 304' and dados_aba6['opcoes6' + str(s + 2)] == 'G 1"' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            i = aba_cx_med['AE35'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Inox 304' and dados_aba6['opcoes6' + str(s + 2)] == 'NPT 1"' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            i = aba_cx_med['AE36'].value
         elif ('opcoes6' + str(s + 3)) in dados_aba6 and ('opcoes6' + str(s + 4)) in dados_aba6 and ('opcoes6' + str(s + 2)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 4)] == 'Inox 304' and dados_aba6['opcoes6' + str(s + 2)] == 'M20x1,5' and dados_aba6['opcoes6' + str(s + 3)] == '2':
            i = aba_cx_med['AE37'].value
         elif ('opcoes6' + str(s)) in dados_aba6 and dados_aba6['opcoes6' + str(s)] == 1:
            i = 'N/A'
         
         # Identificador trilho
         j = aba_trilhos['D3'].value

         # Tag X* (1)
         if ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X1':
            k1 = aba_trilhos['B2'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X2':
            k1 = aba_trilhos['B3'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X3':
            k1 = aba_trilhos['B4'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X4':
            k1 = aba_trilhos['B5'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X5':
            k1 = aba_trilhos['B6'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X6':
            k1 = aba_trilhos['B7'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X7':
            k1 = aba_trilhos['B8'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X8':
            k1 = aba_trilhos['B9'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X9':
            k1 = aba_trilhos['B10'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X10':
            k1 = aba_trilhos['B11'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X11':
            k1 = aba_trilhos['B12'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X12':
            k1 = aba_trilhos['B13'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X13':
            k1 = aba_trilhos['B14'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X14':
            k1 = aba_trilhos['B15'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X15':
            k1 = aba_trilhos['B16'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X16':
            k1 = aba_trilhos['B17'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X17':
            k1 = aba_trilhos['B18'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X18':
            k1 = aba_trilhos['B19'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X19':
            k1 = aba_trilhos['B20'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X20':
            k1 = aba_trilhos['B21'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X21':
            k1 = aba_trilhos['B22'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X22':
            k1 = aba_trilhos['B23'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X23':
            k1 = aba_trilhos['B24'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X24':
            k1 = aba_trilhos['B25'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X25':
            k1 = aba_trilhos['B26'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X26':
            k1 = aba_trilhos['B27'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X27':
            k1 = aba_trilhos['B28'].value
         elif ('opcoes6' + str(s + 6)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 6)] == 'X28':
            k1 = aba_trilhos['B29'].value

         # Tag X* (2)
         if ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X1':
            k2 = aba_trilhos['B2'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X2':
            k2 = aba_trilhos['B3'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X3':
            k2 = aba_trilhos['B4'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X4':
            k2 = aba_trilhos['B5'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X5':
            k2 = aba_trilhos['B6'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X6':
            k2 = aba_trilhos['B7'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X7':
            k2 = aba_trilhos['B8'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X8':
            k2 = aba_trilhos['B9'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X9':
            k2 = aba_trilhos['B10'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X10':
            k2 = aba_trilhos['B11'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X11':
            k2 = aba_trilhos['B12'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X12':
            k2 = aba_trilhos['B13'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X13':
            k2 = aba_trilhos['B14'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X14':
            k2 = aba_trilhos['B15'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X15':
            k2 = aba_trilhos['B16'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X16':
            k2 = aba_trilhos['B17'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X17':
            k2 = aba_trilhos['B18'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X18':
            k2 = aba_trilhos['B19'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X19':
            k2 = aba_trilhos['B20'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X20':
            k2 = aba_trilhos['B21'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X21':
            k2 = aba_trilhos['B22'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X22':
            k2 = aba_trilhos['B23'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X23':
            k2 = aba_trilhos['B24'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X24':
            k2 = aba_trilhos['B25'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X25':
            k2 = aba_trilhos['B26'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X26':
            k2 = aba_trilhos['B27'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X27':
            k2 = aba_trilhos['B28'].value
         elif ('opcoes6' + str(s + 7)) in dados_aba6 and dados_aba6['opcoes6' + str(s + 7)] == 'X28':
            k2 = aba_trilhos['B29'].value
         else:
            k2 = 'N/A'

         # Conj. aterramento
         if 'opcoes63' in dados_aba6 and dados_aba6['opcoes63'] == 'Com aterramento':
            if 'opcoes13' in dados_aba1 and ('opcoes6' + str(s)) in dados_aba6 and dados_aba1['opcoes13'] == 'Aço carbono' and dados_aba6['opcoes6' + str(s)] == 1:
                  l = aba_cx_med['AL3'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes6' + str(s)) in dados_aba6 and dados_aba1['opcoes13'] == 'Inox 304' and dados_aba6['opcoes6' + str(s)] == 1:
                  l = aba_cx_med['AL4'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes6' + str(s)) in dados_aba6 and dados_aba1['opcoes13'] == 'Inox 316' and dados_aba6['opcoes6' + str(s)] == 1:
                  l = aba_cx_med['AL5'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes6' + str(s)) in dados_aba6 and dados_aba1['opcoes13'] == 'Inox 316L' and dados_aba6['opcoes6' + str(s)] == 1:
                  l = aba_cx_med['AL6'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes6' + str(s)) in dados_aba6 and dados_aba1['opcoes13'] == 'Aço carbono' and dados_aba6['opcoes6' + str(s)] == 2:
                  l = aba_cx_med['AL7'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes6' + str(s)) in dados_aba6 and dados_aba1['opcoes13'] == 'Inox 304' and dados_aba6['opcoes6' + str(s)] == 2:
                  l = aba_cx_med['AL8'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes6' + str(s)) in dados_aba6 and dados_aba1['opcoes13'] == 'Inox 316' and dados_aba6['opcoes6' + str(s)] == 2:
                  l = aba_cx_med['AL9'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes6' + str(s)) in dados_aba6 and dados_aba1['opcoes13'] == 'Inox 316L' and dados_aba6['opcoes6' + str(s)] == 2:
                  l = aba_cx_med['AL10'].value
         elif 'opcoes63' in dados_aba6 and dados_aba6['opcoes63'] == 'Sem aterramento':
            l = 'N/A'

         # Caso seja requisitado 4 caixas
         if dados_aba6['opcoes61'] == '4':
            # Preenche o nome e campos da caixa 1
            if s == 4:
               aba_pa['A306'].value = str(dados_aba6['texto61'])
               aba_pa['E307'].value = a
               aba_pa['E309'].value = b
               aba_pa['E311'].value = c
               aba_pa['E320'].value = d
               aba_pa['E322'].value = e
               aba_pa['E316'].value = f
               aba_pa['E310'].value = g
               aba_pa['E308'].value = h
               aba_pa['E321'].value = i
               aba_pa['E317'].value = j
               aba_pa['E318'].value = k1
               aba_pa['E319'].value = k2
               aba_pa['E323'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba6['inc61'])
               pos_inic = int(dados_aba6['lt61'])
               celulas = 'D' + str(307)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(307 + m + 1)
               
               # Imprimindo as quantidades na planilha resposta
               if opcoes69.get() == '1':
                  aba_pa['H310'].value = 1
                  aba_pa['H316'].value = 2
                  aba_pa['H317'].value = 1
                  
               elif opcoes69.get() == '2':
                  aba_pa['H310'].value = 2
                  aba_pa['H316'].value = 4
                  aba_pa['H317'].value = 2

               if opcoes67.get() == '1':
                  aba_pa['H320'].value = 1
                  
               elif opcoes67.get() == '2':
                  aba_pa['H320'].value = 2
               
            # Preenche o nome e campos da caixa 2
            if s == 15:
               aba_pa['A324'].value = str(dados_aba6['texto62'])
               aba_pa['E325'].value = a
               aba_pa['E327'].value = b
               aba_pa['E329'].value = c
               aba_pa['E338'].value = d
               aba_pa['E340'].value = e
               aba_pa['E334'].value = f
               aba_pa['E328'].value = g
               aba_pa['E326'].value = h
               aba_pa['E339'].value = i
               aba_pa['E335'].value = j
               aba_pa['E336'].value = k1
               aba_pa['E337'].value = k2
               aba_pa['E341'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba6['inc62'])
               pos_inic = int(dados_aba6['lt62'])
               celulas = 'D' + str(325)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(325 + m + 1)

               # Imprimindo as quantidades na planilha resposta
               if opcoes620.get() == '1':
                  aba_pa['H328'].value = 1
                  aba_pa['H334'].value = 2
                  aba_pa['H335'].value = 1
                  
               elif opcoes620.get() == '2':
                  aba_pa['H328'].value = 2
                  aba_pa['H334'].value = 4
                  aba_pa['H335'].value = 2

               if opcoes618.get() == '1':
                  aba_pa['H338'].value = 1
                  
               elif opcoes618.get() == '2':
                  aba_pa['H338'].value = 2

            # Preenche o nome e campos da caixa 3
            if s == 26:
               aba_pa['A342'].value = str(dados_aba6['texto63'])
               aba_pa['E343'].value = a
               aba_pa['E345'].value = b
               aba_pa['E347'].value = c
               aba_pa['E356'].value = d
               aba_pa['E358'].value = e
               aba_pa['E352'].value = f
               aba_pa['E346'].value = g
               aba_pa['E344'].value = h
               aba_pa['E357'].value = i
               aba_pa['E353'].value = j
               aba_pa['E354'].value = k1
               aba_pa['E355'].value = k2
               aba_pa['E359'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba6['inc63'])
               pos_inic = int(dados_aba6['lt63'])
               celulas = 'D' + str(343)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(343 + m + 1)

               # Imprimindo as quantidades na planilha resposta
               if opcoes620.get() == '1':
                  aba_pa['H346'].value = 1
                  aba_pa['H352'].value = 2
                  aba_pa['H353'].value = 1
                  
               elif opcoes620.get() == '2':
                  aba_pa['H346'].value = 2
                  aba_pa['H352'].value = 4
                  aba_pa['H353'].value = 2

               if opcoes618.get() == '1':
                  aba_pa['H356'].value = 1
                  
               elif opcoes618.get() == '2':
                  aba_pa['H356'].value = 2

            # Preenche o nome e campos da caixa 4
            if s == 37:
               aba_pa['A360'].value = str(dados_aba6['texto64'])
               aba_pa['E361'].value = a
               aba_pa['E363'].value = b
               aba_pa['E365'].value = c
               aba_pa['E374'].value = d
               aba_pa['E376'].value = e
               aba_pa['E370'].value = f
               aba_pa['E364'].value = g
               aba_pa['E362'].value = h
               aba_pa['E375'].value = i
               aba_pa['E371'].value = j
               aba_pa['E372'].value = k1
               aba_pa['E373'].value = k2
               aba_pa['E377'].value = l

            # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba6['inc64'])
               pos_inic = int(dados_aba6['lt64'])
               celulas = 'D' + str(361)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(361 + m + 1)

            # Imprimindo as quantidades na planilha resposta
               if opcoes620.get() == '1':
                  aba_pa['H364'].value = 1
                  aba_pa['H370'].value = 2
                  aba_pa['H371'].value = 1
                  
               elif opcoes620.get() == '2':
                  aba_pa['H364'].value = 2
                  aba_pa['H370'].value = 4
                  aba_pa['H371'].value = 2

               if opcoes618.get() == '1':
                  aba_pa['H374'].value = 1
                  
               elif opcoes618.get() == '2':
                  aba_pa['H374'].value = 2

         # Caso seja requisitado 3 caixas
         if dados_aba6['opcoes61'] == '3':
            # Preenche o nome e campos da caixa 1
            if s == 4:
               aba_pa['A306'].value = str(dados_aba6['texto61'])
               aba_pa['E307'].value = a
               aba_pa['E309'].value = b
               aba_pa['E311'].value = c
               aba_pa['E320'].value = d
               aba_pa['E322'].value = e
               aba_pa['E316'].value = f
               aba_pa['E310'].value = g
               aba_pa['E308'].value = h
               aba_pa['E321'].value = i
               aba_pa['E317'].value = j
               aba_pa['E318'].value = k1
               aba_pa['E319'].value = k2
               aba_pa['E323'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba6['inc61'])
               pos_inic = int(dados_aba6['lt61'])
               celulas = 'D' + str(307)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(307 + m + 1)

               # Imprimindo as quantidades na planilha resposta
               if opcoes69.get() == '1':
                  aba_pa['H310'].value = 1
                  aba_pa['H316'].value = 2
                  aba_pa['H317'].value = 1
                  
               elif opcoes69.get() == '2':
                  aba_pa['H310'].value = 2
                  aba_pa['H316'].value = 4
                  aba_pa['H317'].value = 2

               if opcoes67.get() == '1':
                  aba_pa['H320'].value = 1
                  
               elif opcoes67.get() == '2':
                  aba_pa['H320'].value = 2

            # Preenche o nome e campos da caixa 2
            if s == 15:
               aba_pa['A324'].value = str(dados_aba6['texto62'])
               aba_pa['E325'].value = a
               aba_pa['E327'].value = b
               aba_pa['E329'].value = c
               aba_pa['E338'].value = d
               aba_pa['E340'].value = e
               aba_pa['E334'].value = f
               aba_pa['E328'].value = g
               aba_pa['E326'].value = h
               aba_pa['E339'].value = i
               aba_pa['E335'].value = j
               aba_pa['E336'].value = k1
               aba_pa['E337'].value = k2
               aba_pa['E341'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba6['inc62'])
               pos_inic = int(dados_aba6['lt62'])
               celulas = 'D' + str(325)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(325 + m + 1)

               # Imprimindo as quantidades na planilha resposta
               if opcoes620.get() == '1':
                  aba_pa['H328'].value = 1
                  aba_pa['H334'].value = 2
                  aba_pa['H335'].value = 1
                  
               elif opcoes620.get() == '2':
                  aba_pa['H328'].value = 2
                  aba_pa['H334'].value = 4
                  aba_pa['H335'].value = 2

               if opcoes618.get() == '1':
                  aba_pa['H338'].value = 1
                  
               elif opcoes618.get() == '2':
                  aba_pa['H338'].value = 2

            # Preenche o nome e campos da caixa 3
            if s == 26:
               aba_pa['A342'].value = str(dados_aba6['texto63'])
               aba_pa['E343'].value = a
               aba_pa['E345'].value = b
               aba_pa['E347'].value = c
               aba_pa['E356'].value = d
               aba_pa['E358'].value = e
               aba_pa['E352'].value = f
               aba_pa['E346'].value = g
               aba_pa['E344'].value = h
               aba_pa['E357'].value = i
               aba_pa['E353'].value = j
               aba_pa['E354'].value = k1
               aba_pa['E355'].value = k2
               aba_pa['E359'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba6['inc63'])
               pos_inic = int(dados_aba6['lt63'])
               celulas = 'D' + str(343)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(343 + m + 1)

               # Imprimindo as quantidades na planilha resposta
               if opcoes620.get() == '1':
                  aba_pa['H346'].value = 1
                  aba_pa['H352'].value = 2
                  aba_pa['H353'].value = 1
                  
               elif opcoes620.get() == '2':
                  aba_pa['H346'].value = 2
                  aba_pa['H352'].value = 4
                  aba_pa['H353'].value = 2

               if opcoes618.get() == '1':
                  aba_pa['H356'].value = 1
                  
               elif opcoes618.get() == '2':
                  aba_pa['H356'].value = 2
            
         # Caso seja requisitado 2 caixas
         if dados_aba6['opcoes61'] == '2':
            # Preenche o nome e campos da caixa 1
            if s == 4:
               aba_pa['A306'].value = str(dados_aba6['texto61'])
               aba_pa['E307'].value = a
               aba_pa['E309'].value = b
               aba_pa['E311'].value = c
               aba_pa['E320'].value = d
               aba_pa['E322'].value = e
               aba_pa['E316'].value = f
               aba_pa['E310'].value = g
               aba_pa['E308'].value = h
               aba_pa['E321'].value = i
               aba_pa['E317'].value = j
               aba_pa['E318'].value = k1
               aba_pa['E319'].value = k2
               aba_pa['E323'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba6['inc61'])
               pos_inic = int(dados_aba6['lt61'])
               celulas = 'D' + str(307)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(307 + m + 1)

               # Imprimindo as quantidades na planilha resposta
               if opcoes69.get() == '1':
                  aba_pa['H310'].value = 1
                  aba_pa['H316'].value = 2
                  aba_pa['H317'].value = 1
                  
               elif opcoes69.get() == '2':
                  aba_pa['H310'].value = 2
                  aba_pa['H316'].value = 4
                  aba_pa['H317'].value = 2

               if opcoes67.get() == '1':
                  aba_pa['H320'].value = 1
                  
               elif opcoes67.get() == '2':
                  aba_pa['H320'].value = 2


            # Preenche o nome e campos da caixa 2
            if s == 15:
               aba_pa['A324'].value = str(dados_aba6['texto62'])
               aba_pa['E325'].value = a
               aba_pa['E327'].value = b
               aba_pa['E329'].value = c
               aba_pa['E338'].value = d
               aba_pa['E340'].value = e
               aba_pa['E334'].value = f
               aba_pa['E328'].value = g
               aba_pa['E326'].value = h
               aba_pa['E339'].value = i
               aba_pa['E335'].value = j
               aba_pa['E336'].value = k1
               aba_pa['E337'].value = k2
               aba_pa['E341'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba6['inc62'])
               pos_inic = int(dados_aba6['lt62'])
               celulas = 'D' + str(325)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(325 + m + 1)

               # Imprimindo as quantidades na planilha resposta
               if opcoes620.get() == '1':
                  aba_pa['H328'].value = 1
                  aba_pa['H334'].value = 2
                  aba_pa['H335'].value = 1
                  
               elif opcoes620.get() == '2':
                  aba_pa['H328'].value = 2
                  aba_pa['H334'].value = 4
                  aba_pa['H335'].value = 2

               if opcoes618.get() == '1':
                  aba_pa['H338'].value = 1
                  
               elif opcoes618.get() == '2':
                  aba_pa['H338'].value = 2

         # Caso seja requisitado 1 caixa
         if dados_aba6['opcoes61'] == '1':
            # Preenche o nome e campos da caixa 1
            if s == 4:
               aba_pa['A306'].value = str(dados_aba6['texto61'])
               aba_pa['E307'].value = a
               aba_pa['E309'].value = b
               aba_pa['E311'].value = c
               aba_pa['E320'].value = d
               aba_pa['E322'].value = e
               aba_pa['E316'].value = f
               aba_pa['E310'].value = g
               aba_pa['E308'].value = h
               aba_pa['E321'].value = i
               aba_pa['E317'].value = j
               aba_pa['E318'].value = k1
               aba_pa['E319'].value = k2
               aba_pa['E323'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba6['inc61'])
               pos_inic = int(dados_aba6['lt61'])
               celulas = 'D' + str(307)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(307 + m + 1)

               # Imprimindo as quantidades na planilha resposta
               if opcoes69.get() == '1':
                  aba_pa['H310'].value = 1
                  aba_pa['H316'].value = 2
                  aba_pa['H317'].value = 1
                  
               elif opcoes69.get() == '2':
                  aba_pa['H310'].value = 2
                  aba_pa['H316'].value = 4
                  aba_pa['H317'].value = 2

               if opcoes67.get() == '1':
                  aba_pa['H320'].value = 1
                  
               elif opcoes67.get() == '2':
                  aba_pa['H320'].value = 2
            
         if dados_aba6['opcoes61'] == '0':
               pass


      # Regras da aba 7 (Cx. Acess. Grande) =========================================================================================================================================
      s = 0
      for n in range(4):
         if (n + 1) == 1:
            s = 4
         elif (n + 1) == 2:
            s = 15
         elif (n + 1) == 3:
            s = 26
         elif (n + 1) == 4:
            s = 37

         # Conjunto tampa caixa de ligação
         if 'opcoes72' in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 304':
            a = aba_cx_gde['C5'].value
         elif 'opcoes72' in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 316':
            a = aba_cx_gde['C4'].value
         elif 'opcoes72' in dados_aba7 and dados_aba7['opcoes72'] == 'Ferro fundido':
            a = aba_cx_gde['C3'].value

         # Placa de fixação de trilho
         if ('opcoes7' + str(s + 8)) in dados_aba7 and ('opcoes7' + str(s + 10)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 8)] == 'Não' and dados_aba7['opcoes7' + str(s + 10)] == 'Área Segura':
            b = aba_cx_gde['F3'].value
         elif ('opcoes7' + str(s + 8)) in dados_aba7 and ('opcoes7' + str(s + 10)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 8)] == 'Sim' and dados_aba7['opcoes7' + str(s + 10)] == 'Área Segura':
            b = aba_cx_gde['F4'].value
         elif ('opcoes7' + str(s + 10)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 10)] == 'Ex-e':
            b = aba_cx_gde['F5'].value
         elif ('opcoes7' + str(s + 8)) in dados_aba7 and ('opcoes7' + str(s + 10)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 8)] == 'Não' and dados_aba7['opcoes7' + str(s + 10)] == 'Ex-p':
            b = aba_cx_gde['F6'].value

         # Conj. fixação caixa de acessórios
         if dados_aba1['opcoes13'] == 'Aço carbono':
            c = aba_cx_gde['I3'].value
         elif dados_aba1['opcoes13'] == 'Inox 304':
            c = aba_cx_gde['I4'].value
         elif dados_aba1['opcoes13'] == 'Inox 316':
            c = aba_cx_gde['I5'].value
         elif dados_aba1['opcoes13'] == 'Inox 316L':
            c = aba_cx_gde['I6'].value

         # Tampão roscado
         if ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"':
            d = aba_cx_gde['L3'].value
         elif ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"':
            d = aba_cx_gde['L4'].value
         elif ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5':
            d = aba_cx_gde['L5'].value
         else:
            d = 'N/A'

         # Parafuso fixação M6x20
         if 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Aço carbono':
            e = aba_cx_gde['O3'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 304':
            e = aba_cx_gde['O4'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316':
            e = aba_cx_gde['O5'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316L':
            e = aba_cx_gde['O6'].value

         # Parafuso fixação triho
         if 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Aço carbono':
            f = aba_cx_gde['R3'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 304':
            f = aba_cx_gde['R4'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316':
            f = aba_cx_gde['R5'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316L':
            f = aba_cx_gde['R6'].value

         # Ilho de borracha
         if ('opcoes7' + str(s + 10)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 10)] == 'Área Segura':
            g = aba_cx_gde['U3'].value
         elif ('opcoes7' + str(s + 8)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 8)] == 'Sim':
            g = aba_cx_gde['U4'].value
         elif ('opcoes7' + str(s + 10)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 10)] == 'Ex-p':
            g = aba_cx_gde['U4'].value
         elif ('opcoes7' + str(s + 10)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 10)] == 'Ex-e':
            g = aba_cx_gde['U4'].value

         # Caixa de ligação fundida
         if 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 304' and dados_aba7['opcoes7' + str(s)] == 2 and dados_aba7['opcoes7' + str(s + 2)] == 'Sem rosca' and dados_aba7['opcoes7' + str(s + 3)] == '0':
            h = aba_cx_gde['AA3'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 316' and dados_aba7['opcoes7' + str(s)] == 2 and dados_aba7['opcoes7' + str(s + 2)] == 'Sem rosca' and dados_aba7['opcoes7' + str(s + 3)] == '0':
            h = aba_cx_gde['AA4'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s)] == 2 and dados_aba7['opcoes7' + str(s + 2)] == 'Sem rosca' and dados_aba7['opcoes7' + str(s + 3)] == '0':
            h = aba_cx_gde['AA5'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 304' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            h = aba_cx_gde['AA6'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 316' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            h = aba_cx_gde['AA7'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            h = aba_cx_gde['AA8'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 304' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            h = aba_cx_gde['AA9'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 316' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            h = aba_cx_gde['AA10'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            h = aba_cx_gde['AA11'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 304' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            h = aba_cx_gde['AA12'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 316' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            h = aba_cx_gde['AA13'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            h = aba_cx_gde['AA14'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 304' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            h = aba_cx_gde['AA15'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 316' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            h = aba_cx_gde['AA16'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            h = aba_cx_gde['AA17'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 304' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            h = aba_cx_gde['AA18'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 316' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            h = aba_cx_gde['AA19'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            h = aba_cx_gde['AA20'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 304' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            h = aba_cx_gde['AA21'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 316' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            h = aba_cx_gde['AA22'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            h = aba_cx_gde['AA23'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 304' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            h = aba_cx_gde['AA24'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 316' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            h = aba_cx_gde['AA25'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            h = aba_cx_gde['AA26'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 304' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            h = aba_cx_gde['AA27'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 316' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            h = aba_cx_gde['AA28'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            h = aba_cx_gde['AA29'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 304' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            h = aba_cx_gde['AA30'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Inox 316' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            h = aba_cx_gde['AA31'].value
         elif 'opcoes72' in dados_aba7 and ('opcoes7' + str(s)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and ('opcoes7' + str(s + 3)) in dados_aba7 and dados_aba7['opcoes72'] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s)] == 1 and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            h = aba_cx_gde['AA32'].value

         # Conj. placa cega s/ parafusos
         if ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s + 2)] == 'Sem rosca' and dados_aba7['opcoes7' + str(s + 3)] == '0':
            i = aba_cx_gde['AE3'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi6Cu4)' and dados_aba7['opcoes7' + str(s + 2)] == 'Sem rosca' and dados_aba7['opcoes7' + str(s + 3)] == '0':
            i = aba_cx_gde['AE4'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi10Mg)' and dados_aba7['opcoes7' + str(s + 2)] == 'Sem rosca' and dados_aba7['opcoes7' + str(s + 3)] == '0':
            i = aba_cx_gde['A5'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 316' and dados_aba7['opcoes7' + str(s + 2)] == 'Sem rosca' and dados_aba7['opcoes7' + str(s + 3)] == '0':
            i = aba_cx_gde['AE6'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 304' and dados_aba7['opcoes7' + str(s + 2)] == 'Sem rosca' and dados_aba7['opcoes7' + str(s + 3)] == '0':
            i = aba_cx_gde['AE7'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            i = aba_cx_gde['AE8'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            i = aba_cx_gde['AE9'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            i = aba_cx_gde['AE10'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi6Cu4)' and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            i = aba_cx_gde['AE11'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi6Cu4)' and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            i = aba_cx_gde['AE12'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi6Cu4)' and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            i = aba_cx_gde['AE13'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi10Mg)' and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            i = aba_cx_gde['AE14'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi10Mg)' and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            i = aba_cx_gde['AE15'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi10Mg)' and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            i = aba_cx_gde['AE16'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 316' and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            i = aba_cx_gde['AE17'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 316' and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            i = aba_cx_gde['AE18'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 316' and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            i = aba_cx_gde['AE19'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 304' and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            i = aba_cx_gde['AE20'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 304' and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            i = aba_cx_gde['AE21'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 304' and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '1':
            i = aba_cx_gde['AE22'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            i = aba_cx_gde['AE23'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            i = aba_cx_gde['AE24'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            i = aba_cx_gde['AE25'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi6Cu4)' and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            i = aba_cx_gde['AE26'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi6Cu4)' and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            i = aba_cx_gde['AE27'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi6Cu4)' and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            i = aba_cx_gde['AE28'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi10Mg)' and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            i = aba_cx_gde['AE29'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi10Mg)' and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            i = aba_cx_gde['AE30'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi10Mg)' and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            i = aba_cx_gde['AE31'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 316' and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            i = aba_cx_gde['AE32'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 316' and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            i = aba_cx_gde['AE33'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 316' and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            i = aba_cx_gde['AE34'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 304' and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            i = aba_cx_gde['AE35'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 304' and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            i = aba_cx_gde['AE36'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 304' and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '2':
            i = aba_cx_gde['AE37'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            i = aba_cx_gde['AE38'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            i = aba_cx_gde['AE39'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Ferro fundido' and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            i = aba_cx_gde['AE40'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi6Cu4)' and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            i = aba_cx_gde['AE41'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi6Cu4)' and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            i = aba_cx_gde['AE42'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi6Cu4)' and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            i = aba_cx_gde['AE43'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi10Mg)' and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            i = aba_cx_gde['AE44'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi10Mg)' and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            i = aba_cx_gde['AE45'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Alumínio (AlSi10Mg)' and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            i = aba_cx_gde['AE46'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 316' and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            i = aba_cx_gde['AE47'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 316' and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            i = aba_cx_gde['AE48'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 316' and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            i = aba_cx_gde['AE49'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 304' and dados_aba7['opcoes7' + str(s + 2)] == 'G 1"' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            i = aba_cx_gde['AE50'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 304' and dados_aba7['opcoes7' + str(s + 2)] == 'NPT 1"' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            i = aba_cx_gde['AE51'].value
         elif ('opcoes7' + str(s + 3)) in dados_aba7 and ('opcoes7' + str(s + 4)) in dados_aba7 and ('opcoes7' + str(s + 2)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 4)] == 'Inox 304' and dados_aba7['opcoes7' + str(s + 2)] == 'M20x1,5' and dados_aba7['opcoes7' + str(s + 3)] == '3':
            i = aba_cx_gde['AE52'].value
         elif ('opcoes7' + str(s)) in dados_aba7 and dados_aba7['opcoes7' + str(s)] == 1:
            i = 'N/A'

         # Identificador trilho
         j = aba_trilhos['D3'].value

         # Tag X* (1)
         if ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X1':
            k1 = aba_trilhos['B2'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X2':
            k1 = aba_trilhos['B3'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X3':
            k1 = aba_trilhos['B4'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X4':
            k1 = aba_trilhos['B5'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X5':
            k1 = aba_trilhos['B6'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X6':
            k1 = aba_trilhos['B7'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X7':
            k1 = aba_trilhos['B8'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X8':
            k1 = aba_trilhos['B9'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X9':
            k1 = aba_trilhos['B10'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X10':
            k1 = aba_trilhos['B11'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X11':
            k1 = aba_trilhos['B12'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X12':
            k1 = aba_trilhos['B13'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X13':
            k1 = aba_trilhos['B14'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X14':
            k1 = aba_trilhos['B15'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X15':
            k1 = aba_trilhos['B16'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X16':
            k1 = aba_trilhos['B17'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X17':
            k1 = aba_trilhos['B18'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X18':
            k1 = aba_trilhos['B19'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X19':
            k1 = aba_trilhos['B20'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X20':
            k1 = aba_trilhos['B21'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X21':
            k1 = aba_trilhos['B22'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X22':
            k1 = aba_trilhos['B23'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X23':
            k1 = aba_trilhos['B24'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X24':
            k1 = aba_trilhos['B25'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X25':
            k1 = aba_trilhos['B26'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X26':
            k1 = aba_trilhos['B27'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X27':
            k1 = aba_trilhos['B28'].value
         elif ('opcoes7' + str(s + 6)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 6)] == 'X28':
            k1 = aba_trilhos['B29'].value

         # Tag X* (2)
         if ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X1':
            k2 = aba_trilhos['B2'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X2':
            k2 = aba_trilhos['B3'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X3':
            k2 = aba_trilhos['B4'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X4':
            k2 = aba_trilhos['B5'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X5':
            k2 = aba_trilhos['B6'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X6':
            k2 = aba_trilhos['B7'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X7':
            k2 = aba_trilhos['B8'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X8':
            k2 = aba_trilhos['B9'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X9':
            k2 = aba_trilhos['B10'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X10':
            k2 = aba_trilhos['B11'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X11':
            k2 = aba_trilhos['B12'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X12':
            k2 = aba_trilhos['B13'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X13':
            k2 = aba_trilhos['B14'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X14':
            k2 = aba_trilhos['B15'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X15':
            k2 = aba_trilhos['B16'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X16':
            k2 = aba_trilhos['B17'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X17':
            k2 = aba_trilhos['B18'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X18':
            k2 = aba_trilhos['B19'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X19':
            k2 = aba_trilhos['B20'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X20':
            k2 = aba_trilhos['B21'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X21':
            k2 = aba_trilhos['B22'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X22':
            k2 = aba_trilhos['B23'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X23':
            k2 = aba_trilhos['B24'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X24':
            k2 = aba_trilhos['B25'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X25':
            k2 = aba_trilhos['B26'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X26':
            k2 = aba_trilhos['B27'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X27':
            k2 = aba_trilhos['B28'].value
         elif ('opcoes7' + str(s + 7)) in dados_aba7 and dados_aba7['opcoes7' + str(s + 7)] == 'X28':
            k2 = aba_trilhos['B29'].value
         
         # Conj. aterramento
         if 'opcoes73' in dados_aba7 and dados_aba7['opcoes73'] == 'Com aterramento':
            if 'opcoes13' in dados_aba1 and ('opcoes7' + str(s)) in dados_aba7 and dados_aba1['opcoes13'] == 'Aço carbono' and dados_aba7['opcoes7' + str(s)] == 1:
                  l = aba_cx_med['AL3'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes7' + str(s)) in dados_aba7 and dados_aba1['opcoes13'] == 'Inox 304' and dados_aba7['opcoes7' + str(s)] == 1:
                  l = aba_cx_med['AL4'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes7' + str(s)) in dados_aba7 and dados_aba1['opcoes13'] == 'Inox 316' and dados_aba7['opcoes7' + str(s)] == 1:
                  l = aba_cx_med['AL5'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes7' + str(s)) in dados_aba7 and dados_aba1['opcoes13'] == 'Inox 316L' and dados_aba7['opcoes7' + str(s)] == 1:
                  l = aba_cx_med['AL6'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes7' + str(s)) in dados_aba7 and dados_aba1['opcoes13'] == 'Aço carbono' and dados_aba7['opcoes7' + str(s)] == 2:
                  l = aba_cx_med['AL7'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes7' + str(s)) in dados_aba7 and dados_aba1['opcoes13'] == 'Inox 304' and dados_aba7['opcoes7' + str(s)] == 2:
                  l = aba_cx_med['AL8'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes7' + str(s)) in dados_aba7 and dados_aba1['opcoes13'] == 'Inox 316' and dados_aba7['opcoes7' + str(s)] == 2:
                  l = aba_cx_med['AL9'].value
            elif 'opcoes13' in dados_aba1 and ('opcoes7' + str(s)) in dados_aba7 and dados_aba1['opcoes13'] == 'Inox 316L' and dados_aba7['opcoes7' + str(s)] == 2:
                  l = aba_cx_med['AL10'].value
         elif 'opcoes73' in dados_aba7 and dados_aba7['opcoes73'] == 'Sem aterramento':
            l = 'N/A'

         # Caso seja requisitado 4 caixas
         if dados_aba7['opcoes71'] == '4':
            # Preenche o nome e campos da caixa 1
            if s == 4:
               aba_pa['A379'].value = str(dados_aba7['texto71'])
               aba_pa['E380'].value = a
               aba_pa['E382'].value = b
               aba_pa['E384'].value = c
               aba_pa['E393'].value = d
               aba_pa['E395'].value = e
               aba_pa['E389'].value = f
               aba_pa['E383'].value = g
               aba_pa['E381'].value = h
               aba_pa['E394'].value = i
               aba_pa['E390'].value = j
               aba_pa['E391'].value = k1
               aba_pa['E392'].value = k2
               aba_pa['E396'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba7['inc71'])
               pos_inic = int(dados_aba7['lt71'])
               celulas = 'D' + str(380)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(380 + m + 1)

               # Imprimindo as quantidades na planilha resposta
               if opcoes77.get() == '1':
                  aba_pa['H393'].value = 1
                  
               elif opcoes77.get() == '2':
                  aba_pa['H393'].value = 2
               
               if opcoes79.get() == '1':
                  aba_pa['H383'].value = 1
                  aba_pa['H389'].value = 2
                  aba_pa['H390'].value = 1
                  
               elif opcoes79.get() == '2':
                  aba_pa['H383'].value = 2
                  aba_pa['H389'].value = 4
                  aba_pa['H390'].value = 2

            # Preenche o nome e campos da caixa 2
            if s == 15:
               aba_pa['A397'].value = str(dados_aba7['texto72'])
               aba_pa['E398'].value = a
               aba_pa['E400'].value = b
               aba_pa['E402'].value = c
               aba_pa['E411'].value = d
               aba_pa['E413'].value = e
               aba_pa['E407'].value = f
               aba_pa['E401'].value = g
               aba_pa['E399'].value = h
               aba_pa['E412'].value = i
               aba_pa['E408'].value = j
               aba_pa['E409'].value = k1
               aba_pa['E410'].value = k2
               aba_pa['E414'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba7['inc72'])
               pos_inic = int(dados_aba7['lt72'])
               celulas = 'D' + str(398)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(398 + m + 1)

               # Imprimindo as quantidades na planilha resposta
               if opcoes717.get() == '1':
                  aba_pa['H411'].value = 1
                  
               elif opcoes717.get() == '2':
                  aba_pa['H411'].value = 2
               
               if opcoes719.get() == '1':
                  aba_pa['H401'].value = 1
                  aba_pa['H407'].value = 2
                  aba_pa['H408'].value = 1
                  
               elif opcoes719.get() == '2':
                  aba_pa['H401'].value = 2
                  aba_pa['H407'].value = 4
                  aba_pa['H408'].value = 2

            # Preenche o nome e campos da caixa 3
            if s == 26:
               aba_pa['A415'].value = str(dados_aba7['texto73'])
               aba_pa['E416'].value = a
               aba_pa['E418'].value = b
               aba_pa['E420'].value = c
               aba_pa['E429'].value = d
               aba_pa['E431'].value = e
               aba_pa['E425'].value = f
               aba_pa['E419'].value = g
               aba_pa['E417'].value = h
               aba_pa['E430'].value = i
               aba_pa['E426'].value = j
               aba_pa['E427'].value = k1
               aba_pa['E428'].value = k2
               aba_pa['E432'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba7['inc73'])
               pos_inic = int(dados_aba7['lt73'])
               celulas = 'D' + str(416)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(416 + m + 1)

               # Imprimindo as quantidades na planilha resposta
               if opcoes727.get() == '1':
                  aba_pa['H429'].value = 1
                  
               elif opcoes727.get() == '2':
                  aba_pa['H429'].value = 2
               
               if opcoes729.get() == '1':
                  aba_pa['H419'].value = 1
                  aba_pa['H425'].value = 2
                  aba_pa['H426'].value = 1
                  
               elif opcoes729.get() == '2':
                  aba_pa['H419'].value = 2
                  aba_pa['H425'].value = 4
                  aba_pa['H426'].value = 2

            # Preenche o nome e campos da caixa 4
            if s == 37:
               aba_pa['A433'].value = str(dados_aba7['texto74'])
               aba_pa['E434'].value = a
               aba_pa['E436'].value = b
               aba_pa['E438'].value = c
               aba_pa['E447'].value = d
               aba_pa['E449'].value = e
               aba_pa['E443'].value = f
               aba_pa['E437'].value = g
               aba_pa['E435'].value = h
               aba_pa['E448'].value = i
               aba_pa['E444'].value = j
               aba_pa['E445'].value = k1
               aba_pa['E446'].value = k2
               aba_pa['E450'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba7['inc74'])
               pos_inic = int(dados_aba7['lt74'])
               celulas = 'D' + str(434)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(434 + m + 1)

               # Imprimindo as quantidades na planilha resposta
               if opcoes737.get() == '1':
                  aba_pa['H447'].value = 1
                  
               elif opcoes737.get() == '2':
                  aba_pa['H447'].value = 2
               
               if opcoes739.get() == '1':
                  aba_pa['H437'].value = 1
                  aba_pa['H443'].value = 2
                  aba_pa['H444'].value = 1
                  
               elif opcoes739.get() == '2':
                  aba_pa['H437'].value = 2
                  aba_pa['H443'].value = 4
                  aba_pa['H444'].value = 2

         # Caso seja requisitado 3 caixas
         if dados_aba7['opcoes71'] == '3':
            # Preenche o nome e campos da caixa 1
            if s == 4:
               aba_pa['A379'].value = str(dados_aba7['texto71'])
               aba_pa['E380'].value = a
               aba_pa['E382'].value = b
               aba_pa['E384'].value = c
               aba_pa['E393'].value = d
               aba_pa['E395'].value = e
               aba_pa['E389'].value = f
               aba_pa['E383'].value = g
               aba_pa['E381'].value = h
               aba_pa['E394'].value = i
               aba_pa['E390'].value = j
               aba_pa['E391'].value = k1
               aba_pa['E392'].value = k2
               aba_pa['E396'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba7['inc71'])
               pos_inic = int(dados_aba7['lt71'])
               celulas = 'D' + str(380)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(380 + m + 1)

               # Imprimindo as quantidades na planilha resposta
               if opcoes77.get() == '1':
                  aba_pa['H393'].value = 1
                  
               elif opcoes77.get() == '2':
                  aba_pa['H393'].value = 2
               
               if opcoes79.get() == '1':
                  aba_pa['H383'].value = 1
                  aba_pa['H389'].value = 2
                  aba_pa['H390'].value = 1
                  
               elif opcoes79.get() == '2':
                  aba_pa['H383'].value = 2
                  aba_pa['H389'].value = 4
                  aba_pa['H390'].value = 2

            # Preenche o nome e campos da caixa 2
            if s == 15:
               aba_pa['A397'].value = str(dados_aba7['texto72'])
               aba_pa['E398'].value = a
               aba_pa['E400'].value = b
               aba_pa['E402'].value = c
               aba_pa['E411'].value = d
               aba_pa['E413'].value = e
               aba_pa['E407'].value = f
               aba_pa['E401'].value = g
               aba_pa['E399'].value = h
               aba_pa['E412'].value = i
               aba_pa['E408'].value = j
               aba_pa['E409'].value = k1
               aba_pa['E410'].value = k2
               aba_pa['E414'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba7['inc72'])
               pos_inic = int(dados_aba7['lt72'])
               celulas = 'D' + str(398)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(398 + m + 1)
               
               # Imprimindo as quantidades na planilha resposta
               if opcoes717.get() == '1':
                  aba_pa['H411'].value = 1
                  
               elif opcoes717.get() == '2':
                  aba_pa['H411'].value = 2
               
               if opcoes719.get() == '1':
                  aba_pa['H401'].value = 1
                  aba_pa['H407'].value = 2
                  aba_pa['H408'].value = 1
                  
               elif opcoes719.get() == '2':
                  aba_pa['H401'].value = 2
                  aba_pa['H407'].value = 4
                  aba_pa['H408'].value = 2

            # Preenche o nome e campos da caixa 3
            if s == 26:
               aba_pa['A415'].value = str(dados_aba7['texto73'])
               aba_pa['E416'].value = a
               aba_pa['E418'].value = b
               aba_pa['E420'].value = c
               aba_pa['E429'].value = d
               aba_pa['E431'].value = e
               aba_pa['E425'].value = f
               aba_pa['E419'].value = g
               aba_pa['E417'].value = h
               aba_pa['E430'].value = i
               aba_pa['E426'].value = j
               aba_pa['E427'].value = k1
               aba_pa['E428'].value = k2
               aba_pa['E432'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba7['inc73'])
               pos_inic = int(dados_aba7['lt73'])
               celulas = 'D' + str(416)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(416 + m + 1)

               # Imprimindo as quantidades na planilha resposta
               if opcoes727.get() == '1':
                  aba_pa['H429'].value = 1
                  
               elif opcoes727.get() == '2':
                  aba_pa['H429'].value = 2
               
               if opcoes729.get() == '1':
                  aba_pa['H419'].value = 1
                  aba_pa['H425'].value = 2
                  aba_pa['H426'].value = 1
                  
               elif opcoes729.get() == '2':
                  aba_pa['H419'].value = 2
                  aba_pa['H425'].value = 4
                  aba_pa['H426'].value = 2
            
         # Caso seja requisitado 2 caixas
         if dados_aba7['opcoes71'] == '2':
            # Preenche o nome e campos da caixa 1
            if s == 4:
               aba_pa['A379'].value = str(dados_aba7['texto71'])
               aba_pa['E380'].value = a
               aba_pa['E382'].value = b
               aba_pa['E384'].value = c
               aba_pa['E393'].value = d
               aba_pa['E395'].value = e
               aba_pa['E389'].value = f
               aba_pa['E383'].value = g
               aba_pa['E381'].value = h
               aba_pa['E394'].value = i
               aba_pa['E390'].value = j
               aba_pa['E391'].value = k1
               aba_pa['E392'].value = k2
               aba_pa['E396'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba7['inc71'])
               pos_inic = int(dados_aba7['lt71'])
               celulas = 'D' + str(380)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(380 + m + 1)

               # Imprimindo as quantidades na planilha resposta
               if opcoes77.get() == '1':
                  aba_pa['H393'].value = 1
                  
               elif opcoes77.get() == '2':
                  aba_pa['H393'].value = 2
               
               if opcoes79.get() == '1':
                  aba_pa['H383'].value = 1
                  aba_pa['H389'].value = 2
                  aba_pa['H390'].value = 1
                  
               elif opcoes79.get() == '2':
                  aba_pa['H383'].value = 2
                  aba_pa['H389'].value = 4
                  aba_pa['H390'].value = 2

            # Preenche o nome e campos da caixa 2
            if s == 15:
               aba_pa['A397'].value = str(dados_aba7['texto72'])
               aba_pa['E398'].value = a
               aba_pa['E400'].value = b
               aba_pa['E402'].value = c
               aba_pa['E411'].value = d
               aba_pa['E413'].value = e
               aba_pa['E407'].value = f
               aba_pa['E401'].value = g
               aba_pa['E399'].value = h
               aba_pa['E412'].value = i
               aba_pa['E408'].value = j
               aba_pa['E409'].value = k1
               aba_pa['E410'].value = k2
               aba_pa['E414'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba7['inc72'])
               pos_inic = int(dados_aba7['lt72'])
               celulas = 'D' + str(398)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(398 + m + 1)

               # Imprimindo as quantidades na planilha resposta
               if opcoes717.get() == '1':
                  aba_pa['H411'].value = 1
                  
               elif opcoes717.get() == '2':
                  aba_pa['H411'].value = 2
               
               if opcoes719.get() == '1':
                  aba_pa['H401'].value = 1
                  aba_pa['H407'].value = 2
                  aba_pa['H408'].value = 1
                  
               elif opcoes719.get() == '2':
                  aba_pa['H401'].value = 2
                  aba_pa['H407'].value = 4
                  aba_pa['H408'].value = 2

         # Caso seja requisitado 1 caixa
         if dados_aba7['opcoes71'] == '1':
            # Preenche o nome e campos da caixa 1
            if s == 4:
               aba_pa['A379'].value = str(dados_aba7['texto71'])
               aba_pa['E380'].value = a
               aba_pa['E382'].value = b
               aba_pa['E384'].value = c
               aba_pa['E393'].value = d
               aba_pa['E395'].value = e
               aba_pa['E389'].value = f
               aba_pa['E383'].value = g
               aba_pa['E381'].value = h
               aba_pa['E394'].value = i
               aba_pa['E390'].value = j
               aba_pa['E391'].value = k1
               aba_pa['E392'].value = k2
               aba_pa['E396'].value = l

               # Imprimindo as posições da lista técnica na planilha resposta
               incremento = int(dados_aba7['inc71'])
               pos_inic = int(dados_aba7['lt71'])
               celulas = 'D' + str(380)
               for m in range(17):
                     aba_pa[celulas].value = pos_inic + (m * incremento)
                     celulas = 'D' + str(380 + m + 1)

               # Imprimindo as quantidades na planilha resposta
               if opcoes77.get() == '1':
                  aba_pa['H393'].value = 1
                  
               elif opcoes77.get() == '2':
                  aba_pa['H393'].value = 2
               
               if opcoes79.get() == '1':
                  aba_pa['H383'].value = 1
                  aba_pa['H389'].value = 2
                  aba_pa['H390'].value = 1
                  
               elif opcoes79.get() == '2':
                  aba_pa['H383'].value = 2
                  aba_pa['H389'].value = 4
                  aba_pa['H390'].value = 2
         
         if dados_aba7['opcoes71'] == '0':
               pass
      
      
      # Regras da aba 9 (Refrigeração) =======================================================================================================================================
      # Regras para Termoresistências (Instalação no ar e água)
      if 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Aço carbono':
         conduite = aba_sensorTemp['B5'].value
         qtd_conduite = aba_sensorTemp['C5'].value
         luva = aba_sensorTemp['B6'].value
         qtd_luva = aba_sensorTemp['C6'].value
         conMacho = aba_sensorTemp['B7'].value
         qtd_conMacho = aba_sensorTemp['C7'].value
         abracadeira = aba_sensorTemp['B8'].value
         qtd_abracadeira = aba_sensorTemp['C8'].value
         parafuso = aba_sensorTemp['B9'].value
         qtd_parafuso = aba_sensorTemp['C9'].value
      elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 304':
         conduite = aba_sensorTemp['B12'].value
         qtd_conduite = aba_sensorTemp['C12'].value
         luva = aba_sensorTemp['B13'].value
         qtd_luva = aba_sensorTemp['C13'].value
         conMacho = aba_sensorTemp['B14'].value
         qtd_conMacho = aba_sensorTemp['C14'].value
         abracadeira = aba_sensorTemp['B15'].value
         qtd_abracadeira = aba_sensorTemp['C15'].value
         parafuso = aba_sensorTemp['B16'].value
         qtd_parafuso = aba_sensorTemp['C16'].value
      elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316':
         conduite = aba_sensorTemp['B19'].value
         qtd_conduite = aba_sensorTemp['C19'].value
         luva = aba_sensorTemp['B20'].value
         qtd_luva = aba_sensorTemp['C20'].value
         conMacho = aba_sensorTemp['B21'].value
         qtd_conMacho = aba_sensorTemp['C21'].value
         abracadeira = aba_sensorTemp['B22'].value
         qtd_abracadeira = aba_sensorTemp['C22'].value
         parafuso = aba_sensorTemp['B23'].value
         qtd_parafuso = aba_sensorTemp['C23'].value
      elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316L':
         conduite = aba_sensorTemp['B26'].value
         qtd_conduite = aba_sensorTemp['C26'].value
         luva = aba_sensorTemp['B27'].value
         qtd_luva = aba_sensorTemp['C27'].value
         conMacho = aba_sensorTemp['B28'].value
         qtd_conMacho = aba_sensorTemp['C28'].value
         abracadeira = aba_sensorTemp['B29'].value
         qtd_abracadeira = aba_sensorTemp['C29'].value
         parafuso = aba_sensorTemp['B30'].value
         qtd_parafuso = aba_sensorTemp['C30'].value

      # Imprimindo células para Termoresistências (Instalação no ar)
      aba_pa['E492'].value = conduite
      aba_pa['H492'].value = qtd_conduite * int(dados_aba9['opcoes91'])
      aba_pa['E493'].value = luva
      aba_pa['H493'].value = qtd_luva * int(dados_aba9['opcoes91'])
      aba_pa['E494'].value = conMacho
      aba_pa['H494'].value = qtd_conMacho * int(dados_aba9['opcoes91'])
      aba_pa['E495'].value = abracadeira
      aba_pa['H495'].value = qtd_abracadeira * int(dados_aba9['opcoes91'])
      aba_pa['E496'].value = parafuso
      aba_pa['H496'].value = qtd_parafuso * int(dados_aba9['opcoes91'])

      incremento = int(dados_aba9['inc91'])
      pos_inic = int(dados_aba9['lt91']) + incremento
      celulas = 'D' + str(491)
      for m in range(6):
         aba_pa[celulas].value = pos_inic + (m * incremento)
         celulas = 'D' + str(491 + m + 1)

      # Imprimindo células para Termoresistências (Instalação na água)
      aba_pa['E499'].value = conduite
      aba_pa['H499'].value = qtd_conduite * int(dados_aba9['opcoes92'])
      aba_pa['E500'].value = luva
      aba_pa['H500'].value = qtd_luva * int(dados_aba9['opcoes92'])
      aba_pa['E501'].value = conMacho
      aba_pa['H501'].value = qtd_conMacho * int(dados_aba9['opcoes92'])
      aba_pa['E502'].value = abracadeira
      aba_pa['H502'].value = qtd_abracadeira * int(dados_aba9['opcoes92'])
      aba_pa['E503'].value = parafuso
      aba_pa['H503'].value = qtd_parafuso * int(dados_aba9['opcoes92'])

      incremento = int(dados_aba9['inc91'])
      pos_inic = int(aba_pa['D496'].value) + incremento
      celulas = 'D' + str(498)
      for m in range(6):
         aba_pa[celulas].value = pos_inic + (m * incremento)
         celulas = 'D' + str(498 + m + 1)
      
      # Regras para Termômetros (Instalação no ar)
      if 'opcoes95' in dados_aba9 and dados_aba9['opcoes95'] == 'Sem contato':
         if 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Aço carbono':
            fixTermom = aba_termometros['D7'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 304':
            fixTermom = aba_termometros['D8'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316':
            fixTermom = aba_termometros['D9'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316L':
            fixTermom = aba_termometros['D10'].value
      
      elif 'opcoes95' in dados_aba9 and dados_aba9['opcoes95'] == 'Com contato':
         if 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Aço carbono':
            fixTermom = aba_termometros['D15'].value
            qtd_fixTermom = int(dados_aba9['opcoes96'])
            conduite = aba_sensorTemp['B5'].value
            qtd_conduite = aba_sensorTemp['C5'].value
            luva = aba_sensorTemp['B6'].value
            qtd_luva = aba_sensorTemp['C6'].value
            conMacho = aba_sensorTemp['B7'].value
            qtd_conMacho = aba_sensorTemp['C7'].value
            abracadeira = aba_sensorTemp['B8'].value
            qtd_abracadeira = aba_sensorTemp['C8'].value
            parafuso = aba_sensorTemp['B9'].value
            qtd_parafuso = aba_sensorTemp['C9'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 304':
            fixTermom = aba_termometros['D16'].value
            qtd_fixTermom = int(dados_aba9['opcoes96'])
            conduite = aba_sensorTemp['B12'].value
            qtd_conduite = aba_sensorTemp['C12'].value
            luva = aba_sensorTemp['B13'].value
            qtd_luva = aba_sensorTemp['C13'].value
            conMacho = aba_sensorTemp['B14'].value
            qtd_conMacho = aba_sensorTemp['C14'].value
            abracadeira = aba_sensorTemp['B15'].value
            qtd_abracadeira = aba_sensorTemp['C15'].value
            parafuso = aba_sensorTemp['B16'].value
            qtd_parafuso = aba_sensorTemp['C16'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316':
            fixTermom = aba_termometros['D17'].value
            qtd_fixTermom = int(dados_aba9['opcoes96'])
            conduite = aba_sensorTemp['B19'].value
            qtd_conduite = aba_sensorTemp['C19'].value
            luva = aba_sensorTemp['B20'].value
            qtd_luva = aba_sensorTemp['C20'].value
            conMacho = aba_sensorTemp['B21'].value
            qtd_conMacho = aba_sensorTemp['C21'].value
            abracadeira = aba_sensorTemp['B22'].value
            qtd_abracadeira = aba_sensorTemp['C22'].value
            parafuso = aba_sensorTemp['B23'].value
            qtd_parafuso = aba_sensorTemp['C23'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316L':
            fixTermom = aba_termometros['D18'].value
            qtd_fixTermom = int(dados_aba9['opcoes96'])
            conduite = aba_sensorTemp['B26'].value
            qtd_conduite = aba_sensorTemp['C26'].value
            luva = aba_sensorTemp['B27'].value
            qtd_luva = aba_sensorTemp['C27'].value
            conMacho = aba_sensorTemp['B28'].value
            qtd_conMacho = aba_sensorTemp['C28'].value
            abracadeira = aba_sensorTemp['B29'].value
            qtd_abracadeira = aba_sensorTemp['C29'].value
            parafuso = aba_sensorTemp['B30'].value
            qtd_parafuso = aba_sensorTemp['C30'].value
         
      elif'opcoes95' in dados_aba9 and dados_aba9['opcoes95'] == 'Sem termômetro':
         fixTermom = 'N/A'
         qtd_fixTermom = 'N/A'
         conduite = 'N/A'
         qtd_conduite = 'N/A'
         luva = 'N/A'
         qtd_luva = 'N/A'
         conMacho = 'N/A'
         qtd_conMacho = 'N/A'
         abracadeira = 'N/A'
         qtd_abracadeira = 'N/A'
         parafuso = 'N/A'
         qtd_parafuso = 'N/A'

      # Imprimindo células para Termômetros (Instalação no ar)
      aba_pa['E507'].value = fixTermom
      aba_pa['H507'].value = qtd_fixTermom
      aba_pa['E508'].value = conduite
      aba_pa['H508'].value = qtd_conduite * int(dados_aba9['opcoes96'])
      aba_pa['E509'].value = luva
      aba_pa['H509'].value = qtd_luva * int(dados_aba9['opcoes96'])
      aba_pa['E510'].value = conMacho
      aba_pa['H510'].value = qtd_conMacho * int(dados_aba9['opcoes96'])
      aba_pa['E511'].value = abracadeira
      aba_pa['H511'].value = qtd_abracadeira * int(dados_aba9['opcoes96'])
      aba_pa['E512'].value = parafuso
      aba_pa['H512'].value = qtd_parafuso * int(dados_aba9['opcoes96'])

      incremento = int(dados_aba9['inc91'])
      pos_inic = int(aba_pa['D503'].value) + incremento
      celulas = 'D' + str(506)
      for m in range(7):
         aba_pa[celulas].value = pos_inic + (m * incremento)
         celulas = 'D' + str(506 + m + 1)

      # Regras para Termômetros (Instalação na água)
      if 'opcoes97' in dados_aba9 and dados_aba9['opcoes97'] == 'Sem contato':
         if 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Aço carbono':
            fixTermom = aba_termometros['D7'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 304':
            fixTermom = aba_termometros['D8'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316':
            fixTermom = aba_termometros['D9'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316L':
            fixTermom = aba_termometros['D10'].value

      elif 'opcoes97' in dados_aba9 and dados_aba9['opcoes97'] == 'Com contato':
         if 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Aço carbono':
            fixTermom = aba_termometros['D15'].value
            qtd_fixTermom = int(dados_aba9['opcoes98'])
            conduite = aba_sensorTemp['B5'].value
            qtd_conduite = aba_sensorTemp['C5'].value
            luva = aba_sensorTemp['B6'].value
            qtd_luva = aba_sensorTemp['C6'].value
            conMacho = aba_sensorTemp['B7'].value
            qtd_conMacho = aba_sensorTemp['C7'].value
            abracadeira = aba_sensorTemp['B8'].value
            qtd_abracadeira = aba_sensorTemp['C8'].value
            parafuso = aba_sensorTemp['B9'].value
            qtd_parafuso = aba_sensorTemp['C9'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 304':
            fixTermom = aba_termometros['D16'].value
            qtd_fixTermom = int(dados_aba9['opcoes98'])
            conduite = aba_sensorTemp['B12'].value
            qtd_conduite = aba_sensorTemp['C12'].value
            luva = aba_sensorTemp['B13'].value
            qtd_luva = aba_sensorTemp['C13'].value
            conMacho = aba_sensorTemp['B14'].value
            qtd_conMacho = aba_sensorTemp['C14'].value
            abracadeira = aba_sensorTemp['B15'].value
            qtd_abracadeira = aba_sensorTemp['C15'].value
            parafuso = aba_sensorTemp['B16'].value
            qtd_parafuso = aba_sensorTemp['C16'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316':
            fixTermom = aba_termometros['D17'].value
            qtd_fixTermom = int(dados_aba9['opcoes98'])
            conduite = aba_sensorTemp['B19'].value
            qtd_conduite = aba_sensorTemp['C19'].value
            luva = aba_sensorTemp['B20'].value
            qtd_luva = aba_sensorTemp['C20'].value
            conMacho = aba_sensorTemp['B21'].value
            qtd_conMacho = aba_sensorTemp['C21'].value
            abracadeira = aba_sensorTemp['B22'].value
            qtd_abracadeira = aba_sensorTemp['C22'].value
            parafuso = aba_sensorTemp['B23'].value
            qtd_parafuso = aba_sensorTemp['C23'].value
         elif 'opcoes13' in dados_aba1 and dados_aba1['opcoes13'] == 'Inox 316L':
            fixTermom = aba_termometros['D18'].value
            qtd_fixTermom = int(dados_aba9['opcoes98'])
            conduite = aba_sensorTemp['B26'].value
            qtd_conduite = aba_sensorTemp['C26'].value
            luva = aba_sensorTemp['B27'].value
            qtd_luva = aba_sensorTemp['C27'].value
            conMacho = aba_sensorTemp['B28'].value
            qtd_conMacho = aba_sensorTemp['C28'].value
            abracadeira = aba_sensorTemp['B29'].value
            qtd_abracadeira = aba_sensorTemp['C29'].value
            parafuso = aba_sensorTemp['B30'].value
            qtd_parafuso = aba_sensorTemp['C30'].value

      elif'opcoes97' in dados_aba9 and dados_aba9['opcoes97'] == 'Sem termômetro':
         fixTermom = 'N/A'
         qtd_fixTermom = 'N/A'
         conduite = 'N/A'
         qtd_conduite = 'N/A'
         luva = 'N/A'
         qtd_luva = 'N/A'
         conMacho = 'N/A'
         qtd_conMacho = 'N/A'
         abracadeira = 'N/A'
         qtd_abracadeira = 'N/A'
         parafuso = 'N/A'
         qtd_parafuso = 'N/A'
      
      # Imprimindo células para Termômetros (Instalação na água)
      aba_pa['E515'].value = fixTermom
      aba_pa['H515'].value = qtd_fixTermom
      aba_pa['E516'].value = conduite
      aba_pa['H516'].value = qtd_conduite * int(dados_aba9['opcoes98'])
      aba_pa['E517'].value = luva
      aba_pa['H517'].value = qtd_luva * int(dados_aba9['opcoes98'])
      aba_pa['E518'].value = conMacho
      aba_pa['H518'].value = qtd_conMacho * int(dados_aba9['opcoes98'])
      aba_pa['E519'].value = abracadeira
      aba_pa['H519'].value = qtd_abracadeira * int(dados_aba9['opcoes98'])
      aba_pa['E520'].value = parafuso
      aba_pa['H520'].value = qtd_parafuso * int(dados_aba9['opcoes98'])

      incremento = int(dados_aba9['inc91'])
      pos_inic = int(aba_pa['D512'].value) + incremento
      celulas = 'D' + str(514)
      for m in range(7):
         aba_pa[celulas].value = pos_inic + (m * incremento)
         celulas = 'D' + str(514 + m + 1)

      # Regras da aba 10 (Comp. Gerais/Avulsos) =======================================================================================================================================
      # Tag X*
      if 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X1':
         tagX_aba10 = aba_trilhos['B2'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X2':
         tagX_aba10 = aba_trilhos['B3'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X3':
         tagX_aba10 = aba_trilhos['B4'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X4':
         tagX_aba10 = aba_trilhos['B5'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X5':
         tagX_aba10 = aba_trilhos['B6'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X6':
         tagX_aba10 = aba_trilhos['B7'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X7':
         tagX_aba10 = aba_trilhos['B8'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X8':
         tagX_aba10 = aba_trilhos['B9'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X9':
         tagX_aba10 = aba_trilhos['B10'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X10':
         tagX_aba10 = aba_trilhos['B11'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X11':
         tagX_aba10 = aba_trilhos['B12'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X12':
         tagX_aba10 = aba_trilhos['B13'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X13':
         tagX_aba10 = aba_trilhos['B14'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X14':
         tagX_aba10 = aba_trilhos['B15'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X15':
         tagX_aba10 = aba_trilhos['B16'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X16':
         tagX_aba10 = aba_trilhos['B17'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X17':
         tagX_aba10 = aba_trilhos['B18'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X18':
         tagX_aba10 = aba_trilhos['B19'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X19':
         tagX_aba10 = aba_trilhos['B20'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X20':
         tagX_aba10 = aba_trilhos['B21'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X21':
         tagX_aba10 = aba_trilhos['B22'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X22':
         tagX_aba10 = aba_trilhos['B23'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X23':
         tagX_aba10 = aba_trilhos['B24'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X24':
         tagX_aba10 = aba_trilhos['B25'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X25':
         tagX_aba10 = aba_trilhos['B26'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X26':
         tagX_aba10 = aba_trilhos['B27'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X27':
         tagX_aba10 = aba_trilhos['B28'].value
      elif 'opcoes102' in dados_aba10 and dados_aba10['opcoes102'] == 'X28':
         tagX_aba10 = aba_trilhos['B29'].value

      # Rele auxiliar
      if 'opcoes101' in dados_aba10 and dados_aba10['opcoes101'] == '24 Vcc':
         releAux = aba_compGerais['B3'].value
      elif 'opcoes101' in dados_aba10 and dados_aba10['opcoes101'] == '125 Vcc':
         releAux = aba_compGerais['B4'].value
      elif 'opcoes101' in dados_aba10 and dados_aba10['opcoes101'] == '110 Vca':
         releAux = aba_compGerais['B5'].value
      elif 'opcoes101' in dados_aba10 and dados_aba10['opcoes101'] == '220 Vca':
         releAux = aba_compGerais['B6'].value
      elif 'opcoes101' in dados_aba10 and dados_aba10['opcoes101'] == 'Sem freio':
         releAux = 'N/A'
      
      # Imprime o campo TagX* da planilha
      aba_pa['E523'].value = tagX_aba10

      # Imprime o campo Rele auxiliar da planilha
      aba_pa['E524'].value = releAux

      incremento = int(dados_aba10['inc101'])
      pos_inic = int(dados_aba10['lt101'])
      celulas = 'D' + str(523)
      for m in range(2):
         aba_pa[celulas].value = pos_inic + (m * incremento)
         celulas = 'D' + str(523 + m + 1)

      # Imprime o campo "texto longo" da planilha
      aba_pa['A530'].value = dados_aba10['textolongo101']

      # Linhas a serem ocultadas por padrão
      try:
         aba_pa.row_dimensions.group(25, 28, hidden=True)
         aba_pa.row_dimensions.group(91, 140, hidden=True)
         aba_pa.row_dimensions.group(191, 241, hidden=True)
         aba_pa.row_dimensions.group(451, 487, hidden=True)
         aba_pa.row_dimensions.group(525, 528, hidden=True)
      except:
         print('Erro ao tentar excluir linhas inutilizadas da planilha.')
         erro = True

      # Ocultando as linhas não utilizadas da planilha resposta
      try:
         # Estator
         if aba_pa['D'+str(11)].value == '*':
            aba_pa.row_dimensions.group(10, 14, hidden=True)
         elif aba_pa['D'+str(16)].value == '*':
            aba_pa.row_dimensions.group(15, 19, hidden=True)
         elif aba_pa['D'+str(11)].value == '*' and aba_pa['D'+str(16)].value == '*':
            aba_pa.row_dimensions.group(10, 19, hidden=True)

         if aba_pa['D'+str(31)].value != '*':
            aba_pa.row_dimensions.group(34, 39, hidden=True)
         elif aba_pa['D'+str(34)].value != '*':
            aba_pa.row_dimensions.group(31, 33, hidden=True)
            aba_pa.row_dimensions.group(36, 39, hidden=True)
         elif aba_pa['D'+str(36)].value != '*':
            aba_pa.row_dimensions.group(31, 35, hidden=True)

         # Caixa pequena
         if aba_pa['D'+str(246)].value == '*' and aba_pa['D'+str(261)].value == '*' and aba_pa['D'+str(276)].value == '*' and aba_pa['D'+str(291)].value == '*':
            aba_pa.row_dimensions.group(243, 304, hidden=True)
         if aba_pa['D'+str(261)].value == '*' and aba_pa['D'+str(276)].value == '*' and aba_pa['D'+str(291)].value == '*':
            aba_pa.row_dimensions.group(260, 304, hidden=True)
         if aba_pa['D'+str(276)].value == '*' and aba_pa['D'+str(291)].value == '*':
            aba_pa.row_dimensions.group(275, 304, hidden=True)
         if aba_pa['D'+str(291)].value == '*':
            aba_pa.row_dimensions.group(290, 304, hidden=True)

         # Caixa média
         if aba_pa['D'+str(307)].value == '*' and aba_pa['D'+str(325)].value == '*' and aba_pa['D'+str(343)].value == '*' and aba_pa['D'+str(361)].value == '*':
            aba_pa.row_dimensions.group(305, 377, hidden=True)
         if aba_pa['D'+str(325)].value == '*' and aba_pa['D'+str(343)].value == '*' and aba_pa['D'+str(361)].value == '*':
            aba_pa.row_dimensions.group(325, 377, hidden=True)
         if aba_pa['D'+str(343)].value == '*' and aba_pa['D'+str(361)].value == '*':
            aba_pa.row_dimensions.group(343, 377, hidden=True)
         if aba_pa['D'+str(361)].value == '*':
            aba_pa.row_dimensions.group(361, 377, hidden=True)

         # Caixa grande
         if aba_pa['D'+str(380)].value == '*' and aba_pa['D'+str(398)].value == '*' and aba_pa['D'+str(416)].value == '*' and aba_pa['D'+str(434)].value == '*':
            aba_pa.row_dimensions.group(378, 450, hidden=True)
         if aba_pa['D'+str(398)].value == '*' and aba_pa['D'+str(416)].value == '*' and aba_pa['D'+str(434)].value == '*':
            aba_pa.row_dimensions.group(397, 450, hidden=True)
         if aba_pa['D'+str(416)].value == '*' and aba_pa['D'+str(434)].value == '*':
            aba_pa.row_dimensions.group(415, 450, hidden=True)
         if aba_pa['D'+str(434)].value == '*':
            aba_pa.row_dimensions.group(433, 450, hidden=True)

      except:
         print('Erro ao tentar ocultar as linhas inutilizadas da planilha.')
         erro = True

      # Salvando a planilha Excel
      pa.save(os.getcwd() + "\Armazenamento" + "\planilha_resposta_mat_" + texto12.get() + ".xlsx")

      # Abre o Excel
      excel = client.Dispatch("Excel.Application")
      
      # Lê o arquivo Excel
      sheets = excel.Workbooks.Open(os.getcwd() + "\Armazenamento" + "\planilha_resposta_mat_" + texto12.get() + ".xlsx")
      work_sheets = sheets.Worksheets[0]
      
      # Converte para PDF
      work_sheets.ExportAsFixedFormat(0, pasta_escolhida + '/planilha_resposta_mat_' + texto12.get() + '.pdf')

      #Fechando a planilha resposta em excel
      pa.close()

      # Deleta popup de sucesso
      def deleta_popup_sucesso():
         if erro == False:
            res.destroy()
      
      # Cria janela de aviso de programa executado com sucesso
      if erro == False:
         res = Toplevel(root)
         res.geometry("500x250")
         res.title("Programa executado com sucesso!")

         if var.get() == 1:
            Label(res, text='Acesse a pasta escolhida para localizar\n o arquivo "planilha_resposta.xlsx" com os resultados.', font= '20').place(x=60,y=80)
         elif var.get()== 2:
            Label(res, text='Access the chosen folder to find\n the file "planilha_resposta.xlsx" with the results.', font= '20').place(x=60,y=80)

         Button(res, text='OK', font='20', command=deleta_popup_sucesso).place(x=220, y=150)


def escolha_idioma():
   global info_texto, btn_imprimir, \
      opcoes11, opcoes12, opcoes13, opcoes14, texto11, texto12, texto13, texto14, \
      opcoes21, opcoes22, opcoes23, opcoes24, varcb1, varcb2, lt21, inc21, \
      opcoes31, opcoes32, opcoes33, opcoes34, opcoes35, opcoes36, opcoes37, opcoes38, opcoes39, opcoes310, opcoes311, opcoes312, opcoes313, opcoes314, opcoes315, opcoes316, opcoes317, opcoes318, opcoes319, opcoes320, opcoes321, opcoes322, lt31, inc31, \
      opcoes41, opcoes42, opcoes43, opcoes44, opcoes45, opcoes46, opcoes47, opcoes48, opcoes49, opcoes410, opcoes411, opcoes412, opcoes413, opcoes414, opcoes415, opcoes416, opcoes417, opcoes418, opcoes419, opcoes420, opcoes421, opcoes422, lt41, inc41, \
      opcoes51, opcoes52, opcoes53, opcoes54, opcoes55, opcoes56, opcoes57, opcoes58, opcoes59, opcoes510, opcoes511, opcoes512, opcoes513, opcoes514, opcoes515, opcoes516, opcoes517, opcoes518, opcoes519, opcoes520, opcoes521, opcoes522, opcoes523, opcoes524, opcoes525, opcoes526, opcoes527, opcoes528, opcoes529, opcoes530, opcoes531, texto51, texto52, texto53, texto54, lt51, lt52, lt53, lt54, inc51, inc52, inc53, inc54, \
      opcoes61, opcoes62, opcoes63, opcoes64, opcoes65, opcoes66, opcoes67, opcoes68, opcoes69, opcoes610, opcoes611, opcoes612, opcoes613, opcoes614, opcoes615, opcoes616, opcoes617, opcoes618, opcoes619, opcoes620, opcoes621, opcoes622, opcoes623, opcoes624, opcoes625, opcoes626, opcoes627, opcoes628, opcoes629, opcoes630, opcoes631, opcoes632, opcoes633, opcoes634, opcoes635, opcoes636, opcoes637, opcoes638, opcoes639, opcoes640, opcoes641, opcoes642, opcoes643, opcoes644, opcoes645, opcoes646, opcoes647, texto61, texto62, texto63, texto64, lt61, lt62, lt63, lt64, inc61, inc62, inc63, inc64, \
      opcoes71, opcoes72, opcoes73, opcoes74, opcoes75, opcoes76, opcoes77, opcoes78, opcoes79, opcoes710, opcoes711, opcoes712, opcoes713, opcoes714, opcoes715, opcoes716, opcoes717, opcoes718, opcoes719, opcoes720, opcoes721, opcoes722, opcoes723, opcoes724, opcoes725, opcoes726, opcoes727, opcoes728, opcoes729, opcoes730, opcoes731, opcoes732, opcoes733, opcoes734, opcoes735, opcoes736, opcoes737, opcoes738, opcoes739, opcoes740, opcoes741, opcoes742, opcoes743, opcoes744, opcoes745, opcoes746, opcoes747, texto71, texto72, texto73, texto74, lt71, lt72, lt73, lt74, inc71, inc72, inc73, inc74, \
      opcoes81, opcoes82, opcoes83, opcoes84, opcoes85, opcoes86, opcoes87, opcoes88, opcoes89, opcoes810, opcoes811, opcoes812, opcoes813, opcoes814, opcoes815, opcoes816, opcoes817, texto81, texto82, lt81, inc81, \
      opcoes91, opcoes92, opcoes93, opcoes94, opcoes95, opcoes96, opcoes97, opcoes98, lt91, inc91, \
      opcoes101, opcoes102, textolongo101, lt101, inc101, \
      texto_padrao11, texto_padrao12, texto_padrao13, texto_padrao14, \
      texto_padrao_lt21, texto_padrao_lt31

   if var.get() == 1:

      info_texto.destroy()
      info_texto = Label(fr_info, text='Todos os campos devem ser preenchidos antes de imprimir os dados.', font= '20')
      info_texto.place(x=10, y=20, width=800)

      nb = ttk.Notebook(root)
      nb.place(x=0, y=81, width=largura, height=altura-130)

      # Conteúdo da Aba 1 pt ==================================================================================================
      aba1 = Frame(nb)
      nb.add(aba1, text='Informações do projeto')

      fr_info_proj = LabelFrame(aba1, borderwidth=1, relief='solid', text='  Informações:  ')
      fr_info_proj.place(x=5, y=5, width=450, height=200)

      def texto_temp11(e):
         if texto11.get() == 'Digitar nome do projeto...':
            texto11.delete(0, END)
         elif texto11.get() == '':
            texto11.insert(END, 'Digitar nome do projeto...')

      Label(fr_info_proj, text='Nome do projeto:').place(x=5, y=25)
      texto11 = Entry(fr_info_proj, takefocus = 0)
      texto11.insert(END, "Digitar nome do projeto...")
      texto11.place(x=200, y=25, width=200, height=20)
      texto11.bind("<FocusIn>", texto_temp11)
      texto11.bind("<FocusOut>", texto_temp11)

      def texto_temp12(e):
         if texto12.get() == "Digitar material...":
            texto12.delete(0, END)
         elif texto12.get() == '':
            texto12.insert(END, "Digitar material...")

      Label(fr_info_proj, text='Material do gerador/motor:').place(x=5, y=65)
      texto12 = Entry(fr_info_proj, takefocus = 0)
      texto12.insert(END, "Digitar material...")
      texto12.place(x=200, y=65, width=150, height=20)
      texto12.bind("<FocusIn>", texto_temp12)
      texto12.bind("<FocusOut>", texto_temp12)

      def texto_temp13(e):
         if texto13.get() == "Digitar login...":
            texto13.delete(0, END)
         elif texto13.get() == '':
            texto13.insert(END, "Digitar login...")

      Label(fr_info_proj, text='Login do projetista responsável:').place(x=5, y=105)
      texto13 = Entry(fr_info_proj, takefocus = 0)
      texto13.insert(END, "Digitar login...")
      texto13.place(x=200, y=105, width=150, height=20)
      texto13.bind("<FocusIn>", texto_temp13)
      texto13.bind("<FocusOut>", texto_temp13)

      def texto_temp14(e):
         if texto14.get() == "Digitar ordem...":
            texto14.delete(0, END)
         elif texto14.get() == '':
            texto14.insert(END, "Digitar ordem...")

      Label(fr_info_proj, text='Ordem de vendas:').place(x=5, y=145)
      texto14 = Entry(fr_info_proj, takefocus = 0)
      texto14.insert(END, "Digitar ordem...")
      texto14.place(x=200, y=145, width=150, height=20)
      texto14.bind("<FocusIn>", texto_temp14)
      texto14.bind("<FocusOut>", texto_temp14)

      fr_car_proj = LabelFrame(aba1, borderwidth=1, relief='solid', text='  Características do projeto:  ')
      fr_car_proj.place(x=5, y=205, width=450, height=200)

      Label(fr_car_proj, text='Tipo de projeto:').place(x=5, y=25)
      lista11 = ['Hidrogerador Horizontal', 'Hidrogerador Vertical', 'Turbogerador (ou ST40/ST41)', 'Diesel', 'SH10', 'SH11', 'GH11']
      opcoes11 = StringVar()
      opcoes11.set('Selecionar...')
      OptionMenu(fr_car_proj, opcoes11, *lista11).place(x=200, y=25)

      Label(fr_car_proj, text='Método de proteção do \nmotor/gerador:').place(x=5, y=65)
      lista12 = ['Classificada (Ex)', 'Segura']
      opcoes12 = StringVar()
      opcoes12.set('Selecionar...')
      OptionMenu(fr_car_proj, opcoes12, *lista12).place(x=200, y=65)

      Label(fr_car_proj, text='Material do elemento\n de fixação:').place(x=5, y=105)
      lista13 = ['Inox 304', 'Inox 316', 'Inox 316L', 'Aço carbono']
      opcoes13 = StringVar()
      opcoes13.set('Selecionar...')
      OptionMenu(fr_car_proj, opcoes13, *lista13).place(x=200, y=105)

      Label(fr_car_proj, text='Tamanho da carcaça:').place(x=5, y=145)
      lista14 = ['280', '315', '355', '400', '450', '500', '560', '630', '710', '800', '900', '1000', '1120', '1250', '1400', '1600', '1800', '2000', '2250']
      opcoes14 = StringVar()
      opcoes14.set('Selecionar...')
      OptionMenu(fr_car_proj, opcoes14, *lista14).place(x=200, y=145)

      # Conteúdo da Aba 2 pt ====================================================================================================
      aba2 = Frame(nb)
      nb.add(aba2, text='Carcaça/Estator')

      fr_cb = LabelFrame(aba2, borderwidth=1, relief='solid', text='  Localização do Pt-100:  ')
      fr_cb.place(x=5, y=5, width=520, height=105)

      varcb1 = IntVar()
      varcb1.set(1)

      cb21 = Checkbutton(fr_cb, text='Pt-100 no estator', variable=varcb1, onvalue=1, offvalue=0)
      cb21.pack(padx=5, pady= 30, side=LEFT)
      
      varcb2 = IntVar()
      varcb2.set(1)

      cb22 = Checkbutton(fr_cb, text='Pt-100 no núcleo', variable=varcb2, onvalue=1, offvalue=0)
      cb22.pack(padx=5, pady= 15, side=LEFT)

      fr_res = LabelFrame(aba2, borderwidth=1, relief='solid', text='  Resistores de aquecimento:  ')
      fr_res.place(x=5, y=110, width=520, height=200)

      Label(fr_res, text='Tipo de resistor:').place(x=5, y=25)
      lista21 = ['Tipo "W"', 'Roscado tipo "U" sem cabeçote', 'Roscado tipo "U" com cabeçote']
      opcoes21 = StringVar()
      opcoes21.set('Selecionar...')
      OptionMenu(fr_res, opcoes21, *lista21).place(x=300, y=25)

      Label(fr_res, text='Quantidade total de resistores:').place(x=5, y=65)
      lista22 = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10']
      opcoes22 = StringVar()
      opcoes22.set('Selecionar...')
      OptionMenu(fr_res, opcoes22, *lista22).place(x=300, y=65)

      Label(fr_res, text='Quantidade de resistores por tampa (Lado esquerdo):').place(x=5, y=105)
      lista23 = ['0', '1', '2', '3', '4']
      opcoes23 = StringVar()
      opcoes23.set('Selecionar...')
      OptionMenu(fr_res, opcoes23, *lista23).place(x=300, y=105)

      Label(fr_res, text='Quantidade de resistores por tampa (Lado direito):').place(x=5, y=145)
      lista24 = ['0', '1', '2', '3', '4']
      opcoes24 = StringVar()
      opcoes24.set('Selecionar...')
      OptionMenu(fr_res, opcoes24, *lista24).place(x=300, y=145)

      fr_lt21 = LabelFrame(aba2, borderwidth=1, relief='solid', text='  Posições na lista técnica:  ')
      fr_lt21.place(x=5, y=310, width=300, height=120)

      def texto_temp_lt21(e):
         if lt21.get() == "Digitar número...":
            lt21.delete(0, END)
         elif lt21.get() == '':
            lt21.insert(END, "Digitar número...")
            
      Label(fr_lt21, text='Posição inicial na\n lista técnica:').place(x=5, y=10)
      lt21 = Entry(fr_lt21, takefocus = 0)
      lt21.insert(END, "Digitar número...")
      lt21.place(x=150, y=15, width=105, height=20)
      lt21.bind("<FocusIn>", texto_temp_lt21)
      lt21.bind("<FocusOut>", texto_temp_lt21)

      Label(fr_lt21, text='Incremento:').place(x=5, y=50)
      lis21 = ['1', '5', '10']
      inc21 = StringVar()
      inc21.set('Selecionar...')
      OptionMenu(fr_lt21, inc21, *lis21).place(x=150, y=50)

      # Conteúdo da Aba 3 pt ==================================================================================================
      aba3 = Frame(nb)
      nb.add(aba3, text='Mancal L.A.')

      def ativa_termom_aba3(r):
         if opcoes310.get() == 'Sem termômetro':
            opt311.configure(state='disabled')
            opt312.configure(state='disabled')
            opt313.configure(state='disabled')
            opt314.configure(state='disabled')
         else:
            opt311.configure(state='normal')
            opt312.configure(state='normal')
            opt313.configure(state='normal')
            opt314.configure(state='normal')
      
      def ativa_sensor_oleo_termos_aba3(r):
         if opcoes35.get() == '0':
            opt36.configure(state='disabled')
         else:
            opt36.configure(state='normal')
         
      def ativa_sensor_oleo_termom_aba3(r):
         if opcoes313.get() == '0':
            opt314.configure(state='disabled')
         else:
            opt314.configure(state='normal')

      fr_infos31 = LabelFrame(aba3, borderwidth=1, relief='solid', text='  Informações:  ')
      fr_infos31.place(x=5, y=5, width=450, height=120)

      Label(fr_infos31, text='Isolação do mancal:').place(x=5, y=15)
      lista31 = ['Isolado', 'Não isolado']
      opcoes31 = StringVar()
      opcoes31.set('Selecionar...')
      OptionMenu(fr_infos31, opcoes31, *lista31).place(x=250, y=15)

      Label(fr_infos31, text='Tipo de mancal:').place(x=5, y=50)
      lista32 = ['Deslizamento horizontal', 'Deslizamento vertical', 'Rolamento à óleo', 'Rolamento']
      opcoes32 = StringVar()
      opcoes32.set('Selecionar...')
      OptionMenu(fr_infos31, opcoes32, *lista32).place(x=250, y=50)

      fr_termosensor = LabelFrame(aba3, borderwidth=1, relief='solid', text='  Termoresistor:  ')
      fr_termosensor.place(x=5, y=125, width=450, height=200)

      Label(fr_termosensor, text='Qtde sensor axial - Escora e contra-escora:').place(x=5, y=10)
      lista33 = ['0', '1', '2', '3', '4']
      opcoes33 = StringVar()
      opcoes33.set('Selecionar...')
      OptionMenu(fr_termosensor, opcoes33, *lista33).place(x=300, y=10)

      Label(fr_termosensor, text='Quantidade sensor radial:').place(x=5, y=50)
      lista34 = ['1', '2']
      opcoes34 = StringVar()
      opcoes34.set('Selecionar...')
      OptionMenu(fr_termosensor, opcoes34, *lista34).place(x=300, y=50)

      Label(fr_termosensor, text='Quantidade sensor óleo:').place(x=5, y=90)
      lista35 = ['0', '1', '2']
      opcoes35 = StringVar()
      opcoes35.set('Selecionar...')
      OptionMenu(fr_termosensor, opcoes35, *lista35, command=ativa_sensor_oleo_termos_aba3).place(x=300, y=90)

      Label(fr_termosensor, text='Rosca de fixação do reservatório:').place(x=5, y=130)
      lista36 = ['N/A', 'G 1"', 'G 3/4"', 'G 1.1/4"']
      opcoes36 = StringVar()
      opcoes36.set('Selecionar...')
      opt36 = OptionMenu(fr_termosensor, opcoes36, *lista36)
      opt36.place(x=300, y=130)

      fr_termometro = LabelFrame(aba3, borderwidth=1, relief='solid', text='  Termômetro:  ')
      fr_termometro.place(x=5, y=325, width=450, height=230)

      Label(fr_termometro, text='Contato elétrico:').place(x=5, y=15)
      lista310 = ['Sem contato', 'Com contato', 'Sem termômetro']
      opcoes310 = StringVar()
      opcoes310.set('Selecionar...')
      OptionMenu(fr_termometro, opcoes310, *lista310, command=ativa_termom_aba3).place(x=300, y=10)

      Label(fr_termometro, text='Suporte inclinado para fixação na base:').place(x=5, y=45)
      lista311 = ['Sim', 'Não']
      opcoes311 = StringVar()
      opcoes311.set('Selecionar...')
      opt311 = OptionMenu(fr_termometro, opcoes311, *lista311)
      opt311.place(x=300, y=45)

      Label(fr_termometro, text='Qtde termômetro radial:').place(x=5, y=80)
      lista312 = ['1', '2']
      opcoes312 = StringVar()
      opcoes312.set('Selecionar...')
      opt312 = OptionMenu(fr_termometro, opcoes312, *lista312)
      opt312.place(x=300, y=80)

      Label(fr_termometro, text='Qtde termômetro óleo:').place(x=5, y=115)
      lista313 = ['0', '1', '2']
      opcoes313 = StringVar()
      opcoes313.set('Selecionar...')
      opt313 = OptionMenu(fr_termometro, opcoes313, *lista313, command=ativa_sensor_oleo_termom_aba3)
      opt313.place(x=300, y=115)

      Label(fr_termometro, text='Rosca de fixação do reservatório:').place(x=5, y=150)
      lista314 = ['N/A', 'G 1"', 'G 3/4"', 'G 1.1/4"']
      opcoes314 = StringVar()
      opcoes314.set('Selecionar...')
      opt314 = OptionMenu(fr_termometro, opcoes314, *lista314)
      opt314.place(x=300, y=150)

      fr_lt31 = LabelFrame(aba3, borderwidth=1, relief='solid', text='  Posições na lista técnica:  ')
      fr_lt31.place(x=460, y=5, width=300, height=120)

      def texto_temp_lt31(e):
         if lt31.get() == "Digitar número...":
            lt31.delete(0, END)
         elif lt31.get() == '':
            lt31.insert(END, "Digitar número...")
            
      Label(fr_lt31, text='Posição inicial na\n lista técnica:').place(x=5, y=10)
      lt31 = Entry(fr_lt31, takefocus = 0)
      lt31.insert(END, "Digitar número...")
      lt31.place(x=150, y=15, width=105, height=20)
      lt31.bind("<FocusIn>", texto_temp_lt31)
      lt31.bind("<FocusOut>", texto_temp_lt31)

      Label(fr_lt31, text='Incremento:').place(x=5, y=50)
      lis31 = ['1', '5', '10']
      inc31 = StringVar()
      inc31.set('Selecionar...')
      OptionMenu(fr_lt31, inc31, *lis31).place(x=150, y=50)

      # Conteúdo da Aba 4 pt ==================================================================================================
      aba4 = Frame(nb)
      nb.add(aba4, text='Mancal L.N.A.')

      def ativa_termom_aba4(r):
         if opcoes410.get() == 'Sem termômetro':
            opt411.configure(state='disabled')
            opt412.configure(state='disabled')
            opt413.configure(state='disabled')
            opt414.configure(state='disabled')
         else:
            opt411.configure(state='normal')
            opt412.configure(state='normal')
            opt413.configure(state='normal')
            opt414.configure(state='normal')
      
      def ativa_sensor_oleo_termos_aba4(r):
         if opcoes45.get() == '0':
            opt46.configure(state='disabled')
         else:
            opt46.configure(state='normal')
         
      def ativa_sensor_oleo_termom_aba4(r):
         if opcoes413.get() == '0':
            opt414.configure(state='disabled')
         else:
            opt414.configure(state='normal')

      fr_infos41 = LabelFrame(aba4, borderwidth=1, relief='solid', text='  Informações:  ')
      fr_infos41.place(x=5, y=5, width=450, height=120)

      Label(fr_infos41, text='Isolação do mancal:').place(x=5, y=15)
      lista41 = ['Isolado', 'Não isolado']
      opcoes41 = StringVar()
      opcoes41.set('Selecionar...')
      OptionMenu(fr_infos41, opcoes41, *lista41).place(x=250, y=15)

      Label(fr_infos41, text='Tipo de mancal:').place(x=5, y=50)
      lista42 = ['Deslizamento horizontal', 'Deslizamento vertical', 'Rolamento à óleo', 'Rolamento']
      opcoes42 = StringVar()
      opcoes42.set('Selecionar...')
      OptionMenu(fr_infos41, opcoes42, *lista42).place(x=250, y=50)

      fr_termosensor2 = LabelFrame(aba4, borderwidth=1, relief='solid', text='  Termoresistor:  ')
      fr_termosensor2.place(x=5, y=125, width=450, height=200)

      Label(fr_termosensor2, text='Qtde sensor axial - Escora e contra-escora:').place(x=5, y=10)
      lista43 = ['0', '1', '2', '3', '4']
      opcoes43 = StringVar()
      opcoes43.set('Selecionar...')
      OptionMenu(fr_termosensor2, opcoes43, *lista43).place(x=300, y=10)

      Label(fr_termosensor2, text='Quantidade sensor radial:').place(x=5, y=50)
      lista44 = ['1', '2']
      opcoes44 = StringVar()
      opcoes44.set('Selecionar...')
      OptionMenu(fr_termosensor2, opcoes44, *lista44).place(x=300, y=50)

      Label(fr_termosensor2, text='Quantidade sensor óleo:').place(x=5, y=90)
      lista45 = ['0', '1', '2']
      opcoes45 = StringVar()
      opcoes45.set('Selecionar...')
      OptionMenu(fr_termosensor2, opcoes45, *lista45, command=ativa_sensor_oleo_termos_aba4).place(x=300, y=90)

      Label(fr_termosensor2, text='Rosca de fixação do reservatório:').place(x=5, y=130)
      lista46 = ['N/A', 'G 1"', 'G 3/4"', 'G 1.1/4"']
      opcoes46 = StringVar()
      opcoes46.set('Selecionar...')
      opt46 = OptionMenu(fr_termosensor2, opcoes46, *lista46)
      opt46.place(x=300, y=130)

      fr_termometro2 = LabelFrame(aba4, borderwidth=1, relief='solid', text='  Termômetro:  ')
      fr_termometro2.place(x=5, y=325, width=450, height=230)

      Label(fr_termometro2, text='Contato elétrico:').place(x=5, y=15)
      lista410 = ['Sem contato', 'Com contato', 'Sem termômetro']
      opcoes410 = StringVar()
      opcoes410.set('Selecionar...')
      OptionMenu(fr_termometro2, opcoes410, *lista410, command=ativa_termom_aba4).place(x=300, y=10)

      Label(fr_termometro2, text='Suporte inclinado para fixação na base:').place(x=5, y=45)
      lista411 = ['Sim', 'Não']
      opcoes411 = StringVar()
      opcoes411.set('Selecionar...')
      opt411 = OptionMenu(fr_termometro2, opcoes411, *lista411)
      opt411.place(x=300, y=45)

      Label(fr_termometro2, text='Qtde termômetro radial:').place(x=5, y=80)
      lista412 = ['1', '2']
      opcoes412 = StringVar()
      opcoes412.set('Selecionar...')
      opt412 = OptionMenu(fr_termometro2, opcoes412, *lista412)
      opt412.place(x=300, y=80)

      Label(fr_termometro2, text='Qtde termômetro óleo:').place(x=5, y=115)
      lista413 = ['0', '1', '2']
      opcoes413 = StringVar()
      opcoes413.set('Selecionar...')
      opt413 = OptionMenu(fr_termometro2, opcoes413, *lista413, command=ativa_sensor_oleo_termom_aba4)
      opt413.place(x=300, y=115)

      Label(fr_termometro2, text='Rosca de fixação do reservatório:').place(x=5, y=150)
      lista414 = ['N/A', 'G 1"', 'G 3/4"', 'G 1.1/4"']
      opcoes414 = StringVar()
      opcoes414.set('Selecionar...')
      opt414 = OptionMenu(fr_termometro2, opcoes414, *lista414)
      opt414.place(x=300, y=150)

      fr_lt41 = LabelFrame(aba4, borderwidth=1, relief='solid', text='  Posições na lista técnica:  ')
      fr_lt41.place(x=460, y=5, width=300, height=120)

      def texto_temp_lt41(e):
         if lt41.get() == "Digitar número...":
            lt41.delete(0, END)
         elif lt41.get() == '':
            lt41.insert(END, "Digitar número...")

      Label(fr_lt41, text='Posição inicial na\n lista técnica:').place(x=5, y=10)
      lt41 = Entry(fr_lt41, takefocus = 0)
      lt41.insert(END, "Digitar número...")
      lt41.place(x=150, y=15, width=105, height=20)
      lt41.bind("<FocusIn>", texto_temp_lt41)
      lt41.bind("<FocusOut>", texto_temp_lt41)

      Label(fr_lt41, text='Incremento:').place(x=5, y=50)
      lis41 = ['1', '5', '10']
      inc41 = StringVar()
      inc41.set('Selecionar...')
      OptionMenu(fr_lt41, inc41, *lis41).place(x=150, y=50)

      # Conteúdo da Aba 5 pt ==================================================================================================
      aba5 = Frame(nb)
      nb.add(aba5, text='Cx. Acess. Pequena')
       
      def ativa_cx_peq(sel):
         global flag51, flag52, flag53, flag54
         if opcoes51.get() == '0':
            menu52.configure(state='disable')
            menu53.configure(state='disable')
            
            for child in fr_cx_peq1.winfo_children():
               child.configure(state='disable')
            for child in fr_cx_peq2.winfo_children():
               child.configure(state='disable')
            for child in fr_cx_peq3.winfo_children():
               child.configure(state='disable')
            for child in fr_cx_peq4.winfo_children():
               child.configure(state='disable')
            flag51 = False
            flag52 = False
            flag53 = False
            flag54 = False

         elif opcoes51.get() == '1':
            menu52.configure(state='normal')
            menu53.configure(state='normal')

            for child in fr_cx_peq1.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_peq2.winfo_children():
               child.configure(state='disable')
            for child in fr_cx_peq3.winfo_children():
               child.configure(state='disable')
            for child in fr_cx_peq4.winfo_children():
               child.configure(state='disable')
            flag51 = True
            flag52 = False
            flag53 = False
            flag54 = False
            
         elif opcoes51.get() == '2':
            menu52.configure(state='normal')
            menu53.configure(state='normal')

            for child in fr_cx_peq1.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_peq2.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_peq3.winfo_children():
               child.configure(state='disable')
            for child in fr_cx_peq4.winfo_children():
               child.configure(state='disable')
            flag51 = True
            flag52 = True
            flag53 = False
            flag54 = False
            
         elif opcoes51.get() == '3':
            menu52.configure(state='normal')
            menu53.configure(state='normal')

            for child in fr_cx_peq1.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_peq2.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_peq3.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_peq4.winfo_children():
               child.configure(state='disable')
            flag51 = True
            flag52 = True
            flag53 = True
            flag54 = False
         
         elif opcoes51.get() == '4':
            menu52.configure(state='normal')
            menu53.configure(state='normal')

            for child in fr_cx_peq1.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_peq2.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_peq3.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_peq4.winfo_children():
               child.configure(state='normal')
            flag51 = True
            flag52 = True
            flag53 = True
            flag54 = True


      def seta_qtd_rosca51(*r):
         if opcoes55.get() == 'Sem rosca':
            opcoes56.set('0')
            opt59.configure(state='disabled')

         else:
            opt59.configure(state='normal')
            opcoes56.set('1')

      def seta_qtd_rosca52(*r):
         if opcoes512.get() == 'Sem rosca':
            opcoes513.set('0')
            opt510.configure(state='disabled')

         else:
            opt510.configure(state='normal')
            opcoes513.set('1')

      def seta_qtd_rosca53(*r):
         if opcoes519.get() == 'Sem rosca':
            opcoes520.set('0')
            opt511.configure(state='disabled')

         else:
            opt511.configure(state='normal')
            opcoes520.set('1')

      def seta_qtd_rosca54(*r):
         if opcoes526.get() == 'Sem rosca':
            opcoes527.set('0')
            opt512.configure(state='disabled')

         else:
            opt512.configure(state='normal')
            opcoes527.set('1')
   

      def ativa_rosca_cx_peq1(*z):
         global flag513
         lista55 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
         opt513 = OptionMenu(fr_cx_peq1, opcoes55, *lista55, command=seta_qtd_rosca51)

         if opcoes54.get() == 1:
            opt517.configure(state='disable') # Desabilita 'Material placa cega' se selecionado 'Rosca'
            opcoes57.set('N/A')

            opt513['menu'].delete(0,'end') # Remove todas as opções (reset 1)
            for opt in lista55: # Insere todas as opções (reset2)
                  opt513['menu'].add_command(label=opt, command=tkinter._setit(opcoes55, opt))
            opt513['menu'].delete(0,0) # Remove opção 'Sem rosca'
            flag513 = True

         else:
            opt517.configure(state='normal') # Habilita 'Material placa cega' se selecionado 'Placa'
            opcoes57.set('Selecionar...')

            opt513['menu'].delete(0,'end') # Remove todas as opções (reset1)
            for opt in lista55: # Insere todas as opções
               opt513['menu'].add_command(label=opt, command=tkinter._setit(opcoes55, opt))
            flag513 = False

      def ativa_rosca_cx_peq2():
         global flag514
         lista512 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
         opt514 = OptionMenu(fr_cx_peq1, opcoes512, *lista512, command=seta_qtd_rosca52)

         if opcoes511.get() == 1:
            opt518.configure(state='disable') # Desabilita 'Material placa cega' se selecionado 'Rosca'
            opcoes514.set('N/A')

            opt514['menu'].delete(0,'end') # Remove todas as opções (reset 1)
            for opt in lista512: # Insere todas as opções (reset2)
                  opt514['menu'].add_command(label=opt, command=tkinter._setit(opcoes512, opt))
            opt514['menu'].delete(0,0) # Remove opção 'Sem rosca'
            flag514 = True

         else:
            opt518.configure(state='normal') # Habilita 'Material placa cega' se selecionado 'Placa'
            opcoes514.set('Selecionar...')

            opt514['menu'].delete(0,'end') # Remove todas as opções (reset1)
            for opt in lista512: # Insere todas as opções
                  opt514['menu'].add_command(label=opt, command=tkinter._setit(opcoes512, opt))
            flag514 = False

      def ativa_rosca_cx_peq3():
         global flag515
         lista519 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
         opt515 = OptionMenu(fr_cx_peq1, opcoes519, *lista519, command=seta_qtd_rosca53)

         if opcoes518.get() == 1:
            opt519.configure(state='disable') # Desabilita 'Material placa cega' se selecionado 'Rosca'
            opcoes521.set('N/A')

            opt515['menu'].delete(0,'end') # Remove todas as opções (reset 1)
            for opt in lista519: # Insere todas as opções (reset2)
                  opt515['menu'].add_command(label=opt, command=tkinter._setit(opcoes519, opt))
            opt515['menu'].delete(0,0) # Remove opção 'Sem rosca'
            flag515 = True

         else:
            opt519.configure(state='normal') # Habilita 'Material placa cega' se selecionado 'Placa'
            opcoes521.set('Selecionar...')

            opt515['menu'].delete(0,'end') # Remove todas as opções (reset1)
            for opt in lista519: # Insere todas as opções
                  opt515['menu'].add_command(label=opt, command=tkinter._setit(opcoes519, opt))
            flag515 = False
         

      def ativa_rosca_cx_peq4():
         global flag516
         lista526 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
         opt516 = OptionMenu(fr_cx_peq1, opcoes526, *lista526, command=seta_qtd_rosca54)

         if opcoes525.get() == 1:
            opt520.configure(state='disable') # Desabilita 'Material placa cega' se selecionado 'Rosca'
            opcoes528.set('N/A')

            opt516['menu'].delete(0,'end') # Remove todas as opções (reset 1)
            for opt in lista526: # Insere todas as opções (reset2)
                  opt516['menu'].add_command(label=opt, command=tkinter._setit(opcoes526, opt))
            opt516['menu'].delete(0,0) # Remove opção 'Sem rosca'
            flag516 = True

         else:
            opt520.configure(state='normal') # Habilita 'Material placa cega' se selecionado 'Placa'
            opcoes528.set('Selecionar...')

            opt516['menu'].delete(0,'end') # Remove todas as opções (reset1)
            for opt in lista526: # Insere todas as opções
                  opt516['menu'].add_command(label=opt, command=tkinter._setit(opcoes526, opt))
            flag516 = False


      fr_cx_peq = LabelFrame(aba5, borderwidth=1, relief='solid', text='  Caixa de acessórios pequena:  ')
      fr_cx_peq.place(x=5, y=5, width=1175, height=60)

      Label(fr_cx_peq, text='Quantidade caixas de acessórios:').place(x=5, y=5)
      lista51 = ['0', '1', '2', '3', '4']
      opcoes51 = StringVar()
      opcoes51.set('Selecionar...')
      OptionMenu(fr_cx_peq, opcoes51, *lista51, command=ativa_cx_peq).place(x=200, y=2)

      Label(fr_cx_peq, text='Material caixas de acessórios:').place(x=370, y=5)
      lista52 = ['Inox 304', 'Inox 316', 'Ferro fundido']
      opcoes52 = StringVar()
      opcoes52.set('Selecionar...')
      menu52 = OptionMenu(fr_cx_peq, opcoes52, *lista52)
      menu52.place(x=545, y=2)

      Label(fr_cx_peq, text='Aterramento:').place(x=715, y=5)
      lista53 = ['Com aterramento', 'Sem aterramento']
      opcoes53 = StringVar()
      opcoes53.set('Selecionar...')
      menu53 = OptionMenu(fr_cx_peq, opcoes53, *lista53)
      menu53.place(x=805, y=2)

      fr_cx_peq1 = LabelFrame(aba5, borderwidth=1, relief='solid', text='  Unidade 1:  ')
      fr_cx_peq1.place(x=5, y=70, width=290, height=420)

      def texto_temp_nome51(e):
         if texto51.get() == "Digitar nome...":
            texto51.delete(0, END)
         elif texto51.get() == '':
            texto51.insert(END, "Digitar nome...")
             
      Label(fr_cx_peq1, text='Nome da caixa:').place(x=5, y=5)
      texto51 = Entry(fr_cx_peq1, takefocus = 0)
      texto51.insert(END, "Digitar nome...")
      texto51.place(x=105, y=5, width=150, height=20)
      texto51.bind("<FocusIn>", texto_temp_nome51)
      texto51.bind("<FocusOut>", texto_temp_nome51)

      Label(fr_cx_peq1, text='Tipo de saída de \ncabos do cliente:').place(x=5, y=40)
      opcoes54 = IntVar()
      Radiobutton(fr_cx_peq1, text='Rosca', value=1, variable=opcoes54, command=ativa_rosca_cx_peq1).place(x=120, y=50)
      Radiobutton(fr_cx_peq1, text='Placa', value=2, variable=opcoes54, command=ativa_rosca_cx_peq1).place(x=200, y=50)

      Label(fr_cx_peq1, text='Tipo de rosca:').place(x=5, y=80)
      lista55 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
      opcoes55 = StringVar()
      opcoes55.set('Selecionar...')
      opt513 = OptionMenu(fr_cx_peq1, opcoes55, *lista55, command=seta_qtd_rosca51)
      opt513.place(x=130, y=80)

      Label(fr_cx_peq1, text='Quantidade de rosca:').place(x=5, y=120)
      lista56 = ['1']
      opcoes56 = StringVar()
      opcoes56.set('Selecionar...')
      opt59 = OptionMenu(fr_cx_peq1, opcoes56, *lista56)
      opt59.place(x=130, y=120)

      Label(fr_cx_peq1, text='Material da placa cega:').place(x=5, y=160)
      lista57 = ['Inox 304', 'Inox 316', 'Ferro fundido', 'Alumínio (AlSi6Cu4)', 'Alumínio (AlSi10Mg)']
      opcoes57 = StringVar()
      opcoes57.set('Selecionar...')
      opt517 = OptionMenu(fr_cx_peq1, opcoes57, *lista57)
      opt517.place(x=130, y=160)

      Label(fr_cx_peq1, text='Identificador trilho 1:').place(x=5, y=200)
      lista58 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes58 = StringVar()
      opcoes58.set('Selecionar...')
      OptionMenu(fr_cx_peq1, opcoes58, *lista58).place(x=130, y=200)

      Label(fr_cx_peq1, text='Montada no mancal:').place(x=5, y=240)
      lista59 = ['Sim', 'Não']
      opcoes59 = StringVar()
      opcoes59.set('Selecionar...')
      OptionMenu(fr_cx_peq1, opcoes59, *lista59).place(x=130, y=240)

      Label(fr_cx_peq1, text='Método de proteção:').place(x=5, y=280)
      lista510 = ['Área Segura', 'Ex-e', 'Ex-p']
      opcoes510 = StringVar()
      opcoes510.set('Selecionar...')
      OptionMenu(fr_cx_peq1, opcoes510, *lista510).place(x=130, y=280)

      def texto_temp_lt51(e):
         if lt51.get() == "Digitar número...":
            lt51.delete(0, END)
         elif lt51.get() == '':
            lt51.insert(END, "Digitar número...")

      Label(fr_cx_peq1, text='Posição inicial da\n lista técnica:').place(x=5, y=320)
      lt51 = Entry(fr_cx_peq1, takefocus = 0)
      lt51.insert(END, "Digitar número...")
      lt51.place(x=132.5, y=325, width=102.5, height=20)
      lt51.bind("<FocusIn>", texto_temp_lt51)
      lt51.bind("<FocusOut>", texto_temp_lt51)

      Label(fr_cx_peq1, text='Incremento:').place(x=5, y=360)
      lis51 = ['1', '5', '10']
      inc51 = StringVar()
      inc51.set('Selecionar...')
      OptionMenu(fr_cx_peq1, inc51, *lis51).place(x=130, y=355)

      fr_cx_peq2 = LabelFrame(aba5, borderwidth=1, relief='solid', text='  Unidade 2:  ')
      fr_cx_peq2.place(x=300, y=70, width=290, height=420)

      def texto_temp_nome52(e):
         if texto52.get() == "Digitar nome...":
            texto52.delete(0, END)
         elif texto52.get() == '':
            texto52.insert(END, "Digitar nome...")

      Label(fr_cx_peq2, text='Nome da caixa:').place(x=5, y=5)
      texto52 = Entry(fr_cx_peq2, takefocus = 0)
      texto52.insert(END, "Digitar nome...")
      texto52.place(x=105, y=5, width=150, height=20)
      texto52.bind("<FocusIn>", texto_temp_nome52)
      texto52.bind("<FocusOut>", texto_temp_nome52)

      Label(fr_cx_peq2, text='Tipo de saída de \ncabos do cliente:').place(x=5, y=40)
      opcoes511 = IntVar()
      Radiobutton(fr_cx_peq2, text='Rosca', value=1, variable=opcoes511, command=ativa_rosca_cx_peq2).place(x=120, y=50)
      Radiobutton(fr_cx_peq2, text='Placa', value=2, variable=opcoes511, command=ativa_rosca_cx_peq2).place(x=200, y=50)

      Label(fr_cx_peq2, text='Tipo de rosca:').place(x=5, y=80)
      lista512 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
      opcoes512 = StringVar()
      opcoes512.set('Selecionar...')
      opt514 = OptionMenu(fr_cx_peq2, opcoes512, *lista512, command=seta_qtd_rosca52)
      opt514.place(x=130, y=80)

      Label(fr_cx_peq2, text='Quantidade de rosca:').place(x=5, y=120)
      lista513 = ['1']
      opcoes513 = StringVar()
      opcoes513.set('Selecionar...')
      opt510 = OptionMenu(fr_cx_peq2, opcoes513, *lista513)
      opt510.place(x=130, y=120)

      Label(fr_cx_peq2, text='Material da placa cega:').place(x=5, y=160)
      lista514 = ['Inox 304', 'Inox 316', 'Ferro fundido', 'Alumínio (AlSi6Cu4)', 'Alumínio (AlSi10Mg)']
      opcoes514 = StringVar()
      opcoes514.set('Selecionar...')
      opt518 = OptionMenu(fr_cx_peq2, opcoes514, *lista514)
      opt518.place(x=130, y=160)

      Label(fr_cx_peq2, text='Identificador trilho 1:').place(x=5, y=200)
      lista515 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes515 = StringVar()
      opcoes515.set('Selecionar...')
      OptionMenu(fr_cx_peq2, opcoes515, *lista515).place(x=130, y=200)

      Label(fr_cx_peq2, text='Montada no mancal:').place(x=5, y=240)
      lista516 = ['Sim', 'Não']
      opcoes516 = StringVar()
      opcoes516.set('Selecionar...')
      OptionMenu(fr_cx_peq2, opcoes516, *lista516).place(x=130, y=240)

      Label(fr_cx_peq2, text='Método de proteção:').place(x=5, y=280)
      lista517 = ['Área Segura', 'Ex-e', 'Ex-p']
      opcoes517 = StringVar()
      opcoes517.set('Selecionar...')
      OptionMenu(fr_cx_peq2, opcoes517, *lista517).place(x=130, y=280)

      def texto_temp_lt52(e):
         if lt52.get() == "Digitar número...":
            lt52.delete(0, END)
         elif lt52.get() == '':
            lt52.insert(END, "Digitar número...")

      Label(fr_cx_peq2, text='Posição inicial da\n lista técnica:').place(x=5, y=320)
      lt52 = Entry(fr_cx_peq2, takefocus = 0)
      lt52.insert(END, "Digitar número...")
      lt52.place(x=132.5, y=325, width=102.5, height=20)
      lt52.bind("<FocusIn>", texto_temp_lt52)
      lt52.bind("<FocusOut>", texto_temp_lt52)

      Label(fr_cx_peq2, text='Incremento:').place(x=5, y=360)
      lis52 = ['1', '5', '10']
      inc52 = StringVar()
      inc52.set('Selecionar...')
      OptionMenu(fr_cx_peq2, inc52, *lis52).place(x=130, y=355)

      fr_cx_peq3 = LabelFrame(aba5, borderwidth=1, relief='solid', text='  Unidade 3:  ')
      fr_cx_peq3.place(x=595, y=70, width=290, height=420)

      def texto_temp_nome53(e):
         if texto53.get() == "Digitar nome...":
            texto53.delete(0, END)
         elif texto53.get() == '':
            texto53.insert(END, "Digitar nome...")

      Label(fr_cx_peq3, text='Nome da caixa:').place(x=5, y=5)
      texto53 = Entry(fr_cx_peq3, takefocus = 0)
      texto53.insert(END, "Digitar nome...")
      texto53.place(x=105, y=5, width=150, height=20)
      texto53.bind("<FocusIn>", texto_temp_nome53)
      texto53.bind("<FocusOut>", texto_temp_nome53)

      Label(fr_cx_peq3, text='Tipo de saída de \ncabos do cliente:').place(x=5, y=40)
      opcoes518 = IntVar()
      Radiobutton(fr_cx_peq3, text='Rosca', value=1, variable=opcoes518, command=ativa_rosca_cx_peq3).place(x=120, y=50)
      Radiobutton(fr_cx_peq3, text='Placa', value=2, variable=opcoes518, command=ativa_rosca_cx_peq3).place(x=200, y=50)

      Label(fr_cx_peq3, text='Tipo de rosca:').place(x=5, y=80)
      lista519 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
      opcoes519 = StringVar()
      opcoes519.set('Selecionar...')
      opt515 = OptionMenu(fr_cx_peq3, opcoes519, *lista519, command=seta_qtd_rosca53)
      opt515.place(x=130, y=80)

      Label(fr_cx_peq3, text='Quantidade de rosca:').place(x=5, y=120)
      lista520 = ['1']
      opcoes520 = StringVar()
      opcoes520.set('Selecionar...')
      opt511 = OptionMenu(fr_cx_peq3, opcoes520, *lista520)
      opt511.place(x=130, y=120)

      Label(fr_cx_peq3, text='Material da placa cega:').place(x=5, y=160)
      lista521 = ['Inox 304', 'Inox 316', 'Ferro fundido', 'Alumínio (AlSi6Cu4)', 'Alumínio (AlSi10Mg)']
      opcoes521 = StringVar()
      opcoes521.set('Selecionar...')
      opt519 = OptionMenu(fr_cx_peq3, opcoes521, *lista521)
      opt519.place(x=130, y=160)

      Label(fr_cx_peq3, text='Identificador trilho 1:').place(x=5, y=200)
      lista522 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes522 = StringVar()
      opcoes522.set('Selecionar...')
      OptionMenu(fr_cx_peq3, opcoes522, *lista522).place(x=130, y=200)

      Label(fr_cx_peq3, text='Montada no mancal:').place(x=5, y=240)
      lista523 = ['Sim', 'Não']
      opcoes523 = StringVar()
      opcoes523.set('Selecionar...')
      OptionMenu(fr_cx_peq3, opcoes523, *lista523).place(x=130, y=240)

      Label(fr_cx_peq3, text='Método de proteção:').place(x=5, y=280)
      lista524 = ['Área Segura', 'Ex-e', 'Ex-p']
      opcoes524 = StringVar()
      opcoes524.set('Selecionar...')
      OptionMenu(fr_cx_peq3, opcoes524, *lista524).place(x=130, y=280)

      def texto_temp_lt53(e):
         if lt53.get() == "Digitar número...":
            lt53.delete(0, END)
         elif lt53.get() == '':
            lt53.insert(END, "Digitar número...")

      Label(fr_cx_peq3, text='Posição inicial da\n lista técnica:').place(x=5, y=320)
      lt53 = Entry(fr_cx_peq3, takefocus = 0)
      lt53.insert(END, "Digitar número...")
      lt53.place(x=132.5, y=325, width=102.5, height=20)
      lt53.bind("<FocusIn>", texto_temp_lt53)
      lt53.bind("<FocusOut>", texto_temp_lt53)

      Label(fr_cx_peq3, text='Incremento:').place(x=5, y=360)
      lis53 = ['1', '5', '10']
      inc53 = StringVar()
      inc53.set('Selecionar...')
      OptionMenu(fr_cx_peq3, inc53, *lis53).place(x=130, y=355)

      fr_cx_peq4 = LabelFrame(aba5, borderwidth=1, relief='solid', text='  Unidade 4:  ')
      fr_cx_peq4.place(x=890, y=70, width=290, height=420)

      def texto_temp_nome54(e):
         if texto54.get() == "Digitar nome...":
            texto54.delete(0, END)
         elif texto54.get() == '':
            texto54.insert(END, "Digitar nome...")

      Label(fr_cx_peq4, text='Nome da caixa:').place(x=5, y=5)
      texto54 = Entry(fr_cx_peq4, takefocus = 0)
      texto54.insert(END, "Digitar nome...")
      texto54.place(x=105, y=5, width=150, height=20)
      texto54.bind("<FocusIn>", texto_temp_nome54)
      texto54.bind("<FocusOut>", texto_temp_nome54)

      Label(fr_cx_peq4, text='Tipo de saída de \ncabos do cliente:').place(x=5, y=40)
      opcoes525 = IntVar()
      Radiobutton(fr_cx_peq4, text='Rosca', value=1, variable=opcoes525, command=ativa_rosca_cx_peq4).place(x=120, y=50)
      Radiobutton(fr_cx_peq4, text='Placa', value=2, variable=opcoes525, command=ativa_rosca_cx_peq4).place(x=200, y=50)

      Label(fr_cx_peq4, text='Tipo de rosca:').place(x=5, y=80)
      lista526 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
      opcoes526 = StringVar()
      opcoes526.set('Selecionar...')
      opt516 = OptionMenu(fr_cx_peq4, opcoes526, *lista526, command=seta_qtd_rosca54)
      opt516.place(x=130, y=80)

      Label(fr_cx_peq4, text='Quantidade de rosca:').place(x=5, y=120)
      lista527 = ['1']
      opcoes527 = StringVar()
      opcoes527.set('Selecionar...')
      opt512 = OptionMenu(fr_cx_peq4, opcoes527, *lista527)
      opt512.place(x=130, y=120)

      Label(fr_cx_peq4, text='Material da placa cega:').place(x=5, y=160)
      lista528 = ['Inox 304', 'Inox 316', 'Ferro fundido', 'Alumínio (AlSi6Cu4)', 'Alumínio (AlSi10Mg)']
      opcoes528 = StringVar()
      opcoes528.set('Selecionar...')
      opt520 = OptionMenu(fr_cx_peq4, opcoes528, *lista528)
      opt520.place(x=130, y=160)

      Label(fr_cx_peq4, text='Identificador trilho 1:').place(x=5, y=200)
      lista529 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes529 = StringVar()
      opcoes529.set('Selecionar...')
      OptionMenu(fr_cx_peq4, opcoes529, *lista529).place(x=130, y=200)

      Label(fr_cx_peq4, text='Montada no mancal:').place(x=5, y=240)
      lista530 = ['Sim', 'Não']
      opcoes530 = StringVar()
      opcoes530.set('Selecionar...')
      OptionMenu(fr_cx_peq4, opcoes530, *lista530).place(x=130, y=240)

      Label(fr_cx_peq4, text='Método de proteção:').place(x=5, y=280)
      lista531 = ['Área Segura', 'Ex-e', 'Ex-p']
      opcoes531 = StringVar()
      opcoes531.set('Selecionar...')
      OptionMenu(fr_cx_peq4, opcoes531, *lista531).place(x=130, y=280)

      def texto_temp_lt54(e):
         if lt54.get() == "Digitar número...":
            lt54.delete(0, END)
         elif lt54.get() == '':
            lt54.insert(END, "Digitar número...")

      Label(fr_cx_peq4, text='Posição inicial da\n lista técnica:').place(x=5, y=320)
      lt54 = Entry(fr_cx_peq4, takefocus = 0)
      lt54.insert(END, "Digitar número...")
      lt54.place(x=132.5, y=325, width=102.5, height=20)
      lt54.bind("<FocusIn>", texto_temp_lt54)
      lt54.bind("<FocusOut>", texto_temp_lt54)

      Label(fr_cx_peq4, text='Incremento:').place(x=5, y=360)
      lis54 = ['1', '5', '10']
      inc54 = StringVar()
      inc54.set('Selecionar...')
      OptionMenu(fr_cx_peq4, inc54, *lis54).place(x=130, y=355)

      # Conteúdo da Aba 6 pt (Cx Acessórios Média) ===============================================================================
      aba6 = Frame(nb)
      nb.add(aba6, text='Cx. Acess. Média')

      def ativa_cx_med(sel):
         global flag61, flag62, flag63, flag64
         if opcoes61.get() == '0':
            menu62.configure(state='disable')
            menu63.configure(state='disable')

            for child in fr_cx_med1.winfo_children():
               child.configure(state='disable')
            for child in fr_cx_med2.winfo_children():
               child.configure(state='disable')
            for child in fr_cx_med3.winfo_children():
               child.configure(state='disable')
            for child in fr_cx_med4.winfo_children():
               child.configure(state='disable')
            flag61 = False
            flag62 = False
            flag63 = False
            flag64 = False

         elif opcoes61.get() == '1':
            menu62.configure(state='normal')
            menu63.configure(state='normal')

            for child in fr_cx_med1.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_med2.winfo_children():
               child.configure(state='disable')
            for child in fr_cx_med3.winfo_children():
               child.configure(state='disable')
            for child in fr_cx_med4.winfo_children():
               child.configure(state='disable')
            flag61 = True
            flag62 = False
            flag63 = False
            flag64 = False
            
         elif opcoes61.get() == '2':
            menu62.configure(state='normal')
            menu63.configure(state='normal')

            for child in fr_cx_med1.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_med2.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_med3.winfo_children():
               child.configure(state='disable')
            for child in fr_cx_med4.winfo_children():
               child.configure(state='disable')
            flag61 = True
            flag62 = True
            flag63 = False
            flag64 = False
            
         elif opcoes61.get() == '3':
            menu62.configure(state='normal')
            menu63.configure(state='normal')

            for child in fr_cx_med1.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_med2.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_med3.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_med4.winfo_children():
               child.configure(state='disable')
            flag61 = True
            flag62 = True
            flag63 = True
            flag64 = False
         
         elif opcoes61.get() == '4':
            menu62.configure(state='normal')
            menu63.configure(state='normal')

            for child in fr_cx_med1.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_med2.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_med3.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_med4.winfo_children():
               child.configure(state='normal')
            flag61 = True
            flag62 = True
            flag63 = True
            flag64 = True


      def ativa_ident_Cx_med1(sel):
         global flag65, tagX2_61
         if opcoes69.get() == '1':
            opt62.configure(state='disable')
            opcoes611.set('N/A')
            #tagX2_61 = 'N/A'
            flag65 = False

         elif opcoes69.get() == '2':
            opt62.configure(state='normal')
            opcoes611.set('Selecionar...')
            #tagX2_61 = opcoes611.get()
            flag65 = True
      
      def ativa_ident_Cx_med2(sel):
         global flag66, tagX2_62
         if opcoes620.get() == '1':
            opt64.configure(state='disable')
            opcoes622.set('N/A')
            #tagX2_62 = 'N/A'
            flag66 = False

         elif opcoes620.get() == '2':
            opt64.configure(state='normal')
            opcoes622.set('Selecionar...')
            #tagX2_62 = opcoes622.get()
            flag66 = True

      def ativa_ident_Cx_med3(sel):
         global flag67, tagX2_63
         if opcoes631.get() == '1':
            opt66.configure(state='disable')
            opcoes633.set('N/A')
            #tagX2_63 = 'N/A'
            flag67 = False

         elif opcoes631.get() == '2':
            opt66.configure(state='normal')
            opcoes633.set('Selecionar...')
            #tagX2_63 = opcoes633.get()
            flag67 = True

      def ativa_ident_Cx_med4(sel):
         global flag68, tagX2_64
         if opcoes642.get() == '1':
            opt68.configure(state='disable')
            opcoes644.set('N/A')
            #tagX2_64 = 'N/A'
            flag68 = False

         elif opcoes642.get() == '2':
            opt68.configure(state='normal')
            opcoes644.set('Selecionar...')
            #tagX2_64 = opcoes644.get()
            flag68 = True

      
      def seta_qtd_rosca61(*r):
         if opcoes66.get() == 'Sem rosca':
            opcoes67.set('0')
            opt69.configure(state='disabled')

         else:
            opt69.configure(state='normal')
            opcoes67.set('Selecionar...')

      def seta_qtd_rosca62(*r):
         if opcoes617.get() == 'Sem rosca':
            opcoes618.set('0')
            opt610.configure(state='disabled')

         else:
            opt610.configure(state='normal')
            opcoes618.set('Selecionar...')

      def seta_qtd_rosca63(*r):
         if opcoes628.get() == 'Sem rosca':
            opcoes629.set('0')
            opt611.configure(state='disabled')

         else:
            opt611.configure(state='normal')
            opcoes629.set('Selecionar...')

      def seta_qtd_rosca64(*r):
         if opcoes639.get() == 'Sem rosca':
            opcoes640.set('0')
            opt612.configure(state='disabled')

         else:
            opt612.configure(state='normal')
            opcoes640.set('Selecionar...')

      def ativa_rosca_cx_med1():
         global flag613
         lista66 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
         opt613 = OptionMenu(fr_cx_med1, opcoes66, *lista66, command=seta_qtd_rosca61)

         if opcoes64.get() == 1:
            opt617.configure(state='disable') # Desabilita 'Material placa cega' se selecionado 'Rosca'
            opcoes68.set('N/A')

            opt613['menu'].delete(0,'end') # Remove todas as opções (reset 1)
            for opt in lista66: # Insere todas as opções (reset2)
                  opt613['menu'].add_command(label=opt, command=tkinter._setit(opcoes66, opt))
            opt613['menu'].delete(0,0) # Remove opção 'Sem rosca'
            flag613 = True

         else:
            opt617.configure(state='normal') # Habilita 'Material placa cega' se selecionado 'Placa'
            opcoes68.set('Selecionar...')

            opt613['menu'].delete(0,'end') # Remove todas as opções (reset1)
            for opt in lista66: # Insere todas as opções
                  opt613['menu'].add_command(label=opt, command=tkinter._setit(opcoes66, opt))
            flag613 = False

      def ativa_rosca_cx_med2():
         global flag614
         lista617 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
         opt614 = OptionMenu(fr_cx_med2, opcoes617, *lista617, command=seta_qtd_rosca62)

         if opcoes615.get() == 1:
            opt618.configure(state='disable') # Desabilita 'Material placa cega' se selecionado 'Rosca'
            opcoes619.set('N/A')

            opt614['menu'].delete(0,'end') # Remove todas as opções (reset 1)
            for opt in lista617: # Insere todas as opções (reset2)
                  opt614['menu'].add_command(label=opt, command=tkinter._setit(opcoes617, opt))
            opt614['menu'].delete(0,0) # Remove opção 'Sem rosca'
            flag614 = True

         else:
            opt618.configure(state='normal') # Habilita 'Material placa cega' se selecionado 'Placa'
            opcoes619.set('Selecionar...')

            opt614['menu'].delete(0,'end') # Remove todas as opções (reset1)
            for opt in lista617: # Insere todas as opções
                  opt614['menu'].add_command(label=opt, command=tkinter._setit(opcoes617, opt))
            flag614 = False

      def ativa_rosca_cx_med3():
         global flag615
         lista628 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
         opt615 = OptionMenu(fr_cx_med3, opcoes628, *lista628, command=seta_qtd_rosca63)

         if opcoes626.get() == 1:
            opt619.configure(state='disable') # Desabilita 'Material placa cega' se selecionado 'Rosca'
            opcoes630.set('N/A')

            opt615['menu'].delete(0,'end') # Remove todas as opções (reset 1)
            for opt in lista628: # Insere todas as opções (reset2)
                  opt615['menu'].add_command(label=opt, command=tkinter._setit(opcoes628, opt))
            opt615['menu'].delete(0,0) # Remove opção 'Sem rosca'
            flag615 = True

         else:
            opt619.configure(state='normal') # Habilita 'Material placa cega' se selecionado 'Placa'
            opcoes630.set('Selecionar...')

            opt615['menu'].delete(0,'end') # Remove todas as opções (reset1)
            for opt in lista628: # Insere todas as opções
                  opt615['menu'].add_command(label=opt, command=tkinter._setit(opcoes628, opt))
            flag615 = False

      def ativa_rosca_cx_med4():
         global flag616
         lista639 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
         opt616 = OptionMenu(fr_cx_med4, opcoes639, *lista639, command=seta_qtd_rosca64)

         if opcoes637.get() == 1:
            opt620.configure(state='disable') # Desabilita 'Material placa cega' se selecionado 'Rosca'
            opcoes641.set('N/A')

            opt616['menu'].delete(0,'end') # Remove todas as opções (reset 1)
            for opt in lista639: # Insere todas as opções (reset2)
                  opt616['menu'].add_command(label=opt, command=tkinter._setit(opcoes639, opt))
            opt616['menu'].delete(0,0) # Remove opção 'Sem rosca'
            flag616 = True

         else:
            opt620.configure(state='normal') # Habilita 'Material placa cega' se selecionado 'Placa'
            opcoes641.set('Selecionar...')

            opt616['menu'].delete(0,'end') # Remove todas as opções (reset1)
            for opt in lista639: # Insere todas as opções
                  opt616['menu'].add_command(label=opt, command=tkinter._setit(opcoes639, opt))
            flag616 = False

      fr_cx_med = LabelFrame(aba6, borderwidth=1, relief='solid', text='  Caixa de acessórios média:  ')
      fr_cx_med.place(x=5, y=5, width=1175, height=60)

      Label(fr_cx_med, text='Quantidade caixas de acessórios:').place(x=5, y=5)
      lista61 = ['0', '1', '2', '3', '4']
      opcoes61 = StringVar()
      opcoes61.set('Selecionar...')
      OptionMenu(fr_cx_med, opcoes61, *lista61, command=ativa_cx_med).place(x=200, y=2)

      Label(fr_cx_med, text='Material caixas de acessórios:').place(x=370, y=5)
      lista62 = ['Inox 304', 'Inox 316', 'Ferro fundido']
      opcoes62 = StringVar()
      opcoes62.set('Selecionar...')
      menu62 = OptionMenu(fr_cx_med, opcoes62, *lista62)
      menu62.place(x=545, y=2)

      Label(fr_cx_med, text='Aterramento:').place(x=715, y=5)
      lista63 = ['Com aterramento', 'Sem aterramento']
      opcoes63 = StringVar()
      opcoes63.set('Selecionar...')
      menu63 = OptionMenu(fr_cx_med, opcoes63, *lista63)
      menu63.place(x=805, y=2)

      fr_cx_med1 = LabelFrame(aba6, borderwidth=1, relief='solid', text='  Unidade 1:  ')
      fr_cx_med1.place(x=5, y=70, width=290, height=520)

      def texto_temp_nome61(e):
         if texto61.get() == "Digitar nome...":
            texto61.delete(0, END)
         elif texto61.get() == '':
            texto61.insert(END, "Digitar nome...")

      Label(fr_cx_med1, text='Nome da caixa:').place(x=5, y=5)
      texto61 = Entry(fr_cx_med1, takefocus = 0)
      texto61.insert(END, "Digitar nome...")
      texto61.place(x=105, y=5, width=150, height=20)
      texto61.bind("<FocusIn>", texto_temp_nome61)
      texto61.bind("<FocusOut>", texto_temp_nome61)

      Label(fr_cx_med1, text='Tipo de saída de \ncabos do cliente:').place(x=5, y=40)
      opcoes64 = IntVar()
      Radiobutton(fr_cx_med1, text='Rosca', value=1, variable=opcoes64, command=ativa_rosca_cx_med1).place(x=120, y=50)
      Radiobutton(fr_cx_med1, text='Placa', value=2, variable=opcoes64, command=ativa_rosca_cx_med1).place(x=200, y=50)

      Label(fr_cx_med1, text='Previsão termostato:').place(x=5, y=80)
      lista65 = ['Nenhum', 'Esquerdo', 'Direito', 'Ambos']
      opcoes65 = StringVar()
      opcoes65.set('Selecionar...')
      OptionMenu(fr_cx_med1, opcoes65, *lista65).place(x=130, y=80)

      Label(fr_cx_med1, text='Tipo de rosca:').place(x=5, y=115)
      lista66 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
      opcoes66 = StringVar()
      opcoes66.set('Selecionar...')
      opt613 = OptionMenu(fr_cx_med1, opcoes66, *lista66, command=seta_qtd_rosca61)
      opt613.place(x=130, y=115)

      Label(fr_cx_med1, text='Quantidade de rosca:').place(x=5, y=150)
      lista67 = ['0', '1', '2']
      opcoes67 = StringVar()
      opcoes67.set('Selecionar...')
      opt69 = OptionMenu(fr_cx_med1, opcoes67, *lista67)
      opt69.place(x=130, y=150)

      Label(fr_cx_med1, text='Material da placa cega:').place(x=5, y=185)
      lista68 = ['Inox 304', 'Inox 316', 'Ferro fundido', 'Alumínio (AlSi6Cu4)', 'Alumínio (AlSi10Mg)']
      opcoes68 = StringVar()
      opcoes68.set('Selecionar...')
      opt617 = OptionMenu(fr_cx_med1, opcoes68, *lista68)
      opt617.place(x=130, y=185)

      Label(fr_cx_med1, text='Quantidade de trilhos:').place(x=5, y=220)
      lista69 = ['1', '2']
      opcoes69 = StringVar()
      opcoes69.set('Selecionar...')
      OptionMenu(fr_cx_med1, opcoes69, *lista69, command=ativa_ident_Cx_med1).place(x=130, y=220)

      Label(fr_cx_med1, text='Identificador trilho 1:').place(x=5, y=255)
      lista610 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes610 = StringVar()
      opcoes610.set('Selecionar...')
      opt61 = OptionMenu(fr_cx_med1, opcoes610, *lista610)
      opt61.place(x=130, y=255)

      Label(fr_cx_med1, text='Identificador trilho 2:').place(x=5, y=290)
      lista611 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes611 = StringVar()
      opcoes611.set('Selecionar...')
      opt62 = OptionMenu(fr_cx_med1, opcoes611, *lista611)
      opt62.place(x=130, y=290)

      Label(fr_cx_med1, text='Montada no mancal:').place(x=5, y=325)
      lista612 = ['Sim', 'Não']
      opcoes612 = StringVar()
      opcoes612.set('Selecionar...')
      OptionMenu(fr_cx_med1, opcoes612, *lista612).place(x=130, y=325)

      Label(fr_cx_med1, text='Tampa profundidade \naumentada:').place(x=5, y=360)
      lista613 = ['Sim', 'Não']
      opcoes613 = StringVar()
      opcoes613.set('Selecionar...')
      OptionMenu(fr_cx_med1, opcoes613, *lista613).place(x=130, y=360)

      Label(fr_cx_med1, text='Método de proteção:').place(x=5, y=400)
      lista614 = ['Área Segura', 'Ex-e', 'Ex-p']
      opcoes614 = StringVar()
      opcoes614.set('Selecionar...')
      OptionMenu(fr_cx_med1, opcoes614, *lista614).place(x=130, y=395)

      def texto_temp_lt61(e):
         if lt61.get() == "Digitar número...":
            lt61.delete(0, END)
         elif lt61.get() == '':
            lt61.insert(END, "Digitar número...")

      Label(fr_cx_med1, text='Posição inicial da\n lista técnica:').place(x=5, y=430)
      lt61 = Entry(fr_cx_med1, takefocus = 0)
      lt61.insert(END, "Digitar número...")
      lt61.place(x=132.5, y=435, width=102.5, height=20)
      lt61.bind("<FocusIn>", texto_temp_lt61)
      lt61.bind("<FocusOut>", texto_temp_lt61)

      Label(fr_cx_med1, text='Incremento:').place(x=5, y=470)
      lis61 = ['1', '5', '10']
      inc61 = StringVar()
      inc61.set('Selecionar...')
      OptionMenu(fr_cx_med1, inc61, *lis61).place(x=130, y=465)

      fr_cx_med2 = LabelFrame(aba6, borderwidth=1, relief='solid', text='  Unidade 2:  ')
      fr_cx_med2.place(x=300, y=70, width=290, height=520)

      def texto_temp_nome62(e):
         if texto62.get() == "Digitar nome...":
            texto62.delete(0, END)
         elif texto62.get() == '':
            texto62.insert(END, "Digitar nome...")

      Label(fr_cx_med2, text='Nome da caixa:').place(x=5, y=5)
      texto62 = Entry(fr_cx_med2, takefocus = 0)
      texto62.insert(END, "Digitar nome...")
      texto62.place(x=105, y=5, width=150, height=20)
      texto62.bind("<FocusIn>", texto_temp_nome62)
      texto62.bind("<FocusOut>", texto_temp_nome62)

      Label(fr_cx_med2, text='Tipo de saída de \ncabos do cliente:').place(x=5, y=40)
      opcoes615 = IntVar()
      Radiobutton(fr_cx_med2, text='Rosca', value=1, variable=opcoes615, command=ativa_rosca_cx_med2).place(x=120, y=50)
      Radiobutton(fr_cx_med2, text='Placa', value=2, variable=opcoes615, command=ativa_rosca_cx_med2).place(x=200, y=50)

      Label(fr_cx_med2, text='Previsão termostato:').place(x=5, y=80)
      lista616 = ['Nenhum', 'Esquerdo', 'Direito', 'Ambos']
      opcoes616 = StringVar()
      opcoes616.set('Selecionar...')
      OptionMenu(fr_cx_med2, opcoes616, *lista616).place(x=130, y=80)

      Label(fr_cx_med2, text='Tipo de rosca:').place(x=5, y=115)
      lista617 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
      opcoes617 = StringVar()
      opcoes617.set('Selecionar...')
      opt614 = OptionMenu(fr_cx_med2, opcoes617, *lista617, command=seta_qtd_rosca62)
      opt614.place(x=130, y=115)

      Label(fr_cx_med2, text='Quantidade de rosca:').place(x=5, y=150)
      lista618 = ['0', '1', '2']
      opcoes618 = StringVar()
      opcoes618.set('Selecionar...')
      opt610 = OptionMenu(fr_cx_med2, opcoes618, *lista618)
      opt610.place(x=130, y=150)

      Label(fr_cx_med2, text='Material da placa cega:').place(x=5, y=185)
      lista619 = ['Inox 304', 'Inox 316', 'Ferro fundido', 'Alumínio (AlSi6Cu4)', 'Alumínio (AlSi10Mg)']
      opcoes619 = StringVar()
      opcoes619.set('Selecionar...')
      opt618 = OptionMenu(fr_cx_med2, opcoes619, *lista619)
      opt618.place(x=130, y=185)

      Label(fr_cx_med2, text='Quantidade de trilhos:').place(x=5, y=220)
      lista620 = ['1', '2']
      opcoes620 = StringVar()
      opcoes620.set('Selecionar...')
      OptionMenu(fr_cx_med2, opcoes620, *lista620, command=ativa_ident_Cx_med2).place(x=130, y=220)

      Label(fr_cx_med2, text='Identificador trilho 1:').place(x=5, y=255)
      lista621 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes621 = StringVar()
      opcoes621.set('Selecionar...')
      opt63 = OptionMenu(fr_cx_med2, opcoes621, *lista621)
      opt63.place(x=130, y=255)

      Label(fr_cx_med2, text='Identificador trilho 2:').place(x=5, y=290)
      lista622 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes622 = StringVar()
      opcoes622.set('Selecionar...')
      opt64 = OptionMenu(fr_cx_med2, opcoes622, *lista622)
      opt64.place(x=130, y=290)

      Label(fr_cx_med2, text='Montada no mancal:').place(x=5, y=325)
      lista623 = ['Sim', 'Não']
      opcoes623 = StringVar()
      opcoes623.set('Selecionar...')
      OptionMenu(fr_cx_med2, opcoes623, *lista623).place(x=130, y=325)

      Label(fr_cx_med2, text='Tampa profundidade \naumentada:').place(x=5, y=360)
      lista624 = ['Sim', 'Não']
      opcoes624 = StringVar()
      opcoes624.set('Selecionar...')
      OptionMenu(fr_cx_med2, opcoes624, *lista624).place(x=130, y=360)

      Label(fr_cx_med2, text='Método de proteção:').place(x=5, y=400)
      lista625 = ['Área Segura', 'Ex-e', 'Ex-p']
      opcoes625 = StringVar()
      opcoes625.set('Selecionar...')
      OptionMenu(fr_cx_med2, opcoes625, *lista625).place(x=130, y=395)

      def texto_temp_lt62(e):
         if lt62.get() == "Digitar número...":
            lt62.delete(0, END)
         elif lt62.get() == '':
            lt62.insert(END, "Digitar número...")

      Label(fr_cx_med2, text='Posição inicial da\n lista técnica:').place(x=5, y=430)
      lt62 = Entry(fr_cx_med2, takefocus = 0)
      lt62.insert(END, "Digitar número...")
      lt62.place(x=132.5, y=435, width=102.5, height=20)
      lt62.bind("<FocusIn>", texto_temp_lt62)
      lt62.bind("<FocusOut>", texto_temp_lt62)

      Label(fr_cx_med2, text='Incremento:').place(x=5, y=470)
      lis62 = ['1', '5', '10']
      inc62 = StringVar()
      inc62.set('Selecionar...')
      OptionMenu(fr_cx_med2, inc62, *lis62).place(x=130, y=465)

      fr_cx_med3 = LabelFrame(aba6, borderwidth=1, relief='solid', text='  Unidade 3:  ')
      fr_cx_med3.place(x=595, y=70, width=290, height=520)

      def texto_temp_nome63(e):
         if texto63.get() == "Digitar nome...":
            texto63.delete(0, END)
         elif texto63.get() == '':
            texto63.insert(END, "Digitar nome...")

      Label(fr_cx_med3, text='Nome da caixa:').place(x=5, y=5)
      texto63 = Entry(fr_cx_med3, takefocus = 0)
      texto63.insert(END, "Digitar nome...")
      texto63.place(x=105, y=5, width=150, height=20)
      texto63.bind("<FocusIn>", texto_temp_nome63)
      texto63.bind("<FocusOut>", texto_temp_nome63)

      Label(fr_cx_med3, text='Tipo de saída de \ncabos do cliente:').place(x=5, y=40)
      opcoes626 = IntVar()
      Radiobutton(fr_cx_med3, text='Rosca', value=1, variable=opcoes626, command=ativa_rosca_cx_med3).place(x=120, y=50)
      Radiobutton(fr_cx_med3, text='Placa', value=2, variable=opcoes626, command=ativa_rosca_cx_med3).place(x=200, y=50)

      Label(fr_cx_med3, text='Previsão termostato:').place(x=5, y=80)
      lista627 = ['Nenhum', 'Esquerdo', 'Direito', 'Ambos']
      opcoes627 = StringVar()
      opcoes627.set('Selecionar...')
      OptionMenu(fr_cx_med3, opcoes627, *lista627).place(x=130, y=80)

      Label(fr_cx_med3, text='Tipo de rosca:').place(x=5, y=115)
      lista628 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
      opcoes628 = StringVar()
      opcoes628.set('Selecionar...')
      opt615 = OptionMenu(fr_cx_med3, opcoes628, *lista628, command=seta_qtd_rosca63)
      opt615.place(x=130, y=115)

      Label(fr_cx_med3, text='Quantidade de rosca:').place(x=5, y=150)
      lista629 = ['0', '1', '2']
      opcoes629 = StringVar()
      opcoes629.set('Selecionar...')
      opt611 = OptionMenu(fr_cx_med3, opcoes629, *lista629)
      opt611.place(x=130, y=150)

      Label(fr_cx_med3, text='Material da placa cega:').place(x=5, y=185)
      lista630 = ['Inox 304', 'Inox 316', 'Ferro fundido', 'Alumínio (AlSi6Cu4)', 'Alumínio (AlSi10Mg)']
      opcoes630 = StringVar()
      opcoes630.set('Selecionar...')
      opt619 = OptionMenu(fr_cx_med3, opcoes630, *lista630)
      opt619.place(x=130, y=185)

      Label(fr_cx_med3, text='Quantidade de trilhos:').place(x=5, y=220)
      lista631 = ['1', '2']
      opcoes631 = StringVar()
      opcoes631.set('Selecionar...')
      OptionMenu(fr_cx_med3, opcoes631, *lista631, command=ativa_ident_Cx_med3).place(x=130, y=220)

      Label(fr_cx_med3, text='Identificador trilho 1:').place(x=5, y=255)
      lista632 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes632 = StringVar()
      opcoes632.set('Selecionar...')
      opt65 = OptionMenu(fr_cx_med3, opcoes632, *lista632)
      opt65.place(x=130, y=255)

      Label(fr_cx_med3, text='Identificador trilho 2:').place(x=5, y=290)
      lista633 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes633 = StringVar()
      opcoes633.set('Selecionar...')
      opt66 = OptionMenu(fr_cx_med3, opcoes633, *lista633)
      opt66.place(x=130, y=290)

      Label(fr_cx_med3, text='Montada no mancal:').place(x=5, y=325)
      lista634 = ['Sim', 'Não']
      opcoes634 = StringVar()
      opcoes634.set('Selecionar...')
      OptionMenu(fr_cx_med3, opcoes634, *lista634).place(x=130, y=325)

      Label(fr_cx_med3, text='Tampa profundidade \naumentada:').place(x=5, y=360)
      lista635 = ['Sim', 'Não']
      opcoes635 = StringVar()
      opcoes635.set('Selecionar...')
      OptionMenu(fr_cx_med3, opcoes635, *lista635).place(x=130, y=360)

      Label(fr_cx_med3, text='Método de proteção:').place(x=5, y=400)
      lista636 = ['Área Segura', 'Ex-e', 'Ex-p']
      opcoes636 = StringVar()
      opcoes636.set('Selecionar...')
      OptionMenu(fr_cx_med3, opcoes636, *lista636).place(x=130, y=395)

      def texto_temp_lt63(e):
         if lt63.get() == "Digitar número...":
            lt63.delete(0, END)
         elif lt63.get() == '':
            lt63.insert(END, "Digitar número...")

      Label(fr_cx_med3, text='Posição inicial da\n lista técnica:').place(x=5, y=430)
      lt63 = Entry(fr_cx_med3, takefocus = 0)
      lt63.insert(END, "Digitar número...")
      lt63.place(x=132.5, y=435, width=102.5, height=20)
      lt63.bind("<FocusIn>", texto_temp_lt63)
      lt63.bind("<FocusOut>", texto_temp_lt63)

      Label(fr_cx_med3, text='Incremento:').place(x=5, y=470)
      lis63 = ['1', '5', '10']
      inc63 = StringVar()
      inc63.set('Selecionar...')
      OptionMenu(fr_cx_med3, inc63, *lis63).place(x=130, y=465)

      fr_cx_med4 = LabelFrame(aba6, borderwidth=1, relief='solid', text='  Unidade 4:  ')
      fr_cx_med4.place(x=890, y=70, width=290, height=520)

      def texto_temp_nome64(e):
         if texto64.get() == "Digitar nome...":
            texto64.delete(0, END)
         elif texto64.get() == '':
            texto64.insert(END, "Digitar nome...")

      Label(fr_cx_med4, text='Nome da caixa:').place(x=5, y=5)
      texto64 = Entry(fr_cx_med4, takefocus = 0)
      texto64.insert(END, "Digitar nome...")
      texto64.place(x=105, y=5, width=150, height=20)
      texto64.bind("<FocusIn>", texto_temp_nome64)
      texto64.bind("<FocusOut>", texto_temp_nome64)

      Label(fr_cx_med4, text='Tipo de saída de \ncabos do cliente:').place(x=5, y=40)
      opcoes637 = IntVar()
      Radiobutton(fr_cx_med4, text='Rosca', value=1, variable=opcoes637, command=ativa_rosca_cx_med4).place(x=120, y=50)
      Radiobutton(fr_cx_med4, text='Placa', value=2, variable=opcoes637, command=ativa_rosca_cx_med4).place(x=200, y=50)

      Label(fr_cx_med4, text='Previsão termostato:').place(x=5, y=80)
      lista638 = ['Nenhum', 'Esquerdo', 'Direito', 'Ambos']
      opcoes638 = StringVar()
      opcoes638.set('Selecionar...')
      OptionMenu(fr_cx_med4, opcoes638, *lista638).place(x=130, y=80)

      Label(fr_cx_med4, text='Tipo de rosca:').place(x=5, y=115)
      lista639 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
      opcoes639 = StringVar()
      opcoes639.set('Selecionar...')
      opt616 = OptionMenu(fr_cx_med4, opcoes639, *lista639, command=seta_qtd_rosca64)
      opt616.place(x=130, y=115)

      Label(fr_cx_med4, text='Quantidade de rosca:').place(x=5, y=150)
      lista640 = ['0', '1', '2']
      opcoes640 = StringVar()
      opcoes640.set('Selecionar...')
      opt612 = OptionMenu(fr_cx_med4, opcoes640, *lista640)
      opt612.place(x=130, y=150)

      Label(fr_cx_med4, text='Material da placa cega:').place(x=5, y=185)
      lista641 = ['Inox 304', 'Inox 316', 'Ferro fundido', 'Alumínio (AlSi6Cu4)', 'Alumínio (AlSi10Mg)']
      opcoes641 = StringVar()
      opcoes641.set('Selecionar...')
      opt620 = OptionMenu(fr_cx_med4, opcoes641, *lista641)
      opt620.place(x=130, y=185)

      Label(fr_cx_med4, text='Quantidade de trilhos:').place(x=5, y=220)
      lista642 = ['1', '2']
      opcoes642 = StringVar()
      opcoes642.set('Selecionar...')
      OptionMenu(fr_cx_med4, opcoes642, *lista642, command=ativa_ident_Cx_med4).place(x=130, y=220)

      Label(fr_cx_med4, text='Identificador trilho 1:').place(x=5, y=255)
      lista643 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes643 = StringVar()
      opcoes643.set('Selecionar...')
      opt67 = OptionMenu(fr_cx_med4, opcoes643, *lista643)
      opt67.place(x=130, y=255)

      Label(fr_cx_med4, text='Identificador trilho 2:').place(x=5, y=290)
      lista644 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes644 = StringVar()
      opcoes644.set('Selecionar...')
      opt68 = OptionMenu(fr_cx_med4, opcoes644, *lista644)
      opt68.place(x=130, y=290)

      Label(fr_cx_med4, text='Montada no mancal:').place(x=5, y=325)
      lista645 = ['Sim', 'Não']
      opcoes645 = StringVar()
      opcoes645.set('Selecionar...')
      OptionMenu(fr_cx_med4, opcoes645, *lista645).place(x=130, y=325)

      Label(fr_cx_med4, text='Tampa profundidade \naumentada:').place(x=5, y=360)
      lista646 = ['Sim', 'Não']
      opcoes646 = StringVar()
      opcoes646.set('Selecionar...')
      OptionMenu(fr_cx_med4, opcoes646, *lista646).place(x=130, y=360)

      Label(fr_cx_med4, text='Método de proteção:').place(x=5, y=400)
      lista647 = ['Área Segura', 'Ex-e', 'Ex-p']
      opcoes647 = StringVar()
      opcoes647.set('Selecionar...')
      OptionMenu(fr_cx_med4, opcoes647, *lista647).place(x=130, y=395)

      def texto_temp_lt64(e):
         if lt64.get() == "Digitar número...":
            lt64.delete(0, END)
         elif lt64.get() == '':
            lt64.insert(END, "Digitar número...")

      Label(fr_cx_med4, text='Posição inicial da\n lista técnica:').place(x=5, y=430)
      lt64 = Entry(fr_cx_med4, takefocus = 0)
      lt64.insert(END, "Digitar número...")
      lt64.place(x=132.5, y=435, width=102.5, height=20)
      lt64.bind("<FocusIn>", texto_temp_lt64)
      lt64.bind("<FocusOut>", texto_temp_lt64)

      Label(fr_cx_med4, text='Incremento:').place(x=5, y=470)
      lis64 = ['1', '5', '10']
      inc64 = StringVar()
      inc64.set('Selecionar...')
      OptionMenu(fr_cx_med4, inc64, *lis64).place(x=130, y=465)

      # Conteúdo da Aba 7 pt (Cx Acessórios Grande) ===============================================================================
      aba7 = Frame(nb)
      nb.add(aba7, text='Cx. Acess. Grande')

      def ativa_cx_gde(sel):
         global flag71, flag72, flag73, flag74, flag75, flag76, flag77, flag78
         if opcoes71.get() == '0':
            menu72.configure(state='disable')
            menu73.configure(state='disable')

            for child in fr_cx_gde1.winfo_children():
               child.configure(state='disable')
            for child in fr_cx_gde2.winfo_children():
               child.configure(state='disable')
            for child in fr_cx_gde3.winfo_children():
               child.configure(state='disable')
            for child in fr_cx_gde4.winfo_children():
               child.configure(state='disable')
            flag71 = False
            flag72 = False
            flag73 = False
            flag74 = False

         elif opcoes71.get() == '1':
            menu72.configure(state='normal')
            menu73.configure(state='normal')

            for child in fr_cx_gde1.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_gde2.winfo_children():
               child.configure(state='disable')
            for child in fr_cx_gde3.winfo_children():
               child.configure(state='disable')
            for child in fr_cx_gde4.winfo_children():
               child.configure(state='disable')
            flag71 = True
            flag72 = False
            flag73 = False
            flag74 = False
            
         elif opcoes71.get() == '2':
            menu72.configure(state='normal')
            menu73.configure(state='normal')

            for child in fr_cx_gde1.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_gde2.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_gde3.winfo_children():
               child.configure(state='disable')
            for child in fr_cx_gde4.winfo_children():
               child.configure(state='disable')
            flag71 = True
            flag72 = True
            flag73 = False
            flag74 = False
            
         elif opcoes71.get() == '3':
            menu72.configure(state='normal')
            menu73.configure(state='normal')

            for child in fr_cx_gde1.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_gde2.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_gde3.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_gde4.winfo_children():
               child.configure(state='disable')
            flag71 = True
            flag72 = True
            flag73 = True
            flag74 = False
         
         elif opcoes71.get() == '4':
            menu72.configure(state='normal')
            menu73.configure(state='normal')

            for child in fr_cx_gde1.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_gde2.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_gde3.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_gde4.winfo_children():
               child.configure(state='normal')
            flag71 = True
            flag72 = True
            flag73 = True
            flag74 = True


      def ativa_ident_Cx_gde1(sel):
         global flag75, tagX2_71

         if opcoes79.get() == '1':
            opt72.configure(state='disable')
            opcoes711.set('N/A')
            #tagX2_71 = 'N/A'
            flag75 = False

         elif opcoes79.get() == '2':
            opt72.configure(state='normal')
            opcoes711.set('Selecionar...')
            #tagX2_71 = opcoes711.get()
            flag75 = True
      
      def ativa_ident_Cx_gde2(sel):
         global flag76, tagX2_72

         if opcoes720.get() == '1':
            opt74.configure(state='disable')
            opcoes722.set('N/A')
            #tagX2_72 = 'N/A'
            flag76 = False

         elif opcoes720.get() == '2':
            opt74.configure(state='normal')
            opcoes722.set('Selecionar...')
            #tagX2_72 = opcoes722.get()
            flag76 = True

      def ativa_ident_Cx_gde3(sel):
         global flag77, tagX2_73

         if opcoes731.get() == '1':
            opt76.configure(state='disable')
            opcoes733.set('N/A')
            #tagX2_73 = 'N/A'
            flag77 = False

         elif opcoes731.get() == '2':
            opt76.configure(state='normal')
            opcoes733.set('Selecionar...')
            #tagX2_73 = opcoes733.get()
            flag77 = True

      def ativa_ident_Cx_gde4(sel):
         global flag78, tagX2_74

         if opcoes742.get() == '1':
            opt78.configure(state='disable')
            opcoes744.set('N/A')
            #tagX2_74 = 'N/A'
            flag78 = False

         elif opcoes742.get() == '2':
            opt78.configure(state='normal')
            opcoes744.set('Selecionar...')
            #tagX2_74 = opcoes744.get()
            flag78 = True

      def seta_qtd_rosca71(*r):
         if opcoes76.get() == 'Sem rosca':
            opcoes77.set('0')
            opt79.configure(state='disabled')

         else:
            opt79.configure(state='normal')
            opcoes77.set('Selecionar...')

      def seta_qtd_rosca72(*r):
         if opcoes717.get() == 'Sem rosca':
            opcoes718.set('0')
            opt710.configure(state='disabled')

         else:
            opt710.configure(state='normal')
            opcoes718.set('Selecionar...')

      def seta_qtd_rosca73(*r):
         if opcoes728.get() == 'Sem rosca':
            opcoes729.set('0')
            opt711.configure(state='disabled')

         else:
            opt711.configure(state='normal')
            opcoes729.set('Selecionar...')

      def seta_qtd_rosca74(*r):
         if opcoes739.get() == 'Sem rosca':
            opcoes740.set('0')
            opt712.configure(state='disabled')

         else:
            opt712.configure(state='normal')
            opcoes740.set('Selecionar...')


      def ativa_rosca_cx_gde1():
         global flag713
         lista76 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
         opt713 = OptionMenu(fr_cx_gde1, opcoes76, *lista76, command=seta_qtd_rosca71)

         if opcoes74.get() == 1:
            opt717.configure(state='disable') # Desabilita 'Material placa cega' se selecionado 'Rosca'
            opcoes78.set('N/A')

            opt713['menu'].delete(0,'end') # Remove todas as opções (reset 1)
            for opt in lista76: # Insere todas as opções (reset2)
                  opt713['menu'].add_command(label=opt, command=tkinter._setit(opcoes76, opt))
            opt713['menu'].delete(0,0) # Remove opção 'Sem rosca'
            flag713 = True

         else:
            opt717.configure(state='normal') # Habilita 'Material placa cega' se selecionado 'Placa'
            opcoes78.set('Selecionar...')

            opt713['menu'].delete(0,'end') # Remove todas as opções (reset1)
            for opt in lista76: # Insere todas as opções
                  opt713['menu'].add_command(label=opt, command=tkinter._setit(opcoes76, opt))
            flag713 = False

      def ativa_rosca_cx_gde2():
         global flag714
         lista717 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
         opt714 = OptionMenu(fr_cx_gde1, opcoes717, *lista717, command=seta_qtd_rosca72)

         if opcoes715.get() == 1:
            opt718.configure(state='disable') # Desabilita 'Material placa cega' se selecionado 'Rosca'
            opcoes719.set('N/A')

            opt714['menu'].delete(0,'end') # Remove todas as opções (reset 1)
            for opt in lista717: # Insere todas as opções (reset2)
                  opt714['menu'].add_command(label=opt, command=tkinter._setit(opcoes717, opt))
            opt714['menu'].delete(0,0) # Remove opção 'Sem rosca'
            flag714 = True

         else:
            opt718.configure(state='normal') # Habilita 'Material placa cega' se selecionado 'Placa'
            opcoes719.set('Selecionar...')

            opt714['menu'].delete(0,'end') # Remove todas as opções (reset1)
            for opt in lista717: # Insere todas as opções
                  opt714['menu'].add_command(label=opt, command=tkinter._setit(opcoes717, opt))
            flag714 = False

      def ativa_rosca_cx_gde3():
         global flag715
         lista728 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
         opt715 = OptionMenu(fr_cx_gde3, opcoes728, *lista728, command=seta_qtd_rosca73)

         if opcoes726.get() == 1:
            opt719.configure(state='disable') # Desabilita 'Material placa cega' se selecionado 'Rosca'
            opcoes730.set('N/A')

            opt715['menu'].delete(0,'end') # Remove todas as opções (reset 1)
            for opt in lista728: # Insere todas as opções (reset2)
                  opt715['menu'].add_command(label=opt, command=tkinter._setit(opcoes728, opt))
            opt715['menu'].delete(0,0) # Remove opção 'Sem rosca'
            flag715 = True

         else:
            opt719.configure(state='normal') # Habilita 'Material placa cega' se selecionado 'Placa'
            opcoes730.set('Selecionar...')

            opt715['menu'].delete(0,'end') # Remove todas as opções (reset1)
            for opt in lista728: # Insere todas as opções
                  opt715['menu'].add_command(label=opt, command=tkinter._setit(opcoes728, opt))
            flag715 = False

      def ativa_rosca_cx_gde4():
         global flag716
         lista739 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
         opt716 = OptionMenu(fr_cx_gde4, opcoes739, *lista739, command=seta_qtd_rosca74)

         if opcoes737.get() == 1:
            opt720.configure(state='disable') # Desabilita 'Material placa cega' se selecionado 'Rosca'
            opcoes741.set('N/A')

            opt716['menu'].delete(0,'end') # Remove todas as opções (reset 1)
            for opt in lista739: # Insere todas as opções (reset2)
                  opt716['menu'].add_command(label=opt, command=tkinter._setit(opcoes739, opt))
            opt716['menu'].delete(0,0) # Remove opção 'Sem rosca'
            flag716 = True

         else:
            opt720.configure(state='normal') # Habilita 'Material placa cega' se selecionado 'Placa'
            opcoes741.set('Selecionar...')

            opt716['menu'].delete(0,'end') # Remove todas as opções (reset1)
            for opt in lista739: # Insere todas as opções
                  opt716['menu'].add_command(label=opt, command=tkinter._setit(opcoes739, opt))
            flag716 = False
      

      fr_cx_gde = LabelFrame(aba7, borderwidth=1, relief='solid', text='  Caixa de acessórios grande:  ')
      fr_cx_gde.place(x=5, y=5, width=1175, height=60)

      Label(fr_cx_gde, text='Quantidade caixas de acessórios:').place(x=5, y=5)
      lista71 = ['0', '1', '2', '3', '4']
      opcoes71 = StringVar()
      opcoes71.set('Selecionar...')
      OptionMenu(fr_cx_gde, opcoes71, *lista71, command=ativa_cx_gde).place(x=200, y=2)

      Label(fr_cx_gde, text='Material caixas de acessórios:').place(x=370, y=5)
      lista72 = ['Inox 304', 'Inox 316', 'Ferro fundido']
      opcoes72 = StringVar()
      opcoes72.set('Selecionar...')
      menu72 = OptionMenu(fr_cx_gde, opcoes72, *lista72)
      menu72.place(x=545, y=2)

      Label(fr_cx_gde, text='Aterramento:').place(x=715, y=5)
      lista73 = ['Com aterramento', 'Sem aterramento']
      opcoes73 = StringVar()
      opcoes73.set('Selecionar...')
      menu73 = OptionMenu(fr_cx_gde, opcoes73, *lista73)
      menu73.place(x=805, y=2)

      fr_cx_gde1 = LabelFrame(aba7, borderwidth=1, relief='solid', text='  Unidade 1:  ')
      fr_cx_gde1.place(x=5, y=70, width=290, height=520)

      def texto_temp_nome71(e):
         if texto71.get() == "Digitar nome...":
            texto71.delete(0, END)
         elif texto71.get() == '':
            texto71.insert(END, "Digitar nome...")

      Label(fr_cx_gde1, text='Nome da caixa:').place(x=5, y=5)
      texto71 = Entry(fr_cx_gde1, takefocus = 0)
      texto71.insert(END, "Digitar nome...")
      texto71.place(x=105, y=5, width=150, height=20)
      texto71.bind("<FocusIn>", texto_temp_nome71)
      texto71.bind("<FocusOut>", texto_temp_nome71)

      Label(fr_cx_gde1, text='Tipo de saída de \ncabos do cliente:').place(x=5, y=40)
      opcoes74 = IntVar()
      Radiobutton(fr_cx_gde1, text='Rosca', value=1, variable=opcoes74, command=ativa_rosca_cx_gde1).place(x=120, y=50)
      Radiobutton(fr_cx_gde1, text='Placa', value=2, variable=opcoes74, command=ativa_rosca_cx_gde1).place(x=200, y=50)

      Label(fr_cx_gde1, text='Previsão termostato:').place(x=5, y=80)
      lista75 = ['Nenhum', 'Esquerdo', 'Direito', 'Ambos']
      opcoes75 = StringVar()
      opcoes75.set('Selecionar...')
      OptionMenu(fr_cx_gde1, opcoes75, *lista75).place(x=130, y=80)

      Label(fr_cx_gde1, text='Tipo de rosca:').place(x=5, y=115)
      lista76 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
      opcoes76 = StringVar()
      opcoes76.set('Selecionar...')
      opt713 = OptionMenu(fr_cx_gde1, opcoes76, *lista76, command=seta_qtd_rosca71)
      opt713.place(x=130, y=115)

      Label(fr_cx_gde1, text='Quantidade de rosca:').place(x=5, y=150)
      lista77 = ['0', '1', '2', '3']
      opcoes77 = StringVar()
      opcoes77.set('Selecionar...')
      opt79 = OptionMenu(fr_cx_gde1, opcoes77, *lista77)
      opt79.place(x=130, y=150)

      Label(fr_cx_gde1, text='Material da placa cega:').place(x=5, y=185)
      lista78 = ['Inox 304', 'Inox 316', 'Ferro fundido', 'Alumínio (AlSi6Cu4)', 'Alumínio (AlSi10Mg)']
      opcoes78 = StringVar()
      opcoes78.set('Selecionar...')
      opt717 = OptionMenu(fr_cx_gde1, opcoes78, *lista78)
      opt717.place(x=130, y=185)

      Label(fr_cx_gde1, text='Quantidade de trilhos:').place(x=5, y=220)
      lista79 = ['1', '2']
      opcoes79 = StringVar()
      opcoes79.set('Selecionar...')
      OptionMenu(fr_cx_gde1, opcoes79, *lista79, command=ativa_ident_Cx_gde1).place(x=130, y=220)

      Label(fr_cx_gde1, text='Identificador trilho 1:').place(x=5, y=255)
      lista710 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes710 = StringVar()
      opcoes710.set('Selecionar...')
      opt71 = OptionMenu(fr_cx_gde1, opcoes710, *lista710)
      opt71.place(x=130, y=255)

      Label(fr_cx_gde1, text='Identificador trilho 2:').place(x=5, y=290)
      lista711 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes711 = StringVar()
      opcoes711.set('Selecionar...')
      opt72 = OptionMenu(fr_cx_gde1, opcoes711, *lista711)
      opt72.place(x=130, y=290)

      Label(fr_cx_gde1, text='Montada no mancal:').place(x=5, y=325)
      lista712 = ['Sim', 'Não']
      opcoes712 = StringVar()
      opcoes712.set('Selecionar...')
      OptionMenu(fr_cx_gde1, opcoes712, *lista712).place(x=130, y=325)

      Label(fr_cx_gde1, text='Método de proteção:').place(x=5, y=400)
      lista714 = ['Área Segura', 'Ex-e', 'Ex-p']
      opcoes714 = StringVar()
      opcoes714.set('Selecionar...')
      OptionMenu(fr_cx_gde1, opcoes714, *lista714).place(x=130, y=395)

      def texto_temp_lt71(e):
         if lt71.get() == "Digitar número...":
            lt71.delete(0, END)
         elif lt71.get() == '':
            lt71.insert(END, "Digitar número...")

      Label(fr_cx_gde1, text='Posição inicial da\n lista técnica:').place(x=5, y=430)
      lt71 = Entry(fr_cx_gde1, takefocus = 0)
      lt71.insert(END, "Digitar número...")
      lt71.place(x=132.5, y=435, width=102.5, height=20)
      lt71.bind("<FocusIn>", texto_temp_lt71)
      lt71.bind("<FocusOut>", texto_temp_lt71)

      Label(fr_cx_gde1, text='Incremento:').place(x=5, y=470)
      lis71 = ['1', '5', '10']
      inc71 = StringVar()
      inc71.set('Selecionar...')
      OptionMenu(fr_cx_gde1, inc71, *lis71).place(x=130, y=465)

      fr_cx_gde2 = LabelFrame(aba7, borderwidth=1, relief='solid', text='  Unidade 2:  ')
      fr_cx_gde2.place(x=300, y=70, width=290, height=520)

      def texto_temp_nome72(e):
            if texto72.get() == "Digitar nome...":
               texto72.delete(0, END)
            elif texto72.get() == '':
               texto72.insert(END, "Digitar nome...")

      Label(fr_cx_gde2, text='Nome da caixa:').place(x=5, y=5)
      texto72 = Entry(fr_cx_gde2, takefocus = 0)
      texto72.insert(END, "Digitar nome...")
      texto72.place(x=105, y=5, width=150, height=20)
      texto72.bind("<FocusIn>", texto_temp_nome72)
      texto72.bind("<FocusOut>", texto_temp_nome72)

      Label(fr_cx_gde2, text='Tipo de saída de \ncabos do cliente:').place(x=5, y=40)
      opcoes715 = IntVar()
      Radiobutton(fr_cx_gde2, text='Rosca', value=1, variable=opcoes715, command=ativa_rosca_cx_gde2).place(x=120, y=50)
      Radiobutton(fr_cx_gde2, text='Placa', value=2, variable=opcoes715, command=ativa_rosca_cx_gde2).place(x=200, y=50)

      Label(fr_cx_gde2, text='Previsão termostato:').place(x=5, y=80)
      lista716 = ['Nenhum', 'Esquerdo', 'Direito', 'Ambos']
      opcoes716 = StringVar()
      opcoes716.set('Selecionar...')
      OptionMenu(fr_cx_gde2, opcoes716, *lista716).place(x=130, y=80)

      Label(fr_cx_gde2, text='Tipo de rosca:').place(x=5, y=115)
      lista717 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
      opcoes717 = StringVar()
      opcoes717.set('Selecionar...')
      opt714 = OptionMenu(fr_cx_gde2, opcoes717, *lista717, command=seta_qtd_rosca72)
      opt714.place(x=130, y=115)

      Label(fr_cx_gde2, text='Quantidade de rosca:').place(x=5, y=150)
      lista718 = ['0', '1', '2', '3']
      opcoes718 = StringVar()
      opcoes718.set('Selecionar...')
      opt710 = OptionMenu(fr_cx_gde2, opcoes718, *lista718)
      opt710.place(x=130, y=150)

      Label(fr_cx_gde2, text='Material da placa cega:').place(x=5, y=185)
      lista719 = ['Inox 304', 'Inox 316', 'Ferro fundido', 'Alumínio (AlSi6Cu4)', 'Alumínio (AlSi10Mg)']
      opcoes719 = StringVar()
      opcoes719.set('Selecionar...')
      opt718 = OptionMenu(fr_cx_gde2, opcoes719, *lista719)
      opt718.place(x=130, y=185)

      Label(fr_cx_gde2, text='Quantidade de trilhos:').place(x=5, y=220)
      lista720 = ['1', '2']
      opcoes720 = StringVar()
      opcoes720.set('Selecionar...')
      OptionMenu(fr_cx_gde2, opcoes720, *lista720, command=ativa_ident_Cx_gde2).place(x=130, y=220)

      Label(fr_cx_gde2, text='Identificador trilho 1:').place(x=5, y=255)
      lista721 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes721 = StringVar()
      opcoes721.set('Selecionar...')
      opt73 = OptionMenu(fr_cx_gde2, opcoes721, *lista721)
      opt73.place(x=130, y=255)

      Label(fr_cx_gde2, text='Identificador trilho 2:').place(x=5, y=290)
      lista722 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes722 = StringVar()
      opcoes722.set('Selecionar...')
      opt74 = OptionMenu(fr_cx_gde2, opcoes722, *lista722)
      opt74.place(x=130, y=290)

      Label(fr_cx_gde2, text='Montada no mancal:').place(x=5, y=325)
      lista723 = ['Sim', 'Não']
      opcoes723 = StringVar()
      opcoes723.set('Selecionar...')
      OptionMenu(fr_cx_gde2, opcoes723, *lista723).place(x=130, y=325)

      Label(fr_cx_gde2, text='Método de proteção:').place(x=5, y=400)
      lista725 = ['Área Segura', 'Ex-e', 'Ex-p']
      opcoes725 = StringVar()
      opcoes725.set('Selecionar...')
      OptionMenu(fr_cx_gde2, opcoes725, *lista725).place(x=130, y=395)

      def texto_temp_lt72(e):
         if lt72.get() == "Digitar número...":
            lt72.delete(0, END)
         elif lt72.get() == '':
            lt72.insert(END, "Digitar número...")

      Label(fr_cx_gde2, text='Posição inicial da\n lista técnica:').place(x=5, y=430)
      lt72 = Entry(fr_cx_gde2, takefocus = 0)
      lt72.insert(END, "Digitar número...")
      lt72.place(x=132.5, y=435, width=102.5, height=20)
      lt72.bind("<FocusIn>", texto_temp_lt72)
      lt72.bind("<FocusOut>", texto_temp_lt72)

      Label(fr_cx_gde2, text='Incremento:').place(x=5, y=470)
      lis72 = ['1', '5', '10']
      inc72 = StringVar()
      inc72.set('Selecionar...')
      OptionMenu(fr_cx_gde2, inc72, *lis72).place(x=130, y=465)

      fr_cx_gde3 = LabelFrame(aba7, borderwidth=1, relief='solid', text='  Unidade 3:  ')
      fr_cx_gde3.place(x=595, y=70, width=290, height=520)

      def texto_temp_nome73(e):
         if texto73.get() == "Digitar nome...":
            texto73.delete(0, END)
         elif texto73.get() == '':
            texto73.insert(END, "Digitar nome...")

      Label(fr_cx_gde3, text='Nome da caixa:').place(x=5, y=5)
      texto73 = Entry(fr_cx_gde3, takefocus = 0)
      texto73.insert(END, "Digitar nome...")
      texto73.place(x=105, y=5, width=150, height=20)
      texto73.bind("<FocusIn>", texto_temp_nome73)
      texto73.bind("<FocusOut>", texto_temp_nome73)

      Label(fr_cx_gde3, text='Tipo de saída de \ncabos do cliente:').place(x=5, y=40)
      opcoes726 = IntVar()
      Radiobutton(fr_cx_gde3, text='Rosca', value=1, variable=opcoes726, command=ativa_rosca_cx_gde3).place(x=120, y=50)
      Radiobutton(fr_cx_gde3, text='Placa', value=2, variable=opcoes726, command=ativa_rosca_cx_gde3).place(x=200, y=50)

      Label(fr_cx_gde3, text='Previsão termostato:').place(x=5, y=80)
      lista727 = ['Nenhum', 'Esquerdo', 'Direito', 'Ambos']
      opcoes727 = StringVar()
      opcoes727.set('Selecionar...')
      OptionMenu(fr_cx_gde3, opcoes727, *lista727).place(x=130, y=80)

      Label(fr_cx_gde3, text='Tipo de rosca:').place(x=5, y=115)
      lista728 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
      opcoes728 = StringVar()
      opcoes728.set('Selecionar...')
      opt715 = OptionMenu(fr_cx_gde3, opcoes728, *lista728, command=seta_qtd_rosca73)
      opt715.place(x=130, y=115)

      Label(fr_cx_gde3, text='Quantidade de rosca:').place(x=5, y=150)
      lista729 = ['0', '1', '2', '3']
      opcoes729 = StringVar()
      opcoes729.set('Selecionar...')
      opt711 = OptionMenu(fr_cx_gde3, opcoes729, *lista729)
      opt711.place(x=130, y=150)

      Label(fr_cx_gde3, text='Material da placa cega:').place(x=5, y=185)
      lista730 = ['Inox 304', 'Inox 316', 'Ferro fundido', 'Alumínio (AlSi6Cu4)', 'Alumínio (AlSi10Mg)']
      opcoes730 = StringVar()
      opcoes730.set('Selecionar...')
      opt719 = OptionMenu(fr_cx_gde3, opcoes730, *lista730)
      opt719.place(x=130, y=185)

      Label(fr_cx_gde3, text='Quantidade de trilhos:').place(x=5, y=220)
      lista731 = ['1', '2']
      opcoes731 = StringVar()
      opcoes731.set('Selecionar...')
      OptionMenu(fr_cx_gde3, opcoes731, *lista731, command=ativa_ident_Cx_gde3).place(x=130, y=220)

      Label(fr_cx_gde3, text='Identificador trilho 1:').place(x=5, y=255)
      lista732 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes732 = StringVar()
      opcoes732.set('Selecionar...')
      opt75 = OptionMenu(fr_cx_gde3, opcoes732, *lista732)
      opt75.place(x=130, y=255)

      Label(fr_cx_gde3, text='Identificador trilho 2:').place(x=5, y=290)
      lista733 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes733 = StringVar()
      opcoes733.set('Selecionar...')
      opt76 = OptionMenu(fr_cx_gde3, opcoes733, *lista733)
      opt76.place(x=130, y=290)

      Label(fr_cx_gde3, text='Montada no mancal:').place(x=5, y=325)
      lista734 = ['Sim', 'Não']
      opcoes734 = StringVar()
      opcoes734.set('Selecionar...')
      OptionMenu(fr_cx_gde3, opcoes734, *lista734).place(x=130, y=325)

      Label(fr_cx_gde3, text='Método de proteção:').place(x=5, y=400)
      lista736 = ['Área Segura', 'Ex-e', 'Ex-p']
      opcoes736 = StringVar()
      opcoes736.set('Selecionar...')
      OptionMenu(fr_cx_gde3, opcoes736, *lista736).place(x=130, y=395)

      def texto_temp_lt73(e):
         if lt73.get() == "Digitar número...":
            lt73.delete(0, END)
         elif lt73.get() == '':
            lt73.insert(END, "Digitar número...")

      Label(fr_cx_gde3, text='Posição inicial da\n lista técnica:').place(x=5, y=430)
      lt73 = Entry(fr_cx_gde3, takefocus = 0)
      lt73.insert(END, "Digitar número...")
      lt73.place(x=132.5, y=435, width=102.5, height=20)
      lt73.bind("<FocusIn>", texto_temp_lt73)
      lt73.bind("<FocusOut>", texto_temp_lt73)

      Label(fr_cx_gde3, text='Incremento:').place(x=5, y=470)
      lis73 = ['1', '5', '10']
      inc73 = StringVar()
      inc73.set('Selecionar...')
      OptionMenu(fr_cx_gde3, inc73, *lis73).place(x=130, y=465)

      fr_cx_gde4 = LabelFrame(aba7, borderwidth=1, relief='solid', text='  Unidade 4:  ')
      fr_cx_gde4.place(x=890, y=70, width=290, height=520)

      def texto_temp_nome74(e):
         if texto74.get() == "Digitar nome...":
            texto74.delete(0, END)
         elif texto74.get() == '':
            texto74.insert(END, "Digitar nome...")

      Label(fr_cx_gde4, text='Nome da caixa:').place(x=5, y=5)
      texto74 = Entry(fr_cx_gde4, takefocus = 0)
      texto74.insert(END, "Digitar nome...")
      texto74.place(x=105, y=5, width=150, height=20)
      texto74.bind("<FocusIn>", texto_temp_nome74)
      texto74.bind("<FocusOut>", texto_temp_nome74)

      Label(fr_cx_gde4, text='Tipo de saída de \ncabos do cliente:').place(x=5, y=40)
      opcoes737 = IntVar()
      Radiobutton(fr_cx_gde4, text='Rosca', value=1, variable=opcoes737, command=ativa_rosca_cx_gde4).place(x=120, y=50)
      Radiobutton(fr_cx_gde4, text='Placa', value=2, variable=opcoes737, command=ativa_rosca_cx_gde4).place(x=200, y=50)

      Label(fr_cx_gde4, text='Previsão termostato:').place(x=5, y=80)
      lista738 = ['Nenhum', 'Esquerdo', 'Direito', 'Ambos']
      opcoes738 = StringVar()
      opcoes738.set('Selecionar...')
      OptionMenu(fr_cx_gde4, opcoes738, *lista738).place(x=130, y=80)

      Label(fr_cx_gde4, text='Tipo de rosca:').place(x=5, y=115)
      lista739 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
      opcoes739 = StringVar()
      opcoes739.set('Selecionar...')
      opt716 = OptionMenu(fr_cx_gde4, opcoes739, *lista739, command=seta_qtd_rosca74)
      opt716.place(x=130, y=115)

      Label(fr_cx_gde4, text='Quantidade de rosca:').place(x=5, y=150)
      lista740 = ['0', '1', '2', '3']
      opcoes740 = StringVar()
      opcoes740.set('Selecionar...')
      opt712 = OptionMenu(fr_cx_gde4, opcoes740, *lista740)
      opt712.place(x=130, y=150)

      Label(fr_cx_gde4, text='Material da placa cega:').place(x=5, y=185)
      lista741 = ['Inox 304', 'Inox 316', 'Ferro fundido', 'Alumínio (AlSi6Cu4)', 'Alumínio (AlSi10Mg)']
      opcoes741 = StringVar()
      opcoes741.set('Selecionar...')
      opt720 = OptionMenu(fr_cx_gde4, opcoes741, *lista741)
      opt720.place(x=130, y=185)

      Label(fr_cx_gde4, text='Quantidade de trilhos:').place(x=5, y=220)
      lista742 = ['1', '2']
      opcoes742 = StringVar()
      opcoes742.set('Selecionar...')
      OptionMenu(fr_cx_gde4, opcoes742, *lista742, command=ativa_ident_Cx_gde4).place(x=130, y=220)

      Label(fr_cx_gde4, text='Identificador trilho 1:').place(x=5, y=255)
      lista743 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes743 = StringVar()
      opcoes743.set('Selecionar...')
      opt77 = OptionMenu(fr_cx_gde4, opcoes743, *lista743)
      opt77.place(x=130, y=255)

      Label(fr_cx_gde4, text='Identificador trilho 2:').place(x=5, y=290)
      lista744 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes744 = StringVar()
      opcoes744.set('Selecionar...')
      opt78 = OptionMenu(fr_cx_gde4, opcoes744, *lista744)
      opt78.place(x=130, y=290)

      Label(fr_cx_gde4, text='Montada no mancal:').place(x=5, y=325)
      lista745 = ['Sim', 'Não']
      opcoes745 = StringVar()
      opcoes745.set('Selecionar...')
      OptionMenu(fr_cx_gde4, opcoes745, *lista745).place(x=130, y=325)

      Label(fr_cx_gde4, text='Método de proteção:').place(x=5, y=400)
      lista747 = ['Área Segura', 'Ex-e', 'Ex-p']
      opcoes747 = StringVar()
      opcoes747.set('Selecionar...')
      OptionMenu(fr_cx_gde4, opcoes747, *lista747).place(x=130, y=395)

      def texto_temp_lt74(e):
         if lt74.get() == "Digitar número...":
            lt74.delete(0, END)
         elif lt74.get() == '':
            lt74.insert(END, "Digitar número...")

      Label(fr_cx_gde4, text='Posição inicial da\n lista técnica:').place(x=5, y=430)
      lt74 = Entry(fr_cx_gde4, takefocus = 0)
      lt74.insert(END, "Digitar número...")
      lt74.place(x=132.5, y=435, width=102.5, height=20)
      lt74.bind("<FocusIn>", texto_temp_lt74)
      lt74.bind("<FocusOut>", texto_temp_lt74)

      Label(fr_cx_gde4, text='Incremento:').place(x=5, y=470)
      lis74 = ['1', '5', '10']
      inc74 = StringVar()
      inc74.set('Selecionar...')
      OptionMenu(fr_cx_gde4, inc74, *lis74).place(x=130, y=465)

      # Conteúdo da Aba 8 pt ==================================================================================================
      aba8 = Frame(nb)
      nb.add(aba8, text='Cx. Acess. Embutida', state='disabled')

      def ativa_cx_emb(sel):
         global flag81, flag82
         if opcoes81.get() == '0':
            for child in fr_cx_emb1.winfo_children():
               child.configure(state='disable')
            for child in fr_cx_emb2.winfo_children():
               child.configure(state='disable')
            flag81 = False
            flag82 = False

         elif opcoes81.get() == '1':
            for child in fr_cx_emb1.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_emb2.winfo_children():
               child.configure(state='disable')
            flag81 = True
            flag82 = False
            
         elif opcoes81.get() == '2':
            for child in fr_cx_emb1.winfo_children():
               child.configure(state='normal')
            for child in fr_cx_emb2.winfo_children():
               child.configure(state='normal')
            flag81 = True
            flag82 = True

      def ativa_ident_Cx_emb1(sel):
         global flag83
         if opcoes87.get() == '1':
            opt81.configure(state='normal')
            opt82.configure(state='disable')
            flag83 = False

         elif opcoes87.get() == '2':
            opt81.configure(state='normal')
            opt82.configure(state='normal')
            flag83 = True
      
      def ativa_ident_Cx_emb2(sel):
         global flag84
         if opcoes814.get() == '1':
            opt83.configure(state='normal')
            opt84.configure(state='disable')
            flag84 = False

         elif opcoes814.get() == '2':
            opt83.configure(state='normal')
            opt84.configure(state='normal')
            flag84 = True

      fr_cx_emb = LabelFrame(aba8, borderwidth=1, relief='solid', text='  Caixa de acessórios embutida:  ')
      fr_cx_emb.place(x=5, y=10, width=850, height=60)

      Label(fr_cx_emb, text='Quantidade caixas de acessórios:').place(x=5, y=5)
      lista81 = ['0', '1', '2']
      opcoes81 = StringVar()
      opcoes81.set('Selecionar...')
      OptionMenu(fr_cx_emb, opcoes81, *lista81, command=ativa_cx_emb).place(x=200, y=2)

      Label(fr_cx_emb, text='Material caixas de acessórios:').place(x=320, y=5)
      lista82 = ['Inox 304', 'Inox 316', 'Ferro fundido']
      opcoes82 = StringVar()
      opcoes82.set('Selecionar...')
      OptionMenu(fr_cx_emb, opcoes82, *lista82).place(x=495, y=2)

      Label(fr_cx_emb, text='Aterramento:').place(x=615, y=5)
      lista83 = ['Com aterramento', 'Sem aterramento']
      opcoes83 = StringVar()
      opcoes83.set('Selecionar...')
      OptionMenu(fr_cx_emb, opcoes83, *lista83).place(x=705, y=2)

      fr_cx_emb1 = LabelFrame(aba8, borderwidth=1, relief='solid', text='  Unidade 1:  ')
      fr_cx_emb1.place(x=5, y=70, width=320, height=370)

      Label(fr_cx_emb1, text='Nome da caixa:').place(x=5, y=5)
      texto81 = Entry(fr_cx_emb1)
      texto81.place(x=105, y=5, width=150, height=20)

      Label(fr_cx_emb1, text='Tipo de rosca:').place(x=5, y=70)
      lista84 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
      opcoes84 = StringVar()
      opcoes84.set('Selecionar...')
      OptionMenu(fr_cx_emb1, opcoes84, *lista84).place(x=160, y=70)

      Label(fr_cx_emb1, text='Quantidade de rosca:').place(x=5, y=105)
      lista85 = ['0', '1', '2', '3']
      opcoes85 = StringVar()
      opcoes85.set('Selecionar...')
      OptionMenu(fr_cx_emb1, opcoes85, *lista85).place(x=160, y=105)

      Label(fr_cx_emb1, text='Material da placa cega:').place(x=5, y=140)
      lista86 = ['Inox 304', 'Inox 316', 'Ferro fundido', 'Alumínio (AlSi6Cu4)', 'Alumínio (AlSi10Mg)']
      opcoes86 = StringVar()
      opcoes86.set('Selecionar...')
      OptionMenu(fr_cx_emb1, opcoes86, *lista86).place(x=160, y=140)

      Label(fr_cx_emb1, text='Quantidade de trilhos:').place(x=5, y=175)
      lista87 = ['1', '2']
      opcoes87 = StringVar()
      opcoes87.set('Selecionar...')
      OptionMenu(fr_cx_emb1, opcoes87, *lista87, command=ativa_ident_Cx_emb1).place(x=160, y=175)

      Label(fr_cx_emb1, text='Identificador trilho 1:').place(x=5, y=210)
      lista88 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes88 = StringVar()
      opcoes88.set('Selecionar...')
      opt81 = OptionMenu(fr_cx_emb1, opcoes88, *lista88)
      opt81.place(x=160, y=210)

      Label(fr_cx_emb1, text='Identificador trilho 2:').place(x=5, y=245)
      lista89 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes89 = StringVar()
      opcoes89.set('Selecionar...')
      opt82 = OptionMenu(fr_cx_emb1, opcoes89, *lista89)
      opt82.place(x=160, y=245)

      Label(fr_cx_emb1, text='Método de proteção:').place(x=5, y=280)
      lista810 = ['Área Segura', 'Ex-p']
      opcoes810 = StringVar()
      opcoes810.set('Selecionar...')
      OptionMenu(fr_cx_emb1, opcoes810, *lista810).place(x=160, y=280)

      fr_cx_emb2 = LabelFrame(aba8, borderwidth=1, relief='solid', text='  Unidade 2:  ')
      fr_cx_emb2.place(x=330, y=70, width=320, height=370)

      Label(fr_cx_emb2, text='Nome da caixa:').place(x=5, y=5)
      texto82 = Entry(fr_cx_emb2, takefocus = 0)
      texto82.place(x=105, y=5, width=150, height=20)

      Label(fr_cx_emb2, text='Tipo de rosca:').place(x=5, y=70)
      lista811 = ['Sem rosca', 'G 1"', 'NPT 1"', 'M20x1,5']
      opcoes811 = StringVar()
      opcoes811.set('Selecionar...')
      OptionMenu(fr_cx_emb2, opcoes811, *lista811).place(x=160, y=70)

      Label(fr_cx_emb2, text='Quantidade de rosca:').place(x=5, y=105)
      lista812 = ['0', '1', '2', '3']
      opcoes812 = StringVar()
      opcoes812.set('Selecionar...')
      OptionMenu(fr_cx_emb2, opcoes812, *lista812).place(x=160, y=105)

      Label(fr_cx_emb2, text='Material da placa cega:').place(x=5, y=140)
      lista813 = ['Inox 304', 'Inox 316', 'Ferro fundido', 'Alumínio (AlSi6Cu4)', 'Alumínio (AlSi10Mg)']
      opcoes813 = StringVar()
      opcoes813.set('Selecionar...')
      OptionMenu(fr_cx_emb2, opcoes813, *lista813).place(x=160, y=140)

      Label(fr_cx_emb2, text='Quantidade de trilhos:').place(x=5, y=175)
      lista814 = ['1', '2']
      opcoes814 = StringVar()
      opcoes814.set('Selecionar...')
      OptionMenu(fr_cx_emb2, opcoes814, *lista814, command=ativa_ident_Cx_emb2).place(x=160, y=175)

      Label(fr_cx_emb2, text='Identificador trilho 1:').place(x=5, y=210)
      lista815 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes815 = StringVar()
      opcoes815.set('Selecionar...')
      opt83 = OptionMenu(fr_cx_emb2, opcoes815, *lista815)
      opt83.place(x=160, y=210)

      Label(fr_cx_emb2, text='Identificador trilho 2:').place(x=5, y=245)
      lista816 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes816 = StringVar()
      opcoes816.set('Selecionar...')
      opt84 = OptionMenu(fr_cx_emb2, opcoes816, *lista816)
      opt84.place(x=160, y=245)

      Label(fr_cx_emb2, text='Método de proteção:').place(x=5, y=280)
      lista817 = ['Área Segura', 'Ex-p']
      opcoes817 = StringVar()
      opcoes817.set('Selecionar...')
      OptionMenu(fr_cx_emb2, opcoes817, *lista817).place(x=160, y=280)

      # Conteúdo da Aba 9 pt ==================================================================================================
      aba9 = Frame(nb)
      nb.add(aba9, text='Refrigeração')

      def ativa_detec(r):
         global flag94
         if opcoes93.get() == 'Sem sensor':
            opt94.configure(state='disable')
            opcoes94.set('0')
            flag94 = False

         else:
            opt94.configure(state='normal')
            opcoes94.set('Selecionar...')
            flag94 = True

      def ativa_termometro_ar(r):
         global flag95
         if opcoes95.get() == 'Sem termômetro':
            opt96.configure(state='disable')
            opcoes96.set('0')
            flag95 = False

         else:
            opt96.configure(state='normal')
            opcoes96.set('Selecionar...')
            flag95 = True
      
      def ativa_termometro_agua(r):
         global flag96
         if opcoes97.get() == 'Sem termômetro':
            opt98.configure(state='disable')
            opcoes98.set('0')
            flag96 = False

         else:
            opt98.configure(state='normal')
            opcoes98.set('Selecionar...')
            flag96 = True

      fr_termosen = LabelFrame(aba9, borderwidth=1, relief='solid', text='  Termosensor:  ')
      fr_termosen.place(x=5, y=5, width=450, height=120)

      Label(fr_termosen, text='Qtde total sensores no ar:').place(x=5, y=20)
      lista91 = ['0', '1', '2', '3', '4']
      opcoes91 = StringVar()
      opcoes91.set('Selecionar...')
      OptionMenu(fr_termosen, opcoes91, *lista91).place(x=300, y=20)

      Label(fr_termosen, text='Qtde total sensores na água:').place(x=5, y=50)
      lista92 = ['0', '1', '2', '3', '4']
      opcoes92 = StringVar()
      opcoes92.set('Selecionar...')
      OptionMenu(fr_termosen, opcoes92, *lista92).place(x=300, y=50)

      fr_vazam = LabelFrame(aba9, borderwidth=1, relief='solid', text='  Detecção de vazamento:  ')
      fr_vazam.place(x=5, y=125, width=450, height=120)

      Label(fr_vazam, text='Modelo de sensor:').place(x=5, y=20)
      lista93 = ['Padrão', 'Especial', 'Sem sensor']
      opcoes93 = StringVar()
      opcoes93.set('Selecionar...')
      OptionMenu(fr_vazam, opcoes93, *lista93, command=ativa_detec).place(x=300, y=20)

      Label(fr_vazam, text='Qtde total de sensores:').place(x=5, y=50)
      lista94 = ['1', '2', '3', '4']
      opcoes94 = StringVar()
      opcoes94.set('Selecionar...')
      opt94 = OptionMenu(fr_vazam, opcoes94, *lista94)
      opt94.place(x=300, y=50)

      fr_termom = LabelFrame(aba9, borderwidth=1, relief='solid', text='  Termômetro:  ')
      fr_termom.place(x=5, y=245, width=450, height=190)

      Label(fr_termom, text='Contato elétrico do termômetro no ar:').place(x=5, y=20)
      lista95 = ['Sem contato', 'Com contato', 'Sem termômetro']
      opcoes95 = StringVar()
      opcoes95.set('Selecionar...')
      OptionMenu(fr_termom, opcoes95, *lista95, command=ativa_termometro_ar).place(x=300, y=20)

      Label(fr_termom, text='Qtde total de termômetros no ar:').place(x=5, y=50)
      lista96 = ['1', '2', '3', '4']
      opcoes96 = StringVar()
      opcoes96.set('Selecionar...')
      opt96 = OptionMenu(fr_termom, opcoes96, *lista96)
      opt96.place(x=300, y=50)

      Label(fr_termom, text='Contato elétrico do termômetro na água:').place(x=5, y=90)
      lista97 = ['Sem contato', 'Com contato', 'Sem termômetro']
      opcoes97 = StringVar()
      opcoes97.set('Selecionar...')
      OptionMenu(fr_termom, opcoes97, *lista97, command=ativa_termometro_agua).place(x=300, y=90)

      Label(fr_termom, text='Qtde total de termômetros na água:').place(x=5, y=120)
      lista98 = ['1', '2', '3', '4']
      opcoes98 = StringVar()
      opcoes98.set('Selecionar...')
      opt98 = OptionMenu(fr_termom, opcoes98, *lista98)
      opt98.place(x=300, y=120)

      fr_lt91 = LabelFrame(aba9, borderwidth=1, relief='solid', text='  Posições na lista técnica:  ')
      fr_lt91.place(x=5, y=435, width=300, height=120)

      def texto_temp_lt91(e):
         if lt91.get() == "Digitar número...":
            lt91.delete(0, END)
         elif lt91.get() == '':
            lt91.insert(END, "Digitar número...")

      Label(fr_lt91, text='Posição inicial na\n lista técnica:').place(x=5, y=10)
      lt91 = Entry(fr_lt91, takefocus = 0)
      lt91.insert(END, "Digitar número...")
      lt91.place(x=150, y=15, width=105, height=20)
      lt91.bind("<FocusIn>", texto_temp_lt91)
      lt91.bind("<FocusOut>", texto_temp_lt91)

      Label(fr_lt91, text='Incremento:').place(x=5, y=50)
      lis91 = ['1', '5', '10']
      inc91 = StringVar()
      inc91.set('Selecionar...')
      OptionMenu(fr_lt91, inc91, *lis91).place(x=150, y=50)

      # Conteúdo da Aba 10 pt ==================================================================================================
      aba10 = Frame(nb)
      nb.add(aba10, text='Comp. Gerais/Avulsos')

      fr_comp = LabelFrame(aba10, borderwidth=1, relief='solid', text='  Componentes Gerais:  ')
      fr_comp.place(x=5, y=5, width=450, height=120)

      Label(fr_comp, text='Tensão do freio:').place(x=5, y=20)
      lista101 = ['Sem freio', '24 Vcc', '125 Vcc', '110 Vca', '220 Vca']
      opcoes101 = StringVar()
      opcoes101.set('Selecionar...')
      OptionMenu(fr_comp, opcoes101, *lista101).place(x=300, y=20)

      Label(fr_comp, text='Identificador do trilho:').place(x=5, y=50)
      lista102 = ['X1', 'X2', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 
                  'X11', 'X12', 'X13', 'X14', 'X15', 'X16', 'X17', 'X18', 'X19', 'X20', 
                  'X21', 'X22', 'X23', 'X24', 'X25', 'X26', 'X27', 'X28']
      opcoes102 = StringVar()
      opcoes102.set('Selecionar...')
      OptionMenu(fr_comp, opcoes102, *lista102).place(x=300, y=50)

      fr_notas = LabelFrame(aba10, borderwidth=1, relief='solid', text='  Notas:  ')
      fr_notas.place(x=5, y=125, width=450, height=200)

      def texto_temp_notas(e):
         if textolongo101.get('1.0','end-1c') == "Digitar notas...":
            textolongo101.delete('1.0','end-1c')
         elif textolongo101.get('1.0','end-1c') == '':
            textolongo101.insert(END, "Digitar notas...")

      textolongo101 = Text(fr_notas)
      textolongo101.insert(END, "Digitar notas...")
      textolongo101.place(x=5, y=5, width=435, height=170)
      textolongo101.bind("<FocusIn>", texto_temp_notas)
      textolongo101.bind("<FocusOut>", texto_temp_notas)

      fr_lt101 = LabelFrame(aba10, borderwidth=1, relief='solid', text='  Posições na lista técnica:  ')
      fr_lt101.place(x=5, y=325, width=300, height=120)

      def texto_temp_lt101(e):
         if lt101.get() == "Digitar número...":
            lt101.delete(0, END)
         elif lt101.get() == '':
            lt101.insert(END, "Digitar número...")

      Label(fr_lt101, text='Posição inicial na\n lista técnica:').place(x=5, y=10)
      lt101 = Entry(fr_lt101, takefocus = 0)
      lt101.insert(END, "Digitar número...")
      lt101.place(x=150, y=15, width=105, height=20)
      lt101.bind("<FocusIn>", texto_temp_lt101)
      lt101.bind("<FocusOut>", texto_temp_lt101)

      Label(fr_lt101, text='Incremento:').place(x=5, y=50)
      lis101 = ['1', '5', '10']
      inc101 = StringVar()
      inc101.set('Selecionar...')
      OptionMenu(fr_lt101, inc101, *lis101).place(x=150, y=50)

      # Conteúdo da Aba 11 pt ==================================================================================================
      aba11 = Frame(nb)
      nb.add(aba11, text='Executar Programa')

      def escolhe_pasta():
         global pasta_escolhida
         texto_pasta.config(text='')
         pasta_escolhida = filedialog.askdirectory()
         Label(fr_salvar, text=str(pasta_escolhida), wraplength=250).place(x=20, y=105)

      fr_salvar = LabelFrame(aba11, borderwidth=1, relief='solid', text='  Gravação dos dados:  ')
      fr_salvar.place(x=5, y=5, width=450, height=200)

      Label(fr_salvar, text='Escolha a pasta na qual você deseja salvar a planilha resposta:').place(x=5, y=20)
      Label(fr_salvar, text='Pasta de destino escolhida:').place(x=10, y=80)
      texto_pasta = Label(fr_salvar, text='(Escolha uma pasta clicando no botão ao lado)')
      texto_pasta.place(x=20, y=105)

      btn_pasta = Button(fr_salvar, text='Escolher pasta...', command=escolhe_pasta)
      btn_pasta.place(x=300, y=100)

      # Cria botão de imprimir planilha
      btn_imprimir = Button(root, text='Imprimir Planilha', command=compila_dados)
      btn_imprimir.place(x=1080, y=710)

      # Imprime apresentação em português
      fr_infos_aba11 = LabelFrame(aba11, borderwidth=1, relief='solid')
      fr_infos_aba11.place(x=600, y=15, width=450, height=400)

      Label(fr_infos_aba11, text='Informações importantes:', font="-weight bold -size 13").place(x=5, y=15)
      Label(fr_infos_aba11, text='Esta é uma versão de testes, podendo ocorrer "bugs".', font= '20').place(x=5, y=60)
      Label(fr_infos_aba11, text='Caso ocorra erro, entre em contato conosco e reporte.', font= '20').place(x=5, y=85)
      Label(fr_infos_aba11, text='Contato: duartec@weg.net', font= '20').place(x=5, y=110)

      Label(fr_infos_aba11, text='Instruções básicas de uso:', font= '20').place(x=5, y=160)
      Label(fr_infos_aba11, text='1) Escolha o idioma desejado no quadro acima;', font= '20').place(x=5, y=190)
      Label(fr_infos_aba11, text='2) Percorra as abas preenchendo todos os dados;', font= '20').place(x=5, y=220)
      Label(fr_infos_aba11, text='3) Escolha uma pasta na aba "Executar Programa";', font= '20').place(x=5, y=250)
      Label(fr_infos_aba11, text='4) Clique no botão "Imprimir Planilha";', font= '20').place(x=5, y=280)
      Label(fr_infos_aba11, text='3) Acesse a pasta de salvamento escolhida;', font= '20').place(x=5, y=310)
      Label(fr_infos_aba11, text='3) Abra a planilha "planilha_resposta_mat_12345678.xlsx".', font= '20').place(x=5, y=340)

#=================================================================================================================================
#============================================= Divisão entre os idiomas ==========================================================
#=================================================================================================================================

   elif var.get() == 2:
      
      btn_imprimir.destroy()

      info_texto.destroy()
      info_texto = Label(fr_info, text='All fields must be filled before printing data.', font= '20')
      info_texto.place(x=10, y=20, width=800)

      nb = ttk.Notebook(root)
      nb.place(x=0, y=81, width=largura, height=altura-130)

      # Conteúdo da Aba 1 en =====================================================================================================
      aba1 = Frame(nb)
      nb.add(aba1, text='Message')

      fr_aviso = LabelFrame(aba1, borderwidth=1, relief='solid')
      fr_aviso.place(x=5, y=5, width=500, height=150)
      
      Label(fr_aviso, text='This option is under development and will be released soon!', font= '20').place(x=30, y=60)


var = IntVar()
rb_pt = Radiobutton(fr_idiomas, text='Português', value=1, variable=var, command=escolha_idioma).place(x=0, y=10)
rb_en = Radiobutton(fr_idiomas, text='English', value=2, variable=var, command=escolha_idioma).place(x=0, y=30)

root.mainloop()

